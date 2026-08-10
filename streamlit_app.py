import streamlit as st
import streamlit.components.v1 as components
import io, os, datetime, math, json, numpy as np, pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# =====================================================================
# KONFIGURACIJA (secrets) + SUPABASE
# =====================================================================
try:
    from supabase import create_client
except Exception:
    create_client = None

def _cfg(key, default=None):
    try:
        if key in st.secrets:
            return st.secrets[key]
    except Exception:
        pass
    return default

APP_PASSWORD   = _cfg("APP_PASSWORD", "vape2024")     # analiticar (pun pristup)
ADMIN_PASSWORD = _cfg("ADMIN_PASSWORD", "aman2024")   # administracija (koleginice)
DIREKTOR_PASSWORD = _cfg("DIREKTOR_PASSWORD", "2026vape")  # direktori (pregled izveštaja)
SUPABASE_URL   = _cfg("SUPABASE_URL", "")
SUPABASE_KEY   = _cfg("SUPABASE_KEY", "")

MESEC_NAZIVI = {1:'Januar',2:'Februar',3:'Mart',4:'April',5:'Maj',6:'Jun',
                7:'Jul',8:'Avgust',9:'Septembar',10:'Oktobar',11:'Novembar',12:'Decembar'}

def mesec_label(key):
    try:
        y, m = str(key).split('-')
        return f"{MESEC_NAZIVI.get(int(m), m)} {y}"
    except Exception:
        return str(key)

@st.cache_resource
def _sb():
    if not (create_client and SUPABASE_URL and SUPABASE_KEY):
        return None
    try:
        return create_client(SUPABASE_URL, SUPABASE_KEY)
    except Exception:
        return None

def sb_dostupan():
    return _sb() is not None

def sb_objavi(mesec_key, sistem, podaci, xlsx_b64=None):
    cli = _sb()
    if cli is None:
        raise RuntimeError("Supabase nije podešen (SUPABASE_URL / SUPABASE_KEY u secrets).")
    payload = {"mesec": mesec_key, "sistem": sistem.strip(), "podaci": podaci}
    if xlsx_b64:
        payload["analitika_xlsx"] = xlsx_b64
    try:
        cli.table("porudzbine").upsert(payload, on_conflict="mesec,sistem").execute()
    except Exception:
        # kolona 'analitika_xlsx' možda ne postoji -> objavi bez nje (ne ruši objavu)
        payload.pop("analitika_xlsx", None)
        cli.table("porudzbine").upsert(payload, on_conflict="mesec,sistem").execute()
    # osvezi kes da koleginice odmah vide
    for fn in (sb_meseci, sb_sisteme, sb_svi_sistemi, sb_ucitaj, sb_pregled, sb_ucitaj_xlsx):
        try: fn.clear()
        except Exception: pass


def sb_obrisi(mesec_key, sistem):
    """Obriši objavljeni izveštaj (mesec + sistem) iz baze. Trajno."""
    cli = _sb()
    if cli is None:
        raise RuntimeError("Supabase nije podešen.")
    cli.table("porudzbine").delete().eq("mesec", mesec_key).eq("sistem", sistem).execute()
    for fn in (sb_meseci, sb_sisteme, sb_svi_sistemi, sb_ucitaj, sb_pregled, sb_ucitaj_xlsx):
        try: fn.clear()
        except Exception: pass


@st.cache_data(ttl=60)
def sb_ucitaj_xlsx(mesec_key, sistem):
    """Vrati base64 analitika Excel-a za sistem/mesec (ili None). Odvojen upit da ne
    opterećuje običan sb_ucitaj."""
    cli = _sb()
    if cli is None:
        return None
    try:
        res = cli.table("porudzbine").select("analitika_xlsx").eq("mesec", mesec_key).eq("sistem", sistem).limit(1).execute()
        if not res.data:
            return None
        return res.data[0].get("analitika_xlsx")
    except Exception:
        return None


def sb_objavi_izvestaj_prodaje(html, xlsx_b64, mesec_label, prodaja_json=""):
    """Sačuvaj (samo poslednji) direktorski Izveštaj prodaje u tabelu izvestaj_prodaje."""
    cli = _sb()
    if cli is None:
        raise RuntimeError("Supabase nije podešen.")
    payload = {"kljuc": "latest", "html": html, "xlsx_b64": xlsx_b64,
               "mesec_label": mesec_label or "", "prodaja_json": prodaja_json or "",
               "generisano": datetime.datetime.now().strftime("%d.%m.%Y %H:%M")}
    try:
        cli.table("izvestaj_prodaje").upsert(payload, on_conflict="kljuc").execute()
    except Exception:
        # kolona prodaja_json možda ne postoji -> sačuvaj bez nje
        payload.pop("prodaja_json", None)
        cli.table("izvestaj_prodaje").upsert(payload, on_conflict="kljuc").execute()
    try:
        sb_ucitaj_izvestaj_prodaje.clear()
    except Exception:
        pass


@st.cache_data(ttl=60)
def sb_ucitaj_izvestaj_prodaje():
    cli = _sb()
    if cli is None:
        return None
    try:
        res = cli.table("izvestaj_prodaje").select("html,xlsx_b64,mesec_label,generisano,prodaja_json").eq("kljuc", "latest").limit(1).execute()
        if res.data:
            return res.data[0]
    except Exception:
        pass
    try:
        res = cli.table("izvestaj_prodaje").select("html,xlsx_b64,mesec_label,generisano").eq("kljuc", "latest").limit(1).execute()
        if not res.data:
            return None
        return res.data[0]
    except Exception:
        return None


@st.cache_data(ttl=30)
def sb_meseci():
    cli = _sb()
    if cli is None: return []
    res = cli.table("porudzbine").select("mesec").execute()
    keys = sorted({r["mesec"] for r in (res.data or [])}, reverse=True)
    return [{"key": k, "label": mesec_label(k)} for k in keys]

@st.cache_data(ttl=30)
def sb_sisteme(mesec_key):
    cli = _sb()
    if cli is None: return []
    res = cli.table("porudzbine").select("sistem").eq("mesec", mesec_key).execute()
    return sorted({r["sistem"] for r in (res.data or [])})

@st.cache_data(ttl=30)
def sb_svi_sistemi():
    cli = _sb()
    if cli is None: return []
    res = cli.table("porudzbine").select("sistem").execute()
    return sorted({r["sistem"] for r in (res.data or [])})

@st.cache_data(ttl=30)
def sb_ucitaj(mesec_key, sistem):
    cli = _sb()
    if cli is None: return None
    res = cli.table("porudzbine").select("podaci").eq("mesec", mesec_key).eq("sistem", sistem).limit(1).execute()
    if not res.data: return None
    return res.data[0]["podaci"]

def sb_predaj(mesec_key, sistem):
    """Označi izveštaj (mesec+sistem) kao PREDAT — zaključava izmene za taj mesec.
    Upisuje meta.predato u JSON reda u tabeli porudzbine."""
    cli = _sb()
    if cli is None:
        raise RuntimeError("Supabase nije podešen.")
    res = cli.table("porudzbine").select("podaci").eq("mesec", mesec_key).eq("sistem", sistem.strip()).limit(1).execute()
    if not res.data:
        raise RuntimeError("Nema objavljenog izveštaja za taj mesec/sistem.")
    podaci = res.data[0].get("podaci") or {}
    meta = podaci.get("meta") or {}
    meta["predato"] = True
    meta["predato_at"] = datetime.datetime.now().strftime("%d.%m.%Y %H:%M")
    podaci["meta"] = meta
    cli.table("porudzbine").update({"podaci": podaci}).eq("mesec", mesec_key).eq("sistem", sistem.strip()).execute()
    try:
        sb_ucitaj.clear()
    except Exception:
        pass


@st.cache_data(ttl=30)
def sb_pregled():
    """Lagani pregled svega objavljenog: mesec, sistem, kada (bez povlacenja stavki)."""
    cli = _sb()
    if cli is None: return []
    res = cli.table("porudzbine").select("mesec,sistem,objavljeno").order("mesec", desc=True).execute()
    return res.data or []

REAKCIJE_OPCIJE = ["Pozvala sam", "Poslala sam mejl", "Obavestila direktorku"]

def _reak_short(r):
    return {"Pozvala sam": "\U0001F4DE Pozvala", "Poslala sam mejl": "\u2709\uFE0F Mejl",
            "Obavestila direktorku": "\U0001F454 Direktorka",
            "Ubačena porudžbina": "\U0001F4E6 Ubačena porudžbina"}.get(r, r)

def _zona_disp(nivo):
    return {"crveno": ("z-red", "\U0001F534 Hitno pozvati", "\U0001F534", "Hitno pozvati"),
            "zuto":   ("z-org", "\U0001F7E0 Iskontrolisati", "\U0001F7E0", "Iskontrolisati"),
            "zeleno": ("z-grn", "\U0001F7E2 Dobra", "\U0001F7E2", "Dobra")}[nivo]

def sb_load_obrada(mesec_key, sistem):
    cli = _sb()
    if cli is None:
        return {}
    try:
        res = cli.table("obrada").select("idk,reakcije,trebovali,trebovali_tip,njihova,napomena").eq("mesec", mesec_key).eq("sistem", sistem).execute()
    except Exception:
        return {}
    out = {}
    for r in (res.data or []):
        out[int(r["idk"])] = {"reakcije": r.get("reakcije") or [], "trebovali_tip": r.get("trebovali_tip") or "", "njihova": r.get("njihova") or {}, "napomena": r.get("napomena") or ""}
    return out

def sb_save_obrada(mesec_key, sistem, idk, reakcije, trebovali_tip, njihova=None, napomena=""):
    cli = _sb()
    if cli is None:
        raise RuntimeError("Supabase nije podešen.")
    cli.table("obrada").upsert({"mesec": mesec_key, "sistem": sistem, "idk": int(idk),
        "reakcije": list(reakcije), "trebovali": bool(trebovali_tip), "trebovali_tip": trebovali_tip or "",
        "njihova": dict(njihova or {}), "napomena": napomena or "",
        "azurirano": datetime.datetime.now().isoformat()}, on_conflict="mesec,sistem,idk").execute()

def sb_bulk_ubaci(mesec_key, sistem, ids):
    cli = _sb()
    if cli is None:
        raise RuntimeError("Supabase nije podešen.")
    _now = datetime.datetime.now().isoformat()
    rows = [{"mesec": mesec_key, "sistem": sistem, "idk": int(i),
             "reakcije": ["Ubačena porudžbina"], "trebovali": True, "trebovali_tip": "nas",
             "njihova": {}, "napomena": "", "azurirano": _now} for i in ids]
    if rows:
        cli.table("obrada").upsert(rows, on_conflict="mesec,sistem,idk").execute()

def sb_bulk_reset(mesec_key, sistem):
    cli = _sb()
    if cli is None:
        raise RuntimeError("Supabase nije podešen.")
    cli.table("obrada").delete().eq("mesec", mesec_key).eq("sistem", sistem).execute()

def sb_load_plan(mesec_key):
    cli = _sb()
    if cli is None:
        return None
    try:
        res = cli.table("plan_objave").select("datum").eq("mesec", mesec_key).limit(1).execute()
    except Exception:
        return None
    if not res.data:
        return None
    return res.data[0].get("datum")

def sb_save_plan(mesec_key, datum):
    cli = _sb()
    if cli is None:
        raise RuntimeError("Supabase nije podešen.")
    cli.table("plan_objave").upsert({"mesec": mesec_key, "datum": datum,
        "azurirano": datetime.datetime.now().isoformat()}, on_conflict="mesec").execute()

def sb_plan_meseci():
    cli = _sb()
    if cli is None:
        return []
    try:
        res = cli.table("plan_objave").select("mesec").execute()
    except Exception:
        return []
    return [r["mesec"] for r in (res.data or [])]

# ---- Rokovi (postavlja direktor; 3 posebna: administracija / sistemi / prodaja) ----
@st.cache_data(ttl=30)
def sb_rokovi_all():
    cli = _sb()
    if cli is None:
        return {}
    try:
        res = cli.table("rokovi").select("mesec,rok_admin,rok_sistemi,rok_prodaja,rok_syx,rok_potraz,napomena").execute()
        return {r["mesec"]: r for r in (res.data or [])}
    except Exception:
        # kolone rok_syx/rok_potraz možda još ne postoje -> učitaj bez njih
        try:
            res = cli.table("rokovi").select("mesec,rok_admin,rok_sistemi,rok_prodaja,napomena").execute()
            return {r["mesec"]: r for r in (res.data or [])}
        except Exception:
            return {}

def sb_rokovi_get(mesec_key):
    try:
        return sb_rokovi_all().get(mesec_key, {}) or {}
    except Exception:
        return {}

def sb_rokovi_set(mesec_key, rok_admin, rok_sistemi, rok_prodaja, napomena, rok_syx=None, rok_potraz=None):
    cli = _sb()
    if cli is None:
        raise RuntimeError("Supabase nije podešen.")
    payload = {"mesec": mesec_key,
               "rok_admin": rok_admin or None, "rok_sistemi": rok_sistemi or None,
               "rok_prodaja": rok_prodaja or None, "rok_syx": rok_syx or None,
               "rok_potraz": rok_potraz or None, "napomena": napomena or "",
               "azurirano": datetime.datetime.now().isoformat()}
    try:
        cli.table("rokovi").upsert(payload, on_conflict="mesec").execute()
    except Exception:
        # fallback ako rok_syx/rok_potraz kolone ne postoje
        payload.pop("rok_syx", None)
        payload.pop("rok_potraz", None)
        cli.table("rokovi").upsert(payload, on_conflict="mesec").execute()
    try:
        sb_rokovi_all.clear()
    except Exception:
        pass

def _rok_fmt(s):
    """'YYYY-MM-DD' -> 'DD.MM.YYYY' (ili prazno)."""
    if not s:
        return ""
    try:
        return datetime.date.fromisoformat(str(s)[:10]).strftime("%d.%m.%Y")
    except Exception:
        return str(s)

def _rok_je_prosao(s):
    """True ako je rok (YYYY-MM-DD) prošao (danas je posle roka)."""
    if not s:
        return False
    try:
        return datetime.date.today() > datetime.date.fromisoformat(str(s)[:10])
    except Exception:
        return False

# ---- Izveštaj SYX (Word dokument po mesecu; analitičar ubacuje, direktori preuzimaju) ----
@st.cache_data(ttl=30)
def sb_syx_list():
    cli = _sb()
    if cli is None:
        return []
    try:
        res = cli.table("izvestaj_syx").select("mesec,filename,azurirano").execute()
        return sorted(res.data or [], key=lambda r: r.get("mesec", ""), reverse=True)
    except Exception:
        return []

@st.cache_data(ttl=300)
def sb_syx_get(mesec_key):
    cli = _sb()
    if cli is None:
        return None
    try:
        res = cli.table("izvestaj_syx").select("filename,docx_b64").eq("mesec", mesec_key).limit(1).execute()
        if res.data:
            return res.data[0]
    except Exception:
        pass
    return None

def sb_syx_set(mesec_key, filename, b64):
    cli = _sb()
    if cli is None:
        raise RuntimeError("Supabase nije podešen.")
    cli.table("izvestaj_syx").upsert({"mesec": mesec_key, "filename": filename, "docx_b64": b64,
        "azurirano": datetime.datetime.now().isoformat()}, on_conflict="mesec").execute()
    for fn in (sb_syx_list, sb_syx_get):
        try:
            fn.clear()
        except Exception:
            pass

def sb_syx_obrisi(mesec_key):
    cli = _sb()
    if cli is None:
        raise RuntimeError("Supabase nije podešen.")
    cli.table("izvestaj_syx").delete().eq("mesec", mesec_key).execute()
    for fn in (sb_syx_list, sb_syx_get):
        try:
            fn.clear()
        except Exception:
            pass

# ---- Izveštaj potraživanja (analitičar uploaduje Excel, administracija dopunjava u aplikaciji, direktor vidi + izvozi) ----
def _potraz_txt(v):
    if v is None:
        return ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%d.%m.%Y")
    return str(v)

def _potraz_tip(naziv):
    n = " ".join(str(naziv).lower().split())
    if "status" in n and "komunik" in n:
        return "dd_a"
    if "status" in n and ("tužb" in n or "tuzb" in n):
        return "dd_b"
    if any(k in n for k in ["vrednost", "ukupni dug", "za uplatu", "dana", "iznos"]):
        return "num"
    return "text"

def _potraz_num(v):
    if isinstance(v, (int, float)):
        return float(v)
    try:
        s = str(v).strip().replace(" ", "")
        return float(s) if s else None
    except Exception:
        return None

def potraz_parse(xlsx_bytes):
    """Rasčlani Excel potraživanja u strukturu (listovi -> sekcije -> kolone/redovi) sa
    koordinatama ćelija, da bi administracija mogla da dopunjava u aplikaciji, a izvoz
    upisuje nazad u originalni fajl (identično formatiranje)."""
    import openpyxl as _ox
    wb = _ox.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    dd_a, dd_b = [], []
    if "_LISTE" in wb.sheetnames:
        for r in wb["_LISTE"].iter_rows(values_only=True):
            if r and len(r) >= 1 and r[0]:
                dd_a.append(_potraz_txt(r[0]))
            if r and len(r) >= 2 and r[1]:
                dd_b.append(_potraz_txt(r[1]))
    out = {"stanje_na_dan": "", "dd_a": dd_a, "dd_b": dd_b, "listovi": []}
    stanje = ""
    for ws in wb.worksheets:
        if ws.title == "_LISTE":
            continue
        rows = list(ws.iter_rows(values_only=False))
        n = len(rows)
        for r in rows[:6]:
            for c in r:
                if c.value and "Stanje na dan" in str(c.value):
                    stanje = str(c.value).replace("Stanje na dan:", "").strip()
        sekcije = []
        i = 0
        last = ""
        while i < n:
            rc = rows[i]
            vals = [_potraz_txt(c.value).strip() for c in rc]
            for m in ("PO FAKTURI", "PO ODJAVI"):
                if any(v == m for v in vals):
                    last = m
            is_hdr = any(v == "Komitent" for v in vals) or any("Naziv komitenta" in v for v in vals)
            if is_hdr:
                kolone = []
                for c in rc:
                    nz = _potraz_txt(c.value).strip()
                    if nz:
                        kolone.append({"col": c.column, "naziv": " ".join(nz.split()), "tip": _potraz_tip(nz)})
                redovi = []
                ukupno_row = None
                sum_cols = []
                j = i + 1
                while j < n:
                    r2 = rows[j]
                    a = _potraz_txt(r2[0].value).strip()
                    if a.upper().startswith("UKUPNO"):
                        ukupno_row = r2[0].row
                        for c in r2:
                            if isinstance(c.value, (int, float)):
                                sum_cols.append(c.column)
                        break
                    if all(_potraz_txt(c.value).strip() == "" for c in r2):
                        break
                    cells = {}
                    for k in kolone:
                        cells[str(k["col"])] = _potraz_txt(r2[k["col"] - 1].value)
                    redovi.append({"r": r2[0].row, "cells": cells})
                    j += 1
                sekcije.append({"naslov": last or ws.title, "kolone": kolone, "redovi": redovi,
                                "ukupno_row": ukupno_row, "sum_cols": sum_cols})
                i = j
                last = ""
                continue
            i += 1
        out["listovi"].append({"sheet": ws.title, "sekcije": sekcije})
    out["stanje_na_dan"] = stanje
    return out

def potraz_init_popuna(struktura):
    """Početna popuna = vrednosti iz uploadovanog fajla (administracija ih dalje menja)."""
    pop = {}
    for L in struktura.get("listovi", []):
        sh = L["sheet"]
        pop[sh] = {}
        for s in L["sekcije"]:
            for rr in s["redovi"]:
                pop[sh].setdefault(str(rr["r"]), {})
                for col, val in rr["cells"].items():
                    pop[sh][str(rr["r"])][str(col)] = val
    return pop

def potraz_export(original_b64, struktura, popuna):
    """Upiši trenutne vrednosti (popuna) u originalni Excel i vrati bytes (identično formatiranje).
    Tip kolone se određuje PO SEKCIJI (ista kolona može biti različita u „po fakturi" i „po odjavi")."""
    import openpyxl as _ox, base64 as _b64
    wb = _ox.load_workbook(io.BytesIO(_b64.b64decode(original_b64)))
    popuna = popuna or {}
    for L in struktura.get("listovi", []):
        sh = L["sheet"]
        if sh not in wb.sheetnames:
            continue
        ws = wb[sh]
        for s in L["sekcije"]:
            coltip = {str(k["col"]): k["tip"] for k in s["kolone"]}
            for rr in s["redovi"]:
                rv = popuna.get(sh, {}).get(str(rr["r"]), {})
                for col, tip in coltip.items():
                    if col not in rv:
                        continue
                    val = rv.get(col)
                    c = ws.cell(row=int(rr["r"]), column=int(col))
                    if tip == "num":
                        nv = _potraz_num(val)
                        c.value = nv if nv is not None else None
                    else:
                        c.value = val if (val is not None and str(val) != "") else None
            if s.get("ukupno_row") and s.get("sum_cols"):
                for scol in s["sum_cols"]:
                    tot = 0.0
                    for rr in s["redovi"]:
                        nv = _potraz_num(popuna.get(sh, {}).get(str(rr["r"]), {}).get(str(scol)))
                        if nv is not None:
                            tot += nv
                    ws.cell(row=int(s["ukupno_row"]), column=int(scol)).value = tot
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

def potraz_section_df(sekcija, sheet, popuna):
    """Napravi DataFrame za jednu sekciju + jedinstvene labele + redosled kolona (col indeksi)."""
    labels = []
    counts = {}
    col_order = []
    for k in sekcija["kolone"]:
        base = k["naziv"]
        c = counts.get(base, 0) + 1
        counts[base] = c
        labels.append(base if c == 1 else base + " (" + str(c) + ")")
        col_order.append(str(k["col"]))
    data = {lab: [] for lab in labels}
    for rr in sekcija["redovi"]:
        rv = popuna.get(sheet, {}).get(str(rr["r"]), {})
        for idx, k in enumerate(sekcija["kolone"]):
            col = str(k["col"])
            val = rv.get(col, rr["cells"].get(col, ""))
            if k["tip"] == "num":
                data[labels[idx]].append(_potraz_num(val))
            else:
                data[labels[idx]].append("" if val is None else str(val))
    return pd.DataFrame(data, columns=labels), labels, col_order

def potraz_col_config(sekcija, labels, df, dd_a, dd_b):
    cfg = {}
    for idx, k in enumerate(sekcija["kolone"]):
        lab = labels[idx]
        tip = k["tip"]
        if tip in ("dd_a", "dd_b"):
            base = list(dd_a if tip == "dd_a" else dd_b)
            try:
                existing = [str(x) for x in df[lab].dropna().unique() if str(x).strip() != ""]
            except Exception:
                existing = []
            opts = [""] + base + [e for e in existing if e not in base]
            cfg[lab] = st.column_config.SelectboxColumn(lab, options=opts, required=False, width="medium")
        elif tip == "num":
            cfg[lab] = st.column_config.NumberColumn(lab, format="%.2f")
        else:
            cfg[lab] = st.column_config.TextColumn(lab)
    return cfg

@st.cache_data(ttl=30)
def sb_potraz_list():
    cli = _sb()
    if cli is None:
        return []
    try:
        res = cli.table("izvestaj_potrazivanja").select("mesec,naziv,predato,azurirano").execute()
        return sorted(res.data or [], key=lambda r: r.get("mesec", ""), reverse=True)
    except Exception:
        return []

@st.cache_data(ttl=120)
def sb_potraz_get(mesec_key):
    cli = _sb()
    if cli is None:
        return None
    try:
        res = cli.table("izvestaj_potrazivanja").select("mesec,naziv,original_b64,struktura,popuna,predato,predato_at").eq("mesec", mesec_key).limit(1).execute()
        if res.data:
            return res.data[0]
    except Exception:
        pass
    return None

def sb_potraz_set(mesec_key, naziv, original_b64, struktura_json, popuna_json):
    cli = _sb()
    if cli is None:
        raise RuntimeError("Supabase nije podešen.")
    cli.table("izvestaj_potrazivanja").upsert({"mesec": mesec_key, "naziv": naziv,
        "original_b64": original_b64, "struktura": struktura_json, "popuna": popuna_json,
        "predato": False, "predato_at": None,
        "azurirano": datetime.datetime.now().isoformat()}, on_conflict="mesec").execute()
    for fn in (sb_potraz_list, sb_potraz_get):
        try:
            fn.clear()
        except Exception:
            pass

def sb_potraz_popuni(mesec_key, popuna_json, predato=False):
    cli = _sb()
    if cli is None:
        raise RuntimeError("Supabase nije podešen.")
    _upd = {"popuna": popuna_json, "azurirano": datetime.datetime.now().isoformat()}
    if predato:
        _upd["predato"] = True
        _upd["predato_at"] = datetime.datetime.now().strftime("%d.%m.%Y %H:%M")
    cli.table("izvestaj_potrazivanja").update(_upd).eq("mesec", mesec_key).execute()
    for fn in (sb_potraz_list, sb_potraz_get):
        try:
            fn.clear()
        except Exception:
            pass

def sb_potraz_obrisi(mesec_key):
    cli = _sb()
    if cli is None:
        raise RuntimeError("Supabase nije podešen.")
    cli.table("izvestaj_potrazivanja").delete().eq("mesec", mesec_key).execute()
    for fn in (sb_potraz_list, sb_potraz_get):
        try:
            fn.clear()
        except Exception:
            pass

def _potraz_collect(edited):
    """Iz izmenjenih data_editor tabela sklopi popunu {sheet: {r: {col: val}}}."""
    newpop = {}
    for (sh, sidx), (ed, col_order, sek) in edited.items():
        newpop.setdefault(sh, {})
        for ri, rr in enumerate(sek["redovi"]):
            newpop[sh].setdefault(str(rr["r"]), {})
            for ci, col in enumerate(col_order):
                try:
                    val = ed.iloc[ri, ci]
                except Exception:
                    val = ""
                if val is None or (isinstance(val, float) and pd.isna(val)):
                    val = ""
                newpop[sh][str(rr["r"])][col] = val
    return newpop

def potraz_admin_ui():
    st.markdown("<div style='font-size:18px;font-weight:800;margin:4px 0 10px;'>💳 Izveštaj potraživanja</div>", unsafe_allow_html=True)
    _lst = sb_potraz_list()
    if not _lst:
        st.info("Analitičar još nije objavio nijedan izveštaj potraživanja.")
        return
    _labels = [mesec_label(r["mesec"]) for r in _lst]
    _keys = [r["mesec"] for r in _lst]
    _sel = st.selectbox("Mesec", _labels, index=0, key="pz_adm_mes")
    _mk = _keys[_labels.index(_sel)]
    rec = sb_potraz_get(_mk)
    if not rec:
        st.info("Nema podataka za ovaj mesec.")
        return
    try:
        struct = json.loads(rec.get("struktura") or "{}")
        pop = json.loads(rec.get("popuna") or "{}")
    except Exception:
        st.error("Greška u podacima izveštaja.")
        return
    _predato = bool(rec.get("predato"))
    _rok_pz = sb_rokovi_get(_mk).get("rok_potraz")
    if _rok_pz and not _predato:
        if _rok_je_prosao(_rok_pz):
            st.warning("⏰ Rok za predaju potraživanja (" + _rok_fmt(_rok_pz) + ") je istekao.")
        else:
            st.info("⏰ Rok za predaju potraživanja: " + _rok_fmt(_rok_pz) + ".")
    if _predato:
        st.success("Ovaj izveštaj je predat direktoru (" + str(rec.get("predato_at") or "") + "). Prikaz je samo za pregled.")
    else:
        st.caption("Stanje na dan: " + str(struct.get("stanje_na_dan", "")) + ". Dopuni iznose, statuse i komentare, pa klikni Prosledi direktoru.")
    dd_a = struct.get("dd_a", [])
    dd_b = struct.get("dd_b", [])
    edited = {}
    for L in struct.get("listovi", []):
        sh = L["sheet"]
        st.markdown("<div style='margin:16px 0 2px;font-size:15px;font-weight:800;color:#7c3aed;'>" + _h_escape(sh) + "</div>", unsafe_allow_html=True)
        for sidx, s in enumerate(L["sekcije"]):
            st.markdown("<div style='font-size:12.5px;font-weight:700;color:#6b7280;margin:8px 0 2px;'>" + _h_escape(str(s["naslov"])) + "</div>", unsafe_allow_html=True)
            df, labels, col_order = potraz_section_df(s, sh, pop)
            cfg = potraz_col_config(s, labels, df, dd_a, dd_b)
            _ed = st.data_editor(df, column_config=cfg, hide_index=True, use_container_width=True,
                                 num_rows="fixed", key="pz_ed_" + _mk + "_" + sh + "_" + str(sidx),
                                 disabled=_predato)
            edited[(sh, sidx)] = (_ed, col_order, s)
    if not _predato:
        _b1, _b2 = st.columns(2)
        with _b1:
            if st.button("💾 Sačuvaj (bez prosleđivanja)", key="pz_adm_save", use_container_width=True):
                try:
                    sb_potraz_popuni(_mk, json.dumps(_potraz_collect(edited), default=str))
                    st.success("Sačuvano.")
                    st.rerun()
                except Exception as _e:
                    st.error("Greška pri čuvanju: " + str(_e))
        with _b2:
            if st.button("📨 Prosledi direktoru", key="pz_adm_send", use_container_width=True, type="primary"):
                try:
                    sb_potraz_popuni(_mk, json.dumps(_potraz_collect(edited), default=str), predato=True)
                    st.success("Prosleđeno direktoru.")
                    st.rerun()
                except Exception as _e:
                    st.error("Greška pri prosleđivanju: " + str(_e))

def potraz_director_ui():
    st.markdown('<div style="font-size:20px;font-weight:800;margin:6px 0 6px;">💳 Izveštaj potraživanja</div>', unsafe_allow_html=True)
    _lst = [r for r in sb_potraz_list() if r.get("predato")]
    if not _lst:
        st.info("Još nema prosleđenih izveštaja potraživanja. Administracija ih popunjava i prosleđuje.")
        return
    _labels = [mesec_label(r["mesec"]) for r in _lst]
    _keys = [r["mesec"] for r in _lst]
    _sel = st.selectbox("Mesec", _labels, index=0, key="pz_dir_mes")
    _mk = _keys[_labels.index(_sel)]
    rec = sb_potraz_get(_mk)
    if not rec:
        st.info("Nema podataka.")
        return
    try:
        struct = json.loads(rec.get("struktura") or "{}")
        pop = json.loads(rec.get("popuna") or "{}")
    except Exception:
        st.error("Greška u podacima.")
        return
    _tc1, _tc2 = st.columns([3, 1])
    with _tc1:
        st.caption("Stanje na dan: " + str(struct.get("stanje_na_dan", ""))
                   + "  ·  predato " + str(rec.get("predato_at") or ""))
    with _tc2:
        try:
            _xb = potraz_export(rec.get("original_b64"), struct, pop)
            st.download_button("⬇️ Izvezi u Excel", _xb,
                file_name="Izvestaj_potrazivanja_" + _mk + ".xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="pz_dir_dl", use_container_width=True)
        except Exception as _e:
            st.caption("Izvoz trenutno nije moguć.")
    for L in struct.get("listovi", []):
        sh = L["sheet"]
        st.markdown("<div style='margin:16px 0 2px;font-size:15px;font-weight:800;color:#7c3aed;'>" + _h_escape(sh) + "</div>", unsafe_allow_html=True)
        for s in L["sekcije"]:
            st.markdown("<div style='font-size:12.5px;font-weight:700;color:#6b7280;margin:8px 0 2px;'>" + _h_escape(str(s["naslov"])) + "</div>", unsafe_allow_html=True)
            df, labels, col_order = potraz_section_df(s, sh, pop)
            st.dataframe(df, hide_index=True, use_container_width=True)

# ---- HITNOST po objektu (na osnovu niskog lagera) ----
# Pragovi su namerno apsolutni i lako se menjaju (dole dve brojke).
HIT_CRVENO_KOM   = 15   # >= ovoliko kom/mesec na artiklima bez lagera -> HITNO
HIT_CRVENO_ART   = 8    # ILI >= ovoliko artikala na nuli -> HITNO
HIT_ZUTO_KOM     = 5
HIT_ZUTO_ART     = 3
def hitnost_objekta(stavke_obj):
    """stavke_obj = lista stavki za jedan objekat (mogu biti i artikli sa porudzbinom 0).
    Urgentnost racunamo SAMO na artiklima koje predlazemo za porudzbinu (kol > 0),
    da mrtvi artikli (bez prodaje, bez porudzbine) ne bi lazno dizali hitnost.
    izgubljeno = predvidjena mesecna prodaja na artiklima koji su TRENUTNO na 0 lagera."""
    _ord = [s for s in stavke_obj if int(s.get('kol', 0)) > 0]
    n_nula = sum(1 for s in _ord if int(s.get('lager', 0)) == 0)
    izgubljeno = sum(int(s.get('pred', 0)) for s in _ord if int(s.get('lager', 0)) == 0)
    if izgubljeno >= HIT_CRVENO_KOM or n_nula >= HIT_CRVENO_ART:
        nivo = 'crveno'
    elif izgubljeno >= HIT_ZUTO_KOM or n_nula >= HIT_ZUTO_ART:
        nivo = 'zuto'
    else:
        nivo = 'zeleno'
    return nivo, n_nula, izgubljeno

HIT_EMOJI = {'crveno': '🔴', 'zuto': '🟡', 'zeleno': '🟢'}
HIT_RANG  = {'crveno': 0, 'zuto': 1, 'zeleno': 2}
HIT_TEKST = {'crveno': 'Hitno', 'zuto': 'Srednje', 'zeleno': 'Može da čeka'}

def stavke_iz_rezultata(result, engine):
    """Napravi listu stavki za cuvanje u Supabase.
    Cuvamo SVE artikle (i one sa porudzbinom 0) za objekte koji imaju bar
    jedan artikal za porudzbinu — da koleginice u detalju vide ceo asortiman
    objekta, a ne samo ono sto se poruci."""
    reg = engine.region_map if hasattr(engine, 'region_map') else {}
    _ordered_ids = set(result[result['Porudzbina_2'] > 0]['ID KOMITENTA'].tolist())
    order = result[result['ID KOMITENTA'].isin(_ordered_ids)]
    stavke = []
    for _, r in order.iterrows():
        idk = int(r['ID KOMITENTA'])
        stavke.append({
            'idk': idk,
            'region': str(reg.get(r['ID KOMITENTA'], '') or ''),
            'ida': int(r['id artikla']),
            'naziv': str(r['Naziv artikla']),
            'grupa': str(r.get('Grupa', '') or ''),
            'pred': int(r.get('Predikcija', 0)),
            'lager': int(r.get('Lager_danas', 0)),
            'kol': int(r['Porudzbina_2']),
        })
    return stavke


def direktor_blok(engine, res):
    """Iz rezultata analitike spakuj podatke za DIREKTORSKI izveštaj (prodaja, trend,
    poređenja, po grupama, OOS). Bezbedno — nikad ne ruši objavu (sve u try/except)."""
    out = {}
    try:
        ml = list(engine.mesec_labels)
    except Exception:
        ml = []

    def _col(lb, suf):
        c = str(lb) + suf
        return c if c in res.columns else None

    # Trend prodaje po mesecu (kom + rsd)
    trend = []
    for lb in ml:
        cp = _col(lb, "_Prodaja")
        cr = _col(lb, "_Promet")
        if cp:
            try:
                trend.append({"mesec": lb, "kom": int(res[cp].sum()),
                              "rsd": int(res[cr].sum()) if cr else 0})
            except Exception:
                pass
    out["prodaja_trend"] = trend
    _last = ml[-1] if ml else None
    out["prodaja_tekuci"] = trend[-1] if trend else {"mesec": _last, "kom": 0, "rsd": 0}

    # Poređenja: prošli mesec, 6-mesečni prosek, isti mesec lani (ako ima podataka)
    comp = {}
    if len(trend) >= 2:
        comp["prosli_mesec"] = trend[-2]
        _prev6 = trend[-7:-1] if len(trend) >= 7 else trend[:-1]
        if _prev6:
            comp["prosek_6m"] = {"kom": int(round(sum(t["kom"] for t in _prev6) / len(_prev6))), "n": len(_prev6)}
    if _last and " " in str(_last):
        _mnaz, _god = str(_last).rsplit(" ", 1)
        if _god.isdigit():
            _lani = _mnaz + " " + str(int(_god) - 1)
            for t in trend:
                if t["mesec"] == _lani:
                    comp["isti_mesec_lani"] = t
                    break
    out["poredjenja"] = comp

    # Prodaja po grupama (tekući mesec)
    grupe = []
    _lc = _col(_last, "_Prodaja") if _last else None
    if _lc and ("Grupa" in res.columns):
        try:
            g = res.groupby("Grupa")[_lc].sum().sort_values(ascending=False)
            for naz, kom in g.items():
                grupe.append({"grupa": str(naz), "kom": int(kom)})
        except Exception:
            pass
    out["po_grupama"] = grupe

    # Out of stock (na osnovu lagera danas)
    try:
        _oos = res[(res["Lager_danas"] == 0) & (res["Predikcija"] > 0)]
        out["oos"] = {"kombinacija_na_0": int(len(_oos)), "izgubljeno_kom": int(_oos["Predikcija"].sum())}
        _pa = (_oos.groupby("Naziv artikla")
               .agg(objekata=("ID KOMITENTA", "nunique"), izgubljeno=("Predikcija", "sum"))
               .sort_values("izgubljeno", ascending=False).head(10))
        out["oos_po_artiklu"] = [{"artikal": str(i), "objekata": int(r["objekata"]),
                                  "izgubljeno": int(r["izgubljeno"])} for i, r in _pa.iterrows()]
    except Exception:
        pass

    # ---- Prosečna prodaja po objektu (mesečno) ----
    try:
        _ppo = []
        for lb in ml:
            cp = _col(lb, "_Prodaja")
            if cp:
                _tot = float(res[cp].sum())
                _no = int((res[cp] > 0).sum())
                _ppo.append({"mesec": lb, "prosek": round(_tot / _no, 1) if _no else 0.0})
        out["prosek_po_objektu"] = _ppo
    except Exception:
        pass

    # ---- Mesečne grupe (za složeni grafikon) iz analitike (fallback ako nema tabele prodaje) ----
    try:
        if "Grupa" in res.columns:
            _gm = {}
            for lb in ml:
                cp = _col(lb, "_Prodaja")
                if cp:
                    _gs = res.groupby("Grupa")[cp].sum()
                    for _gn, _gv in _gs.items():
                        _gm.setdefault(str(_gn), []).append(int(_gv))
            if _gm:
                out["nazivi"] = list(ml)
                out["grupe_mesecno"] = _gm
    except Exception:
        pass

    # ---- OOS po količinama za poslednji mesec ----
    try:
        _dfoos = getattr(engine, "df_oos", None)
        if _dfoos is not None and len(_dfoos) > 0 and ml:
            _last = ml[-1]
            _colo = "OOS_" + str(_last)
            _ok = {"mesec": _last, "izgubljeno_kom": 0, "objekata_na_0": 0, "po_artiklu": []}
            if _colo in _dfoos.columns:
                _sub = _dfoos[_dfoos[_colo] > 0]
                _ok["izgubljeno_kom"] = int(round(_sub[_colo].sum()))
                _per = (_sub.groupby("Naziv artikla")
                        .agg(objekata=("ID KOMITENTA", "nunique"), izg=(_colo, "sum"))
                        .reset_index().sort_values("izg", ascending=False))
                _ok["po_artiklu"] = [{"artikal": str(r["Naziv artikla"]), "objekata": int(r["objekata"]),
                                      "izgubljeno": int(round(r["izg"]))} for _, r in _per.iterrows()]
            try:
                _ok["objekata_na_0"] = int(res[res["Lager_danas"] == 0]["ID KOMITENTA"].nunique())
            except Exception:
                _ok["objekata_na_0"] = int((_dfoos.get("Lager_danas", 0) == 0).sum())
            out["oos_kom"] = _ok
    except Exception:
        pass

    # ---- Predlog porudžbine za sistem + pokrivenost lagera ----
    try:
        if "Porudzbina_2" in res.columns:
            _exc = getattr(engine, "excluded", None) or set()
            _rv = res[~res["ID KOMITENTA"].isin(_exc)] if _exc else res
            _pr = {"ukupno": int(_rv["Porudzbina_2"].sum()),
                   "objekata": int(_rv[_rv["Porudzbina_2"] > 0]["ID KOMITENTA"].nunique()),
                   "po_grupi": []}
            if "Grupa" in _rv.columns:
                _gp = _rv.groupby("Grupa")["Porudzbina_2"].sum().sort_values(ascending=False)
                _pr["po_grupi"] = [{"grupa": str(g), "kom": int(v)} for g, v in _gp.items() if int(v) > 0]
            # prosečna pokrivenost (dani) — ponderisano prodajom, iz df_promo
            _dp = getattr(engine, "df_promo", None)
            if _dp is not None and len(_dp) > 0 and "Dani_pokrivanja" in _dp.columns:
                _dd = _dp[(_dp["Dani_pokrivanja"] < 900) & (_dp["Prodato_kom"] > 0)]
                if len(_dd) > 0:
                    _wsum = float((_dd["Dani_pokrivanja"] * _dd["Prodato_kom"]).sum())
                    _psum = float(_dd["Prodato_kom"].sum())
                    _pr["dani_avg"] = int(round(_wsum / _psum)) if _psum else 0
            out["porudzbina"] = _pr
    except Exception:
        pass

    # ---- Bestseleri i najslabiji artikli (iz res, po ukupnoj prodaji perioda) ----
    try:
        _art = {}
        _grp_of = {}
        for lb in ml:
            cp = _col(lb, "_Prodaja")
            if cp and "Naziv artikla" in res.columns:
                _gsum = res.groupby("Naziv artikla")[cp].sum()
                for _nz, _vv in _gsum.items():
                    _art[str(_nz)] = _art.get(str(_nz), 0) + int(_vv)
        if "Grupa" in res.columns and "Naziv artikla" in res.columns:
            for _nz, _gg in res.groupby("Naziv artikla")["Grupa"].first().items():
                _grp_of[str(_nz)] = str(_gg)
        if _art:
            _srt = sorted(_art.items(), key=lambda kv: kv[1], reverse=True)
            _best = [{"artikal": k, "grupa": _grp_of.get(k, ""), "prodato": v} for k, v in _srt[:8]]
            _slab = [{"artikal": k, "grupa": _grp_of.get(k, ""), "prodato": v}
                     for k, v in sorted(_srt, key=lambda kv: kv[1])[:8]]
            out["artikli_rang"] = {"best": _best, "slab": _slab}
    except Exception:
        pass

    # ---- Uspešnost akcije (iz df_promo; samo ako ima cena) ----
    try:
        _dp = getattr(engine, "df_promo", None)
        if _dp is not None and len(_dp) > 0:
            _ak = {"ukupno_akcija": int(_dp["Profit_akcija"].sum()),
                   "ukupno_redovna": int(_dp["Profit_da_je_redovna"].sum())}
            _ak["razlika"] = _ak["ukupno_redovna"] - _ak["ukupno_akcija"]
            _tp = _dp.sort_values("Prodato_kom", ascending=False).head(12)
            _ak["artikli"] = [{"naziv": str(r["Naziv"]), "grupa": str(r["Grupa"]),
                               "prodato": int(r["Prodato_kom"]), "obrt": float(r["Obrt_x"]),
                               "popust": float(r["Popust_%"]), "profit_akcija": int(r["Profit_akcija"]),
                               "cena_akcije": int(r["Cena_akcije"]), "dani": int(r["Dani_pokrivanja"]) if r["Dani_pokrivanja"] < 900 else 0}
                              for _, r in _tp.iterrows()]
            out["akcija"] = _ak
    except Exception:
        pass

    # ---- Profitabilnost (identično kao u analitici; puni se samo ako ima cena) ----
    try:
        if getattr(engine, "has_prices", False) and len(getattr(engine, "df_profit_obj", [])) > 0:
            prof = engine.df_profit_obj.copy()
            a_labels = list(engine.analitika_labels) if getattr(engine, "analitika_labels", None) else list(ml)
            n_mes = max(len(a_labels), 1)
            pf = {}
            pf["period"] = ", ".join(a_labels) if a_labels else "svi meseci"
            pf["n_mes"] = n_mes
            pf["n_obj"] = int(getattr(engine, "num_komitenti", len(prof)))
            pf["total_trosak"] = int(prof["Trosak_mkt"].sum())
            pf["total_bruto"] = int(prof["Bruto_profit"].sum())
            pf["total_neto"] = int(prof["Neto_profit"].sum())
            _dfoos = getattr(engine, "df_oos", None)
            _has_oos = _dfoos is not None and len(_dfoos) > 0
            pf["total_oos"] = int(_dfoos["Izgubljeni_profit"].sum()) if _has_oos else 0
            # mesečni trend bruto / neto
            _bm = []; _nm = []
            for lb in a_labels:
                cb = "Bruto_" + str(lb); cn = "Neto_" + str(lb)
                _bm.append([lb, int(prof[cb].sum()) if cb in prof.columns else 0])
                _nm.append([lb, int(prof[cn].sum()) if cn in prof.columns else 0])
            pf["bruto_po_mes"] = _bm
            pf["neto_po_mes"] = _nm
            # profitabilnost po objektima
            ukupno = len(prof)
            _neg = prof[prof["Neto_profit"] <= 0]
            _oos_neg = prof[(prof["Neto_profit"] <= 0) & (prof["Potencijalni_profit"] > 0)]
            _pravi_neg = prof[(prof["Neto_profit"] <= 0) & (prof["Potencijalni_profit"] <= 0)]
            pf["obj_ukupno"] = int(ukupno)
            pf["obj_profit"] = int(ukupno - len(_neg))
            pf["obj_oos_neg"] = int(len(_oos_neg))
            pf["obj_pravi_neg"] = int(len(_pravi_neg))
            _tpo = float(getattr(engine, "trosak_po_objektu", 0) or 0)
            _usteda = len(_pravi_neg) * _tpo + (abs(_pravi_neg["Neto_profit"].sum()) if len(_pravi_neg) > 0 else 0)
            pf["usteda_ukupno"] = int(_usteda)
            pf["objekti"] = [{"id": int(r["ID KOMITENTA"]), "neto": int(r["Neto_profit"]),
                              "potencijal": int(r["Potencijalni_profit"]), "bruto": int(r["Bruto_profit"]),
                              "trosak": int(r["Trosak_mkt"]), "oos": int(r["Izgubljeno_OOS"])}
                             for _, r in prof.iterrows()]
            # OOS u dinarima (identično kao analitika)
            if _has_oos:
                pf["oos_0_danas"] = int((_dfoos.get("Lager_danas", 0) == 0).sum()) if "Lager_danas" in _dfoos.columns else 0
                _om = []
                for lb in a_labels:
                    ci = "Izgub_" + str(lb); co = "OOS_" + str(lb)
                    _om.append([lb, int(_dfoos[ci].sum()) if ci in _dfoos.columns else 0,
                                int((_dfoos[co] > 0).sum()) if co in _dfoos.columns else 0])
                pf["oos_po_mes"] = _om
                _oa = (_dfoos.groupby(["id artikla", "Naziv artikla"])
                       .agg(Objekata=("ID KOMITENTA", "nunique"), OOS_meseci=("OOS_meseci", "sum"),
                            Izgubljeni_profit=("Izgubljeni_profit", "sum"))
                       .reset_index().sort_values("Izgubljeni_profit", ascending=False))
                pf["oos_artikli"] = [{"naziv": str(r["Naziv artikla"]), "objekata": int(r["Objekata"]),
                                      "meseci": int(r["OOS_meseci"]), "rsd": int(r["Izgubljeni_profit"])}
                                     for _, r in _oa.iterrows()]
            out["profit"] = pf
    except Exception:
        pass

    return out

# =====================================================================
# PRIJAVA (dve uloge: analitika / administracija)
# =====================================================================
def check_password():
    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False
    if st.session_state.authenticated:
        return True
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Poppins:wght@300;400;500;600;700&display=swap');
    html, body, .stApp { background: #f4f1fb !important; font-family: 'Poppins', sans-serif; }
    .stApp { background: linear-gradient(135deg, #f7f5fc 0%, #efeaf8 55%, #f7f5fc 100%) !important; }
    header[data-testid="stHeader"] { background: transparent !important; }
    .stDeployButton { display: none; }
    footer { display: none; }
    #MainMenu { display: none; }
    .block-container { max-width: 440px !important; margin: 0 auto !important; padding-top: 90px !important; }
    .login-card-wrap { background:#ffffff; border:1px solid #ece7f6; border-radius:20px; padding:34px 30px;
        box-shadow:0 10px 40px rgba(124,58,237,0.10); }
    .stTextInput > div > div > input {
        background: #ffffff !important; border: 1px solid #e3e0ee !important;
        color: #1f2430 !important; border-radius: 12px !important; padding: 13px 16px !important; font-size: 15px !important; }
    .stTextInput > div > div > input::placeholder { color: #9aa0ad !important; }
    .stTextInput > div > div > input:focus {
        border-color: #a855f7 !important; box-shadow: 0 0 0 3px rgba(168,85,247,0.14) !important; }
    .stButton > button {
        background: linear-gradient(135deg, #a855f7 0%, #ec4899 100%) !important; color: white !important;
        border: none !important; border-radius: 12px !important; padding: 13px 32px !important;
        font-weight: 700 !important; font-size: 15px !important; width: 100% !important;
        box-shadow: 0 6px 20px rgba(168,85,247,0.28) !important; transition: opacity 0.2s !important; }
    .stButton > button:hover { opacity: 0.9 !important; }
    .stAlert { border-radius: 10px !important; background: #fdecec !important;
        border: 1px solid #f6c9c9 !important; color: #b42318 !important; }
    </style>
    """, unsafe_allow_html=True)
    st.markdown("""
    <div style="text-align:center; margin-bottom: 32px;">
        <div style="display:inline-flex; align-items:center; gap:10px; margin-bottom: 22px;">
            <div style="width:38px; height:38px; background:linear-gradient(135deg,#a855f7,#ec4899);
                border-radius:10px; display:inline-flex; align-items:center; justify-content:center;
                box-shadow:0 6px 18px rgba(168,85,247,0.3);">
                <div style="width:13px; height:13px; background:white; border-radius:3px; opacity:0.95;"></div>
            </div>
            <span style="font-size:22px; font-weight:800; color:#1f2430; letter-spacing:0.3px;">Vape Shop</span>
            <span style="font-size:22px; font-weight:300; color:#a99bd1;">Porudžbine</span>
        </div>
        <div style="height:1px; background:linear-gradient(90deg, transparent, #e3ddf2, transparent); margin-bottom:26px;"></div>
        <h2 style="color:#1f2430; font-size:23px; font-weight:700; margin:0 0 8px 0; line-height:1.35;">
            Dobrodošli 👋
        </h2>
        <p style="color:#8b8fa0; font-size:14px; margin:0;">
            Unesite šifru za pristup sistemu
        </p>
    </div>
    """, unsafe_allow_html=True)
    pwd = st.text_input("Šifra", type="password", placeholder="Unesite šifru...", label_visibility="collapsed")
    btn = st.button("Prijavi se", use_container_width=True)
    if btn:
        if pwd == APP_PASSWORD:
            st.session_state.authenticated = True
            st.session_state.role = "analitika"
            st.rerun()
        elif pwd == ADMIN_PASSWORD:
            st.session_state.authenticated = True
            st.session_state.role = "administracija"
            st.rerun()
        elif pwd == DIREKTOR_PASSWORD:
            st.session_state.authenticated = True
            st.session_state.role = "direktori"
            st.rerun()
        else:
            st.error("Pogrešna šifra")
    st.markdown("""
    <div style="text-align:center; margin-top:28px;">
        <p style="color:#b9b3c9; font-size:12px; margin:0;">
            Vape Shop · Sistem porudžbina
        </p>
    </div>
    """, unsafe_allow_html=True)
    return False


# =====================================================================
# PREGLED ZA KOLEGINICE (administracija)
# =====================================================================
def _admin_css():
    st.markdown("""<style>
    @import url('https://fonts.googleapis.com/css2?family=Poppins:wght@300;400;500;600;700&display=swap');
    section[data-testid="stSidebar"] { display: none !important; }
    header[data-testid="stHeader"] { display: none !important; }
    #MainMenu { visibility: hidden !important; }
    footer { visibility: hidden !important; }
    .stApp { background: #f5f0ff !important; font-family: 'Poppins', sans-serif; }
    div[data-testid="stMainBlockContainer"], .main .block-container {
        padding: 12px 18px 0 18px !important; max-width: 100% !important; }
    .stButton > button {
        background: linear-gradient(135deg, #a855f7 0%, #ec4899 100%) !important; color: white !important;
        border: none !important; border-radius: 10px !important; font-weight: 600 !important; }
    </style>""", unsafe_allow_html=True)

def _admin_header():
    st.markdown('''<div style="background:#12002a;border-radius:16px;padding:0 28px;height:60px;
        display:flex;align-items:center;justify-content:space-between;margin-bottom:20px;
        border-bottom:3px solid;border-image:linear-gradient(90deg,#a855f7,#ec4899) 1;
        box-shadow:0 4px 20px rgba(18,0,42,0.18);">
        <div style="display:flex;align-items:center;gap:12px;">
            <div style="width:30px;height:30px;background:linear-gradient(135deg,#a855f7,#ec4899);
                border-radius:8px;display:flex;align-items:center;justify-content:center;">
                <div style="width:11px;height:11px;background:white;border-radius:3px;"></div>
            </div>
            <span style="font-size:18px;font-weight:700;color:white;">VAPE</span>
            <span style="font-size:18px;font-weight:300;color:rgba(255,255,255,0.4);">Porudžbine</span>
            <span style="font-size:11px;color:rgba(255,255,255,0.25);margin-left:8px;">·</span>
            <span style="font-size:12px;color:rgba(255,255,255,0.35);">Pregled za administraciju</span>
        </div>
        <div style="display:flex;gap:6px;align-items:center;">
            <div style="width:8px;height:8px;border-radius:50%;background:rgba(168,85,247,0.7);"></div>
            <div style="width:8px;height:8px;border-radius:50%;background:rgba(236,72,153,0.5);"></div>
            <div style="width:8px;height:8px;border-radius:50%;background:rgba(255,255,255,0.15);"></div>
        </div>
    </div>''', unsafe_allow_html=True)

@st.cache_resource
def _pdf_font():
    import os
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    cands = [("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf")]
    try:
        import matplotlib
        _d = os.path.join(matplotlib.get_data_path(), "fonts/ttf")
        cands.append((os.path.join(_d, "DejaVuSans.ttf"), os.path.join(_d, "DejaVuSans-Bold.ttf")))
    except Exception:
        pass
    for _r, _b in cands:
        if os.path.exists(_r):
            try:
                pdfmetrics.registerFont(TTFont("DejaVu", _r))
                if os.path.exists(_b):
                    pdfmetrics.registerFont(TTFont("DejaVu-Bold", _b))
                    return "DejaVu", "DejaVu-Bold"
                return "DejaVu", "DejaVu"
            except Exception:
                pass
    return "Helvetica", "Helvetica-Bold"


def napravi_pdf_izvestaj(mesec_key, mesec_lbl):
    import io as _io
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as _plt
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, KeepTogether, HRFlowable

    FN, FB = _pdf_font()
    H1 = ParagraphStyle('H1', fontName=FB, fontSize=19, leading=23, textColor=colors.HexColor('#3b0764'), spaceAfter=2)
    Hsub = ParagraphStyle('Hsub', fontName=FN, fontSize=10, leading=13, textColor=colors.HexColor('#9188a5'), spaceAfter=14)
    Hsys = ParagraphStyle('Hsys', fontName=FB, fontSize=13, leading=16, textColor=colors.HexColor('#7c3aed'), spaceBefore=2, spaceAfter=4)
    Nar = ParagraphStyle('Nar', fontName=FN, fontSize=10, leading=14.5, textColor=colors.HexColor('#333333'))

    def _chart(nred, norg, ngrn, n, nnas, nnj, nnije, neob, nrev):
        fig, ax = _plt.subplots(1, 2, figsize=(9.0, 2.15))
        ax[0].pie([nred, norg, ngrn], colors=['#e5484d', '#f2820c', '#17a34a'], startangle=90,
                  counterclock=False, wedgeprops=dict(width=0.44, edgecolor='white', linewidth=1.6))
        ax[0].text(0, 0, str(n), ha='center', va='center', fontsize=12, fontweight='bold', color='#3b0764')
        ax[0].set_title('Zone', fontsize=9.5, fontweight='bold', color='#3b0764', pad=3)
        ax[0].legend(["Hitno (" + str(nred) + ")", "Iskontrolisati (" + str(norg) + ")", "Dobra (" + str(ngrn) + ")"],
                     loc='center left', bbox_to_anchor=(0.94, 0.5), frameon=False, fontsize=7.5, handlelength=1)
        ax[1].pie([max(nnas, 0), max(nnj, 0), max(nnije, 0), max(neob, 0)],
                  colors=['#17a34a', '#f2820c', '#c9b8ec', '#e5e2ee'], startangle=90,
                  counterclock=False, wedgeprops=dict(width=0.44, edgecolor='white', linewidth=1.6))
        ax[1].text(0, 0, str(nrev), ha='center', va='center', fontsize=12, fontweight='bold', color='#3b0764')
        ax[1].set_title('Trebovanje', fontsize=9.5, fontweight='bold', color='#3b0764', pad=3)
        ax[1].legend(["Prema našem predlogu (" + str(nnas) + ")", "Po njihovom (" + str(nnj) + ")",
                      "Bez porudžbine (" + str(nnije) + ")", "Neobrađeno (" + str(neob) + ")"],
                     loc='center left', bbox_to_anchor=(0.94, 0.5), frameon=False, fontsize=7.5, handlelength=1)
        fig.subplots_adjust(left=0.01, right=0.72, top=0.84, bottom=0.04, wspace=1.7)
        b = _io.BytesIO()
        fig.savefig(b, format='png', dpi=200, facecolor='white')
        _plt.close(fig)
        b.seek(0)
        return b

    buf = _io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=16 * mm, bottomMargin=14 * mm,
                            leftMargin=18 * mm, rightMargin=18 * mm)
    el = [Paragraph("Izveštaj administracije", H1),
          Paragraph("Mesec: " + mesec_lbl + "   ·   generisano " + datetime.datetime.now().strftime("%d.%m.%Y"), Hsub)]

    sistemi = sb_sisteme(mesec_key)
    if not sistemi:
        el.append(Paragraph("Nema objavljenih sistema za ovaj mesec.", Nar))
        doc.build(el)
        buf.seek(0)
        return buf.getvalue()

    for _sis in sistemi:
        podaci = sb_ucitaj(mesec_key, _sis)
        if not podaci or not podaci.get("stavke"):
            continue
        stavke = podaci["stavke"]
        po = {}
        for s in stavke:
            po.setdefault(int(s["idk"]), []).append(s)
        objekti = []
        for idk, lst in po.items():
            nivo, n0, izg = hitnost_objekta(lst)
            objekti.append({"idk": idk, "nivo": nivo})
        obrada = sb_load_obrada(mesec_key, _sis)

        def _ima(o, naz):
            return naz in (obrada.get(o["idk"], {}).get("reakcije") or [])

        n = len(objekti)
        nred = sum(1 for o in objekti if o["nivo"] == "crveno")
        norg = sum(1 for o in objekti if o["nivo"] == "zuto")
        ngrn = sum(1 for o in objekti if o["nivo"] == "zeleno")
        rev = [o for o in objekti if obrada.get(o["idk"], {}).get("reakcije")]
        nrev = len(rev)
        nPoz = sum(1 for o in objekti if _ima(o, "Pozvala sam"))
        nMej = sum(1 for o in objekti if _ima(o, "Poslala sam mejl"))
        nDir = sum(1 for o in objekti if _ima(o, "Obavestila direktorku"))
        nnas = sum(1 for o in rev if obrada.get(o["idk"], {}).get("trebovali_tip") == "nas")
        nnj = sum(1 for o in rev if obrada.get(o["idk"], {}).get("trebovali_tip") == "njihov")
        nnije = sum(1 for o in rev if not obrada.get(o["idk"], {}).get("trebovali_tip"))
        neob = n - nrev
        pct = round(nrev / max(n, 1) * 100)

        narr = ("Od <b>" + str(n) + "</b> objekata za porudžbinu, <b><font color='#d33'>" + str(nred) +
                "</font></b> je u hitnoj zoni, " + str(norg) + " za kontrolu i " + str(ngrn) + " u dobroj. "
                "Obrađeno je <b>" + str(nrev) + "</b> (" + str(pct) + "%) — pozvano " + str(nPoz) +
                ", mejlova " + str(nMej) + ", direktorki " + str(nDir) + ". "
                "Trebovanje: <b><font color='#158a3f'>" + str(nnas) + " prema našem predlogu</font></b>, "
                "<font color='#c66a00'>" + str(nnj) + " po njihovom</font>, " + str(nnije) +
                " bez porudžbine; <b>" + str(neob) + "</b> neobrađeno.")
        img = Image(_chart(nred, norg, ngrn, n, nnas, nnj, nnije, neob, nrev), width=150 * mm, height=150 * mm * 2.15 / 9.0)
        img.hAlign = 'CENTER'
        el.append(KeepTogether([
            Paragraph(str(_sis), Hsys),
            Paragraph(narr, Nar),
            Spacer(1, 2),
            img,
            Spacer(1, 4),
            HRFlowable(width="100%", thickness=0.6, color=colors.HexColor('#eae4f7'), spaceAfter=6),
        ]))

    doc.build(el)
    buf.seek(0)
    return buf.getvalue()


def _admin_order_xlsx(rows):
    """rows = lista tuplova (id_kupca, id_artikla, kolicina).
    Vraca bajtove .xlsx u formatu koji admin (Nova porudzbina iz Excel-a) ocekuje:
    kolone 'Id kupca', 'Id artikla', 'Kolicina' (tacno kao u sablonu)."""
    import io as _io
    from openpyxl import Workbook as _WB
    _wb = _WB()
    _ws = _wb.active
    _ws.title = "Sheet1"
    _ws.append(["Id kupca", "Id artikla", "Količina"])
    for _k, _a, _q in rows:
        _ws.append([int(_k), int(_a), int(_q)])
    _buf = _io.BytesIO()
    _wb.save(_buf)
    return _buf.getvalue()


def _admin_secret(k, d=""):
    try:
        import streamlit as _s
        if hasattr(_s, "secrets") and k in _s.secrets:
            return str(_s.secrets[k])
    except Exception:
        pass
    import os
    return os.environ.get(k, d)


def _sb_select_all(table, columns, step=1000):
    """Pročitaj SVE redove iz tabele kroz paginaciju (Supabase/PostgREST vraća
    max 1000 po upitu). Vraća listu dict-ova."""
    cli = _sb()
    if cli is None:
        return []
    out = []
    start = 0
    while True:
        res = cli.table(table).select(columns).range(start, start + step - 1).execute()
        batch = res.data or []
        out.extend(batch)
        if len(batch) < step:
            break
        start += step
    return out


def sb_komitenti_map():
    if _sb() is None:
        return {}
    try:
        rows = _sb_select_all("komitenti", "idk,naziv")
        return {int(r["idk"]): (r.get("naziv") or "") for r in rows}
    except Exception:
        return {}


def sb_komitenti_save(mapping):
    cli = _sb()
    if cli is None:
        raise RuntimeError("Supabase nije podešen.")
    rows = [{"idk": int(k), "naziv": v} for k, v in mapping.items()]
    for _i in range(0, len(rows), 500):
        cli.table("komitenti").upsert(rows[_i:_i + 500], on_conflict="idk").execute()


def sb_komitenti_full():
    """Vrati {idk: {naziv,email,telefon,mesto,adresa}} za SVE komitente (paginacija).
    Radi i ako kolone kontakata još ne postoje (tada su prazne)."""
    if _sb() is None:
        return {}
    try:
        rows = _sb_select_all("komitenti", "idk,naziv,email,telefon,mesto,adresa")
        out = {}
        for r in rows:
            out[int(r["idk"])] = {"naziv": r.get("naziv") or "", "email": r.get("email") or "",
                                  "telefon": r.get("telefon") or "", "mesto": r.get("mesto") or "",
                                  "adresa": r.get("adresa") or ""}
        return out
    except Exception:
        try:
            rows = _sb_select_all("komitenti", "idk,naziv")
            return {int(r["idk"]): {"naziv": r.get("naziv") or "", "email": "", "telefon": "",
                                    "mesto": "", "adresa": ""} for r in rows}
        except Exception:
            return {}


def sb_komitenti_upsert_rows(rows):
    """rows: lista dict sa poljima idk,naziv,email,telefon,mesto,adresa. Upsert po idk.
    Vraca broj komitenata koji SU ZAISTA u bazi posle upisa (verifikovano čitanjem).
    Ako baza vrati 0 a poslali smo >0 -> tabela nije dobro podešena (baca grešku)."""
    cli = _sb()
    if cli is None:
        raise RuntimeError("Supabase nije podešen.")
    clean = []
    for r in rows:
        try:
            _id = int(r.get("idk"))
        except Exception:
            continue
        clean.append({"idk": _id, "naziv": r.get("naziv") or "", "email": r.get("email") or "",
                      "telefon": r.get("telefon") or "", "mesto": r.get("mesto") or "",
                      "adresa": r.get("adresa") or ""})
    if not clean:
        return 0

    def _do_upsert(_rows):
        for _i in range(0, len(_rows), 500):
            cli.table("komitenti").upsert(_rows[_i:_i + 500], on_conflict="idk").execute()

    _err = None
    try:
        _do_upsert(clean)
    except Exception as _e1:
        _err = _e1
        # Fallback: možda kolone email/telefon/mesto/adresa još ne postoje -> probaj samo idk+naziv
        slim = [{"idk": r["idk"], "naziv": r["naziv"]} for r in clean]
        try:
            _do_upsert(slim)
            _err = None
        except Exception as _e2:
            _err = _e2

    # Verifikacija: pročitaj nazad koliko redova zaista ima u tabeli posle upisa
    _u_bazi = None
    try:
        _chk = cli.table("komitenti").select("idk", count="exact").limit(1).execute()
        _u_bazi = _chk.count
    except Exception:
        _u_bazi = None

    if _u_bazi is not None:
        if _u_bazi == 0 and len(clean) > 0:
            raise RuntimeError(
                "Upis nije sačuvan — u tabeli 'komitenti' ima 0 redova posle upisa. "
                "Tabela najverovatnije nije podešena kako treba (pokreni SQL setup). "
                + (("Detalj: " + str(_err)) if _err else ""))
        return _u_bazi
    if _err is not None:
        raise RuntimeError("Upis nije uspeo: " + str(_err))
    return len(clean)


def posalji_u_admin(id_kupca, items):
    """Prijavi se u admin i kreira porudzbinu direktno (bez rucnog uvoza).
    items = lista dict {'idArticle': int, 'quantity': int}.
    Vraca (ok: bool, poruka: str)."""
    import re, requests
    base = (_admin_secret("ADMIN_BASE_URL", "https://admin.vapeshop.rs") or "").rstrip("/")
    email = _admin_secret("ADMIN_LOGIN_EMAIL", "")
    pwd = _admin_secret("ADMIN_LOGIN_PASSWORD", "")
    if not email or not pwd:
        return (False, "Nije podešena admin prijava. Analitičar treba da doda ADMIN_LOGIN_EMAIL i ADMIN_LOGIN_PASSWORD u Secrets.")
    items = [{"idArticle": int(i["idArticle"]), "quantity": int(i["quantity"])}
             for i in items if int(i.get("quantity", 0)) > 0]
    if not items:
        return (False, "Nema stavki za slanje (sve količine su 0).")
    _tok = re.compile(r'name="__RequestVerificationToken"[^>]*value="([^"]+)"')
    s = requests.Session()
    s.headers.update({"User-Agent": "Mozilla/5.0", "Accept-Language": "sr,en;q=0.8"})
    try:
        r = s.get(base + "/login", timeout=30)
        m = _tok.search(r.text)
        if not m:
            return (False, "Ne mogu da otvorim login stranicu admina (proveri ADMIN_BASE_URL).")
        s.post(base + "/login",
               data={"Email": email, "Password": pwd,
                     "__RequestVerificationToken": m.group(1)},
               headers={"Referer": base + "/login"}, timeout=30, allow_redirects=True)
        chk = s.get(base + "/orders-processing/new-order-from-excel", timeout=30)
        if ("/login" in chk.url) or ('name="Password"' in chk.text):
            return (False, "Prijava na admin nije uspela. Proveri email/lozinku u Secrets — ili admin blokira pristup sa servera aplikacije.")
        m2 = _tok.search(chk.text)
        if not m2:
            return (False, "Ne mogu da nađem sigurnosni token na stranici za uvoz.")
        xlsx = _admin_order_xlsx([(id_kupca, it["idArticle"], it["quantity"]) for it in items])
        # Jedan korak: uvoz Excel-a = admin ODMAH kreira porudžbinu (kao ručno).
        # Namerno NE zovemo "create-order-from-in-memory-cart" da se ne naprave dve.
        up = s.post(base + "/orders-processing/load-order-from-excel",
                    files={"fileArticles": ("porudzbina.xlsx", xlsx,
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                    data={"__RequestVerificationToken": m2.group(1)},
                    headers={"Referer": base + "/orders-processing/new-order-from-excel"},
                    timeout=90, allow_redirects=True)
        if up.status_code >= 400:
            return (False, "Slanje u admin nije prošlo (status " + str(up.status_code) + ").")
        _low = (up.text or "").lower()
        if ("greška" in _low) or ("greska" in _low) or ("error" in _low) or ("nije prona" in _low):
            return (False, "Admin je prijavio problem pri uvozu (proveri da su šifre artikala ispravne).")
        _mid = re.search(r'/order/details/(\d+)', up.text or "")
        if _mid:
            return (True, "Porudžbina #" + _mid.group(1) + " je kreirana u adminu.")
        return (True, "Porudžbina je poslata u admin (vidi „Pregled porudžbina“).")
    except requests.exceptions.RequestException as _e:
        return (False, "Greška u komunikaciji sa adminom: " + str(_e))


# =====================================================================
# ISTORIJA PORUDŽBINA IZ ADMINA (za detaljnu karticu)
# =====================================================================
def _admin_session():
    """Prijavi se u admin. Vraca (session, base_url, greska)."""
    import re, requests
    base = (_admin_secret("ADMIN_BASE_URL", "https://admin.vapeshop.rs") or "").rstrip("/")
    email = _admin_secret("ADMIN_LOGIN_EMAIL", "")
    pwd = _admin_secret("ADMIN_LOGIN_PASSWORD", "")
    if not email or not pwd:
        return (None, base, "Nije podešena admin prijava (Secrets).")
    _tok = re.compile(r'name="__RequestVerificationToken"[^>]*value="([^"]+)"')
    s = requests.Session()
    s.headers.update({"User-Agent": "Mozilla/5.0", "Accept-Language": "sr,en;q=0.8"})
    try:
        r = s.get(base + "/login", timeout=30)
        m = _tok.search(r.text)
        if not m:
            return (None, base, "Ne mogu da otvorim login admina.")
        s.post(base + "/login",
               data={"Email": email, "Password": pwd, "__RequestVerificationToken": m.group(1)},
               headers={"Referer": base + "/login"}, timeout=30, allow_redirects=True)
        chk = s.get(base + "/orders", timeout=30)
        if ("/login" in chk.url) or ('name="Password"' in chk.text):
            return (None, base, "Prijava na admin nije uspela (proveri Secrets / pristup).")
        return (s, base, "")
    except requests.exceptions.RequestException as _e:
        return (None, base, "Greška u vezi sa adminom: " + str(_e))


def _norm_ws(t):
    import re
    return re.sub(r"\s+", " ", (t or "")).strip()


def _h_escape(t):
    import html
    return html.escape(str(t or ""))


def _clean_komitent_naziv(t):
    """Skini rep tipa ' - [mejl] VP' iz naziva komitenta (kako izgleda u padajucem
    spisku), da bi se poklopio sa imenom u listi porudzbina."""
    import re
    t = _norm_ws(t)
    t = re.sub(r'\s*-\s*\[[^\]]*\]\s*[A-Za-z]{0,3}\s*$', '', t)  # ' - [..] VP'
    t = re.sub(r'\s*\[[^\]]*\]\s*[A-Za-z]{0,3}\s*$', '', t)      # '[..] VP' bez crtice
    return _norm_ws(t)


def admin_build_komitenti(only_ids=None):
    """Napravi mapu {id_kupca: naziv} iz padajuce liste na stranici jedne porudzbine
    i sacuvaj u Supabase tabelu 'komitenti'. Ako je only_ids zadat (skup/lista id-jeva),
    cuva SAMO te (da ne prebrise nazive/kontakte iz fajla). Vraca (broj, greska)."""
    import re, html
    s, base, err = _admin_session()
    if err:
        return (0, err)
    try:
        lst = s.get(base + "/orders", timeout=45)
        _ids = re.findall(r'/order/details/(\d+)', lst.text)
        if not _ids:
            return (0, "Nema porudžbina za čitanje šifara komitenata.")
        det = s.get(base + "/order/details/" + _ids[0], timeout=45)
        _sel = re.search(r'<select[^>]*id="selectUser"[^>]*>(.*?)</select>', det.text, re.S)
        if not _sel:
            return (0, "Ne mogu da nađem listu komitenata.")
        mapa = {}
        for _v, _naz in re.findall(r'<option value="(\d+)"[^>]*>(.*?)</option>', _sel.group(1), re.S):
            mapa[int(_v)] = _clean_komitent_naziv(html.unescape(_naz))
        if only_ids is not None:
            _want = set(int(x) for x in only_ids)
            mapa = {k: v for k, v in mapa.items() if k in _want}
        if not mapa:
            return (0, "Nema naziva za povlačenje (lista prazna ili nijedan ID se ne poklapa).")
        try:
            sb_komitenti_save(mapa)
        except Exception as _e:
            return (len(mapa), "Pročitano " + str(len(mapa)) + ", ali čuvanje nije uspelo: " + str(_e))
        return (len(mapa), "")
    except Exception as _e:
        return (0, "Greška: " + str(_e))


def _parse_order_items(html_text):
    """Iz /order/details/{id} izvuci stavke: [{ida, naziv, kol, cena}] i idUser."""
    import re, html as _h
    _iduser = None
    _mu = re.search(r'data-original-user="(\d+)"', html_text)
    if _mu:
        _iduser = int(_mu.group(1))
    _blok = re.search(r'id="panelOrderItems".*?<tbody>(.*?)</tbody>', html_text, re.S)
    stavke = []
    if _blok:
        for _row in re.findall(r'<tr[^>]*>(.*?)</tr>', _blok.group(1), re.S):
            _a = re.search(r'/article/edit/(\d+)"[^>]*>(.*?)</a>', _row, re.S)
            _pc = re.findall(r'class="price-cell">\s*([^<]*?)\s*</td>', _row)
            if not _a:
                continue
            _kol = _pc[0].strip() if len(_pc) >= 1 else ""
            _cena = _pc[1].strip() if len(_pc) >= 2 else ""
            stavke.append({"ida": int(_a.group(1)),
                           "naziv": _norm_ws(_h.unescape(_a.group(2))),
                           "kol": _kol, "cena": _cena})
    return stavke, _iduser


def admin_istorija_komitenta(id_kupca, naziv, datum_od, datum_do, max_por=40):
    """Vrati listu porudzbina komitenta iz admina u periodu [datum_od, datum_do]
    (format dd.MM.yyyy), sa stavkama. Vraca (lista, greska)."""
    import re, html as _h
    import datetime as _dt
    s, base, err = _admin_session()
    if err:
        return ([], err)
    naziv_n = _clean_komitent_naziv(naziv)
    # granica: datum_od (dd.MM.yyyy) -> date za poredjenje na nasoj strani
    try:
        _cut = _dt.datetime.strptime(datum_od, "%d.%m.%Y").date()
    except Exception:
        _cut = None
    try:
        # Prazan datum na serveru = vrati sve, pa filtriramo kod nas (izbegava format datuma).
        resp = s.post(base + "/orders", data={
            "orderStatuses": ["1", "10", "20", "30", "40", "50", "60", "70", "80", "90", "95"],
            "keyword": "", "idLoad": "", "startDate": "", "endDate": ""},
            headers={"Referer": base + "/orders", "X-Requested-With": "XMLHttpRequest"}, timeout=90)
        body = resp.text or ""
        rezultat = []
        for _row in re.findall(r'<tr[^>]*>(.*?)</tr>', body, re.S):
            if '/order/details/' not in _row:
                continue
            _rowtxt = _norm_ws(_h.unescape(_row))
            if naziv_n and naziv_n not in _rowtxt:
                continue
            _mid = re.search(r'/order/details/(\d+)', _row)
            if not _mid:
                continue
            _oid = _mid.group(1)
            _md = re.search(r'(\d{2}\.\d{2}\.\d{4})(?:\s+(\d{2}:\d{2}))?', _rowtxt)
            _datum = (_md.group(1) + (" " + _md.group(2) if _md.group(2) else "")) if _md else ""
            # filter po datumu (poslednja ~3 meseca)
            if _cut and _md:
                try:
                    _dd = _dt.datetime.strptime(_md.group(1), "%d.%m.%Y").date()
                    if _dd < _cut:
                        continue
                except Exception:
                    pass
            _ms = re.search(r'labelOrderStatus[^"]*"[^>]*>([^<]+)</span>', _row)
            _status = _norm_ws(_h.unescape(_ms.group(1))) if _ms else ""
            _mc = re.search(r'price-cell">\s*([\d.]+)\s*RSD', _row)
            _cena = (_mc.group(1) + " RSD") if _mc else ""
            rezultat.append({"id": _oid, "datum": _datum, "status": _status, "cena": _cena})
        # najnovije prvo, ogranicenje
        rezultat = rezultat[:max_por]
        for _o in rezultat:
            try:
                det = s.get(base + "/order/details/" + _o["id"], timeout=45)
                _st, _iu = _parse_order_items(det.text)
                _o["stavke"] = _st
            except Exception:
                _o["stavke"] = []
        return (rezultat, "")
    except Exception as _e:
        return ([], "Greška pri čitanju istorije: " + str(_e))


def _datum_sort_key(_o):
    """Sortiranje porudžbina po datumu (najnovije prvo)."""
    import datetime as _d
    _s = (_o.get("datum") or "").split(" ")[0]
    try:
        return _d.datetime.strptime(_s, "%d.%m.%Y")
    except Exception:
        return _d.datetime.min


def admin_istorija_bulk(idk_naziv, cutoff_date, max_details=800):
    """Za ceo sistem odjednom: jedan login + jedan /orders POST, pa detalji SAMO za
    porudžbine >= cutoff_date koje se poklapaju sa objektima (po nazivu, potvrda idUser).
    idk_naziv: {idk: naziv}. Vrati ({idk: [ {id,datum,status,cena,stavke} ]}, greska)."""
    import re, html as _h, datetime as _dt
    s, base, err = _admin_session()
    if err:
        return ({}, err)
    name_items = []
    for _idk, _nz in idk_naziv.items():
        _n = _clean_komitent_naziv(_nz or "")
        if _n:
            name_items.append((int(_idk), _n))
    want_ids = set(int(x) for x in idk_naziv.keys())
    try:
        resp = s.post(base + "/orders", data={
            "orderStatuses": ["1", "10", "20", "30", "40", "50", "60", "70", "80", "90", "95"],
            "keyword": "", "idLoad": "", "startDate": "", "endDate": ""},
            headers={"Referer": base + "/orders", "X-Requested-With": "XMLHttpRequest"}, timeout=120)
        body = resp.text or ""
        cand = []
        for _row in re.findall(r'<tr[^>]*>(.*?)</tr>', body, re.S):
            if '/order/details/' not in _row:
                continue
            _rowtxt = _norm_ws(_h.unescape(_row))
            _md = re.search(r'(\d{2}\.\d{2}\.\d{4})(?:\s+(\d{2}:\d{2}))?', _rowtxt)
            if cutoff_date:
                if not _md:
                    continue
                try:
                    if _dt.datetime.strptime(_md.group(1), "%d.%m.%Y").date() < cutoff_date:
                        continue
                except Exception:
                    continue
            _hit = None
            for _idk, _n in name_items:
                if _n in _rowtxt:
                    _hit = _idk
                    break
            if _hit is None:
                continue
            _mid = re.search(r'/order/details/(\d+)', _row)
            if not _mid:
                continue
            _oid = _mid.group(1)
            _datum = (_md.group(1) + ((" " + _md.group(2)) if _md.group(2) else "")) if _md else ""
            _ms = re.search(r'labelOrderStatus[^"]*"[^>]*>([^<]+)</span>', _row)
            _status = _norm_ws(_h.unescape(_ms.group(1))) if _ms else ""
            _sl = _status.lower()
            if ("otkaz" in _sl) or ("storn" in _sl) or ("odbij" in _sl) or ("ponist" in _sl) or ("poništ" in _sl):
                continue  # otkazane/stornirane ne prikazujemo
            _mc = re.search(r'price-cell">\s*([\d.]+)\s*RSD', _row)
            _cena = (_mc.group(1) + " RSD") if _mc else ""
            cand.append((_hit, _oid, _datum, _status, _cena))
        cand = cand[:max_details]
        out = {}
        for (_hit, _oid, _datum, _status, _cena) in cand:
            try:
                det = s.get(base + "/order/details/" + _oid, timeout=45)
                _st, _iu = _parse_order_items(det.text)
                _final = _iu if (_iu in want_ids) else _hit
            except Exception:
                _st = []
                _final = _hit
            out.setdefault(_final, []).append(
                {"id": _oid, "datum": _datum, "status": _status, "cena": _cena, "stavke": _st})
        return (out, "")
    except Exception as _e:
        return ({}, "Greška pri čitanju istorije: " + str(_e))


def _to_int_kol(s):
    """Parsiraj količinu iz stringa (npr. '12', '12,00', '12.0') u int."""
    import re
    s = str(s).strip()
    if not s:
        return 0
    s = re.split(r'[.,]', s)[0]
    s = re.sub(r'[^\d-]', '', s)
    try:
        return int(s)
    except Exception:
        return 0


def _treb_posle_preseka(hist_lst, cutoff_date):
    """Iz učitane istorije porudžbina saberi količine po artiklu za porudžbine
    datirane >= cutoff_date, izuzimajući otkazane/stornirane. Vrati {ida: kom}."""
    import datetime as _dt
    out = {}
    if not hist_lst:
        return out
    for _o in hist_lst:
        _dstr = (_o.get("datum") or "").strip()
        _md = _dstr.split(" ")[0] if _dstr else ""
        _ok_datum = True
        if cutoff_date and _md:
            try:
                _dd = _dt.datetime.strptime(_md, "%d.%m.%Y").date()
                _ok_datum = _dd >= cutoff_date
            except Exception:
                _ok_datum = True
        if not _ok_datum:
            continue
        _stat = (_o.get("status") or "").lower()
        if ("otkaz" in _stat) or ("storn" in _stat) or ("odbij" in _stat) or ("ponist" in _stat) or ("poništ" in _stat):
            continue
        for _s in (_o.get("stavke") or []):
            try:
                _ida = int(_s.get("ida"))
            except Exception:
                continue
            out[_ida] = out.get(_ida, 0) + _to_int_kol(_s.get("kol"))
    return out


def prikazi_administraciju():
    st.set_page_config(page_title="VAPE — Porudžbine", page_icon="📦",
                       layout="wide", initial_sidebar_state="collapsed")
    st.markdown("""<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
    section[data-testid="stSidebar"]{display:none !important;}
    header[data-testid="stHeader"]{display:none !important;}
    #MainMenu{visibility:hidden !important;} footer{visibility:hidden !important;}
    .stApp{background:#fbfbfd !important;font-family:'Inter',sans-serif;}
    div[data-testid="stMainBlockContainer"]{padding:10px 26px 40px !important;max-width:100% !important;}
    /* header */
    .adm-hdr{display:flex;align-items:center;gap:11px;padding:16px 0 16px;border-bottom:1px solid #eef0f4;margin-bottom:22px;}
    .adm-logo{width:26px;height:26px;border-radius:8px;background:linear-gradient(135deg,#a855f7,#ec4899);display:flex;align-items:center;justify-content:center;}
    .adm-logo div{width:9px;height:9px;background:#fff;border-radius:2px;}
    .adm-hdr .t1{font-weight:700;font-size:16px;letter-spacing:-.2px;color:#1f2430;}
    .adm-hdr .t2{color:#9ca3af;font-weight:500;font-size:13px;}
    /* kpi */
    .adm-kpi{display:flex;border:1px solid #eef0f4;border-radius:12px;background:#fff;overflow:hidden;margin:4px 0 14px;}
    .adm-kpi .cell{flex:1;padding:15px 20px;border-right:1px solid #f2f3f7;}
    .adm-kpi .cell:last-child{border-right:none;}
    .adm-kpi .n{font-size:22px;font-weight:700;letter-spacing:-.5px;display:flex;align-items:center;gap:8px;color:#1f2430;}
    .adm-kpi .n .d{width:8px;height:8px;border-radius:50%;}
    .adm-kpi .n .d.r{background:#e5484d;} .adm-kpi .n .d.o{background:#f2820c;} .adm-kpi .n .d.g{background:#17a34a;} .adm-kpi .n .d.p{background:#7c3aed;}
    .adm-kpi .k{font-size:12px;color:#9ca3af;margin-top:3px;font-weight:500;}
    .adm-kpi .cell.c-p{background:#faf7ff;} .adm-kpi .cell.c-r{background:#fff7f7;} .adm-kpi .cell.c-o{background:#fffcf5;} .adm-kpi .cell.c-g{background:#f6fdf9;}
    .adm-kpi .cell.c-p .n{color:#7c3aed;} .adm-kpi .cell.c-r .n{color:#d33;} .adm-kpi .cell.c-o .n{color:#c66a00;} .adm-kpi .cell.c-g .n{color:#158a3f;}
    table.adm-t tr.row-red td{background:#fff8f8;} table.adm-t tr.row-org td{background:#fffcf6;}
    /* progress */
    .adm-prog{display:flex;align-items:center;gap:12px;margin-bottom:8px;}
    .adm-prog .t{font-size:12.5px;color:#6b7280;font-weight:600;white-space:nowrap;}
    .adm-prog .bar{flex:1;height:5px;background:#eef0f4;border-radius:99px;overflow:hidden;}
    .adm-prog .bar>div{height:100%;background:#7c3aed;border-radius:99px;}
    /* tables */
    table.adm-t{width:100%;border-collapse:collapse;}
    table.adm-t th{text-align:left;font-size:11px;color:#b0b4bd;font-weight:600;text-transform:uppercase;letter-spacing:.5px;padding:0 14px 12px;}
    table.adm-t td{padding:13px 14px;border-top:1px solid #f2f3f7;font-size:14px;vertical-align:middle;color:#2a2f3a;}
    table.adm-t td.idc{font-weight:700;} table.adm-t td.ce{text-align:center;} table.adm-t td.mut{color:#b0b4bd;}
    .zona{display:inline-flex;align-items:center;gap:7px;font-weight:600;font-size:13px;white-space:nowrap;}
    .zona .zd{width:8px;height:8px;border-radius:50%;}
    .z-red{color:#d33;} .z-red .zd{background:#e5484d;}
    .z-org{color:#c66a00;} .z-org .zd{background:#f2820c;}
    .z-grn{color:#158a3f;} .z-grn .zd{background:#17a34a;}
    .stat{font-size:12.5px;color:#9ca3af;}
    .stchip{display:inline-block;font-size:11.5px;color:#5b21b6;background:#f2effc;border-radius:6px;padding:2px 8px;margin:1px 3px 1px 0;}
    .tb-nas{color:#158a3f;font-weight:600;font-size:12.5px;} .tb-nj{color:#c66a00;font-weight:600;font-size:12.5px;}
    .np{color:#c4c7cf;}
    .adot{width:9px;height:9px;border-radius:50%;display:inline-block;}
    .ar-red{background:#e5484d;} .ar-yel{background:#f2820c;} .ar-grn{background:#17a34a;}
    /* empty */
    .adm-empty{text-align:center;padding:84px 20px;}
    .adm-empty .ic{font-size:40px;opacity:.5;margin-bottom:14px;}
    .adm-empty h2{font-size:19px;font-weight:600;color:#374151;margin-bottom:6px;}
    .adm-empty .w{color:#9ca3af;font-size:14px;margin-bottom:14px;}
    .adm-empty .plan{display:inline-block;color:#7c3aed;font-size:14px;font-weight:600;border:1px solid #ede9fe;background:#faf8ff;border-radius:99px;padding:8px 18px;}
    .adm-empty p{color:#b0b4bd;font-size:13px;margin-top:14px;}
    /* detail */
    .adm-dh{display:flex;align-items:center;gap:14px;margin:6px 0 18px;}
    .adm-dh .id{font-size:22px;font-weight:700;letter-spacing:-.5px;color:#1f2430;}
    .adm-dh .mut{color:#b0b4bd;font-size:13px;}
    .adm-lbl{font-size:11px;color:#b0b4bd;font-weight:600;text-transform:uppercase;letter-spacing:.5px;margin:2px 0 8px;}
    .revy{color:#158a3f;font-weight:600;font-size:12.5px;} .revn{color:#c4c7cf;font-weight:600;font-size:12.5px;}
    /* streamlit kontrole suptilnije */
    .stButton>button{border-radius:9px;font-weight:600;}
    button[data-testid="baseButton-primary"]{background:#7c3aed !important;border-color:#7c3aed !important;color:#fff !important;}
    /* dugmad za slanje u admin — jasno obojena, kompaktna */
    [class*="st-key-axn_"] button{background:#16a34a !important;border-color:#16a34a !important;color:#fff !important;font-weight:600 !important;font-size:13px !important;padding:7px 12px !important;border-radius:8px !important;box-shadow:none !important;}
    [class*="st-key-axn_"] button:hover{background:#128a3e !important;border-color:#128a3e !important;}
    [class*="st-key-axj_"] button{background:#f59e0b !important;border-color:#f59e0b !important;color:#fff !important;font-weight:600 !important;font-size:13px !important;padding:7px 12px !important;border-radius:8px !important;box-shadow:none !important;}
    [class*="st-key-axj_"] button:hover{background:#d97706 !important;border-color:#d97706 !important;}
    .stButton button[kind="primary"]{background:#7c3aed !important;border-color:#7c3aed !important;color:#fff !important;font-size:13.5px !important;padding:8px 14px !important;border-radius:8px !important;font-weight:600 !important;}
    /* zaglavlje: mala, uredna dugmad */
    [class*="st-key-adm_odjava"] button{font-size:11px !important;padding:3px 8px !important;border-radius:7px !important;min-height:0 !important;line-height:1.2 !important;background:#fff !important;border:1px solid #e5e7eb !important;color:#6b7280 !important;font-weight:600 !important;box-shadow:none !important;}
    [class*="st-key-adm_odjava"] button:hover{background:#f9fafb !important;color:#374151 !important;}
    [class*="st-key-refresh_all_admin"] button{font-size:11px !important;padding:3px 8px !important;border-radius:7px !important;min-height:0 !important;line-height:1.2 !important;background:#7c3aed !important;border-color:#7c3aed !important;color:#fff !important;font-weight:600 !important;box-shadow:none !important;}
    [class*="st-key-refresh_all_admin"] button:hover{background:#6d28d9 !important;border-color:#6d28d9 !important;}
    [class*="st-key-predaj_izvestaj"] button{font-size:11px !important;padding:3px 8px !important;border-radius:7px !important;min-height:0 !important;line-height:1.2 !important;background:#0ea5e9 !important;border-color:#0ea5e9 !important;color:#fff !important;font-weight:600 !important;box-shadow:none !important;}
    [class*="st-key-predaj_izvestaj"] button:hover{background:#0284c7 !important;border-color:#0284c7 !important;}
    .stMultiSelect [data-baseweb="tag"]{background:#f2effc !important;color:#5b21b6 !important;border:none !important;}
    .stMultiSelect [data-baseweb="tag"] span{color:#5b21b6 !important;}
    </style>""", unsafe_allow_html=True)

    _hc1, _hc2 = st.columns([7.5, 1.15])
    with _hc1:
        st.markdown('<div class="adm-hdr"><div class="adm-logo"><div></div></div>'
                    '<span class="t1">VAPE Porudžbine</span><span class="t2">· Administracija</span></div>',
                    unsafe_allow_html=True)
    with _hc2:
        st.markdown("<div style='height:10px;'></div>", unsafe_allow_html=True)
        if st.button("Odjava", key="adm_odjava", use_container_width=True):
            for _k in ("authenticated", "role"):
                st.session_state.pop(_k, None)
            st.rerun()
        if st.button("🔄 Ažuriraj iz admina", key="refresh_all_admin",
                     use_container_width=True, type="primary"):
            st.session_state["_req_refresh_admin"] = True
        if st.button("📝 Predaj izveštaj", key="predaj_izvestaj", use_container_width=True):
            st.session_state["_req_predaj"] = True

    if not sb_dostupan():
        st.error("Veza sa bazom trenutno nije podešena. Javi se analitičaru.")
        return

    _adm_mode = st.radio("Prikaz", ["📦 Porudžbine", "💳 Potraživanja"], horizontal=True,
                         key="adm_mode", label_visibility="collapsed")
    if "Potra" in _adm_mode:
        potraz_admin_ui()
        return

    _pub = sb_meseci()
    _mes_keys = [m["key"] for m in _pub]
    for _k in sb_plan_meseci():
        if _k not in _mes_keys:
            _mes_keys.append(_k)
    _mes_keys = sorted(set(_mes_keys), reverse=True)
    if not _mes_keys:
        st.info("Još nema objavljenih podataka. Analitičar treba prvo da objavi bar jedan sistem.")
        return

    _c1, _c2, _c3 = st.columns([1, 1, 3])
    _mlbls = [mesec_label(k) for k in _mes_keys]
    with _c1:
        _sel_lbl = st.selectbox("Mesec", _mlbls, index=0, key="adm_mes")
    mesec_key = _mes_keys[_mlbls.index(_sel_lbl)]
    _imaju = sb_sisteme(mesec_key)
    _svi = sb_svi_sistemi()
    _sis_opts = sorted(set(_svi) | set(_imaju))
    with _c2:
        if _sis_opts:
            sistem = st.selectbox("Sistem", _sis_opts, index=0, key="adm_sis")
        else:
            sistem = None
            st.selectbox("Sistem", ["(nema)"], disabled=True)

    if not sistem:
        st.warning("Nema dostupnih sistema.")
        return

    # Rok koji je direktor postavio za ovaj mesec (administracija radi do tog roka)
    _rk_adm = sb_rokovi_get(mesec_key).get("rok_admin")
    if _rk_adm:
        _proso = _rok_je_prosao(_rk_adm)
        _bg = "#fef2f2;border-color:#fecaca;color:#b42318" if _proso else "#f0fdf4;border-color:#bbf7d0;color:#166534"
        _txt = ("Rok je istekao (" + _rok_fmt(_rk_adm) + ") — izveštaj je zaključan i predat direktoru."
                if _proso else "Rok za predaju izveštaja (" + _sel_lbl + "): " + _rok_fmt(_rk_adm))
        st.markdown('<div style="background:' + _bg + ';border:1px solid;border-radius:10px;padding:9px 14px;'
                    'font-size:12.5px;font-weight:600;margin:2px 0 12px;">⏰ ' + _txt + '</div>', unsafe_allow_html=True)

    _pcx = st.columns([1.5, 1, 2])
    with _pcx[0]:
        if st.button("📄 Napravi PDF izveštaj (" + _sel_lbl + ")", key="pdf_make", use_container_width=True):
            try:
                with st.spinner("Pravim PDF..."):
                    st.session_state["pdf_bytes"] = napravi_pdf_izvestaj(mesec_key, _sel_lbl)
                    st.session_state["pdf_mes"] = mesec_key
            except Exception as _pe:
                st.session_state["pdf_bytes"] = None
                st.error("Greška pri pravljenju PDF-a: " + str(_pe))
    with _pcx[1]:
        if st.session_state.get("pdf_bytes") and st.session_state.get("pdf_mes") == mesec_key:
            st.download_button("⬇️ Preuzmi PDF", st.session_state["pdf_bytes"],
                file_name="Izvestaj_administracije_" + mesec_key + ".pdf", mime="application/pdf",
                key="pdf_dl", use_container_width=True)

    podaci = sb_ucitaj(mesec_key, sistem)
    if not podaci or not podaci.get("stavke"):
        _plan = sb_load_plan(mesec_key)
        if _plan:
            _ph = '<div class="plan">Planirana objava do ' + str(_plan) + '.</div>'
        else:
            _ph = '<div class="plan">Datum objave još nije zakazan.</div>'
        st.markdown('<div class="adm-empty"><div class="ic">🗓️</div>'
                    '<h2>Izveštaj još nije objavljen</h2>'
                    '<div class="w">' + sistem + ' · ' + _sel_lbl + '</div>' + _ph +
                    '<p>Do tada isplanirajte obilaske.</p></div>', unsafe_allow_html=True)
        return

    stavke = podaci["stavke"]
    meta = podaci.get("meta", {}) or {}
    _mes_kol = meta.get("meseci")
    _por_lbl = ("Porudžbina (" + str(_mes_kol).replace(".", ",") + " mes)") if _mes_kol else "Porudžbina"

    # --- Zaključavanje meseca: predato ručno ili istekao rok (rok postavljaju direktori) ---
    _predato = bool(meta.get("predato"))
    _rok = sb_rokovi_get(mesec_key).get("rok_admin") or meta.get("rok")  # 'YYYY-MM-DD' ako je postavljen
    _rok_prosao = False
    if _rok:
        try:
            _rok_prosao = datetime.date.today() > datetime.date.fromisoformat(str(_rok))
        except Exception:
            _rok_prosao = False
    _zakljucan = _predato or _rok_prosao
    po_obj = {}
    for s in stavke:
        po_obj.setdefault(int(s["idk"]), []).append(s)
    objekti = []
    for idk, lst in po_obj.items():
        nivo, n_nula, izgub = hitnost_objekta(lst)
        objekti.append({"idk": idk, "artikala": len(lst), "na_nuli": n_nula,
                        "izgub": izgub, "nivo": nivo, "lst": lst})
    objekti.sort(key=lambda r: (HIT_RANG[r["nivo"]], -r["izgub"]))
    ids_sorted = [o["idk"] for o in objekti]
    obj_by_id = {o["idk"]: o for o in objekti}

    obrada_map = sb_load_obrada(mesec_key, sistem)
    reviewed = set(idk for idk, v in obrada_map.items() if v.get("reakcije"))
    if st.session_state.get("_komfull") is None:
        st.session_state["_komfull"] = sb_komitenti_full()
    komfull = st.session_state.get("_komfull") or {}

    n_obj = len(objekti)
    n_red = sum(1 for o in objekti if o["nivo"] == "crveno")
    n_org = sum(1 for o in objekti if o["nivo"] == "zuto")
    n_grn = sum(1 for o in objekti if o["nivo"] == "zeleno")
    n_done = len([o for o in objekti if o["idk"] in reviewed])
    _pct = int(n_done / max(n_obj, 1) * 100)

    st.markdown(
        '<div class="adm-kpi">'
        '<div class="cell c-p"><div class="n"><span class="d p"></span>' + str(n_obj) + '</div><div class="k">Objekata za porudžbinu</div></div>'
        '<div class="cell c-r"><div class="n"><span class="d r"></span>' + str(n_red) + '</div><div class="k">Hitno pozvati</div></div>'
        '<div class="cell c-o"><div class="n"><span class="d o"></span>' + str(n_org) + '</div><div class="k">Iskontrolisati</div></div>'
        '<div class="cell c-g"><div class="n"><span class="d g"></span>' + str(n_grn) + '</div><div class="k">Dobra</div></div>'
        '</div>', unsafe_allow_html=True)
    st.markdown('<div class="adm-prog"><span class="t">Pregledano ' + str(n_done) + ' / ' + str(n_obj) + '</span>'
                '<div class="bar"><div style="width:' + str(_pct) + '%"></div></div></div>', unsafe_allow_html=True)

    # --- Predaja / zaključavanje izveštaja ---
    if _zakljucan:
        _pat = meta.get("predato_at")
        if _predato and _pat:
            _kada = " · predato " + str(_pat)
        elif _rok_prosao:
            _kada = " · rok istekao (" + str(_rok) + ")"
        else:
            _kada = ""
        st.markdown('<div style="background:#eef2ff;border:1px solid #c7d2fe;border-radius:10px;'
                    'padding:11px 14px;margin:4px 0 14px;color:#3730a3;font-size:13.5px;font-weight:600;">'
                    '🔒 Izveštaj za ' + _sel_lbl + ' · ' + sistem + ' je zaključan' + _kada +
                    '. Samo pregled — izmene, ažuriranje i slanje nisu mogući za ovaj mesec.</div>',
                    unsafe_allow_html=True)
        st.session_state.pop("_req_predaj", None)
    elif st.session_state.get("_req_predaj"):
        st.warning("Predajom zaključavaš " + _sel_lbl + " · " + sistem
                   + " — posle toga nema izmena, ažuriranja ni slanja u admin za taj mesec. Sigurno?")
        _pcf1, _pcf2, _pcf3 = st.columns([1, 1, 3])
        with _pcf1:
            if st.button("✅ Potvrdi predaju", key="predaj_ok", type="primary", use_container_width=True):
                try:
                    sb_predaj(mesec_key, sistem)
                    st.session_state.pop("_req_predaj", None)
                    st.rerun()
                except Exception as _e:
                    st.error("Greška pri predaji: " + str(_e))
        with _pcf2:
            if st.button("Otkaži", key="predaj_cancel", use_container_width=True):
                st.session_state.pop("_req_predaj", None)
                st.rerun()

    # --- Izvrši osvežavanje iz admina (traženo dugmetom "Ažuriraj" u zaglavlju) ---
    _bez_naziva = [o["idk"] for o in objekti if not ((komfull.get(int(o["idk"]), {}) or {}).get("naziv"))]
    if st.session_state.pop("_req_refresh_admin", False) and not _zakljucan:
        try:
            _cut_all = datetime.date(int(str(mesec_key).split("-")[0]), int(str(mesec_key).split("-")[1]), 1)
        except Exception:
            _cut_all = None
        with st.spinner("Povlačim sve iz admina za ceo sistem (nazivi + prethodne porudžbine + dopuna)... može potrajati par minuta."):
            if _bez_naziva:
                admin_build_komitenti(only_ids=_bez_naziva)
            _kf = sb_komitenti_full()
            st.session_state["_komfull"] = _kf
            _idk_naziv = {o["idk"]: (_kf.get(int(o["idk"]), {}) or {}).get("naziv", "") for o in objekti}
            _bulk, _be_all = admin_istorija_bulk(_idk_naziv, _cut_all)
        if _be_all and not _bulk:
            st.error(_be_all)
        else:
            _n_hist = 0
            for o in objekti:
                _lst = sorted(_bulk.get(o["idk"], []), key=_datum_sort_key, reverse=True)
                st.session_state["hist_" + str(sistem) + "_" + str(o["idk"])] = {"lst": _lst, "err": ""}
                if _lst:
                    _n_hist += 1
            st.session_state["_refresh_done"] = {"sis": sistem, "mes": mesec_key, "n": _n_hist}
            st.rerun()
    _rf = st.session_state.get("_refresh_done")
    if _rf and _rf.get("sis") == sistem and _rf.get("mes") == mesec_key:
        st.success("✅ Ažurirano iz admina — prethodne porudžbine i dopuna su spremni u svakom objektu. "
                   "Objekata sa porudžbinama posle 01.: " + str(_rf.get("n", 0)) + ".")

    with st.expander("📦 Porudžbina ubačena za ceo sistem (grupna akcija)"):
        st.caption("Za sisteme gde mi direktno ubacujemo porudžbine (npr. BB TRADE, KNEZ) — jednim klikom se svi objekti označe kao Ubačena porudžbina i postaju pregledani. Ne koristiti za sisteme gde se objekti zovu pojedinačno.")
        _bc1, _bc2 = st.columns(2)
        with _bc1:
            if st.button("✅ Označi sve: porudžbina ubačena", key="bulk_ubaci", use_container_width=True, disabled=_zakljucan):
                try:
                    sb_bulk_ubaci(mesec_key, sistem, ids_sorted)
                    st.success("Sve označeno kao ubačena porudžbina.")
                    st.rerun()
                except Exception as _e:
                    st.error("Greška: " + str(_e))
        with _bc2:
            if st.button("↩️ Poništi obradu celog sistema", key="bulk_reset", use_container_width=True, disabled=_zakljucan):
                try:
                    sb_bulk_reset(mesec_key, sistem)
                    st.success("Obrada sistema poništena.")
                    st.rerun()
                except Exception as _e:
                    st.error("Greška: " + str(_e))

    def _treb_list_cell(tip, rev):
        if tip == "nas":
            return '<span class="tb-nas">✓ po našem</span>'
        if tip == "njihov":
            return '<span class="tb-nj">po njihovom</span>'
        return '<span class="np">—</span>' if rev else '<span class="np">·</span>'

    tab_lista, tab_detalj = st.tabs(["Lista objekata", "Detalj / obrada"])

    with tab_lista:
        try:
            _cut_list = datetime.date(int(str(mesec_key).split("-")[0]), int(str(mesec_key).split("-")[1]), 1)
        except Exception:
            _cut_list = None
        _rows = ""
        for o in objekti:
            z = _zona_disp(o["nivo"])
            v = obrada_map.get(o["idk"], {})
            reak = v.get("reakcije", [])
            if reak:
                stat = "".join('<span class="stchip">' + _reak_short(r) + '</span>' for r in reak)
            else:
                stat = '<span class="stat">Nepregledano</span>'
            _rc = "row-red" if o["nivo"] == "crveno" else ("row-org" if o["nivo"] == "zuto" else "")
            # oznaka: da li je već trebovano posle 01. (samo za objekte čija je istorija učitana)
            _treb_mark = ""
            _hf = st.session_state.get("hist_" + str(sistem) + "_" + str(o["idk"]))
            if _hf and not _hf.get("err"):
                _tt = int(sum(_treb_posle_preseka(_hf.get("lst") or [], _cut_list).values()))
                if _tt > 0:
                    _treb_mark = ' <span title="Trebovano posle 01." style="color:#b45309;font-weight:700;">⚠️ posle 01. (' + str(_tt) + ')</span>'
            _nazlist = (komfull.get(int(o["idk"]), {}) or {}).get("naziv", "")
            _nazcell = ('<td>' + _h_escape(_nazlist) + _treb_mark + '</td>') if _nazlist else ('<td class="mut">— naknadno' + _treb_mark + '</td>')
            _rows += ('<tr class="' + _rc + '">'
                '<td class="idc">' + str(o["idk"]) + '</td>'
                + _nazcell +
                '<td class="ce" style="color:#d33;">' + str(o["na_nuli"]) + '</td>'
                '<td class="ce">' + str(o["izgub"]) + '</td>'
                '<td><span class="zona ' + z[0] + '"><span class="zd"></span>' + z[3] + '</span></td>'
                '<td>' + stat + '</td>'
                '<td class="ce">' + _treb_list_cell(v.get("trebovali_tip", ""), o["idk"] in reviewed) + '</td>'
                '</tr>')
        st.markdown('<table class="adm-t">'
            '<thead><tr><th>ID</th><th>Naziv komitenta</th><th>Na nuli</th><th>Izgubljeno</th>'
            '<th>Zona</th><th>Status</th><th style="text-align:center;">Trebovali</th></tr></thead>'
            '<tbody>' + _rows + '</tbody></table>', unsafe_allow_html=True)
        st.caption("Status i trebovanje se menjaju u kartici Detalj / obrada.")

    with tab_detalj:
        _labels = []
        for o in objekti:
            _zz = _zona_disp(o["nivo"])
            _nz = (komfull.get(int(o["idk"]), {}) or {}).get("naziv", "")
            _labels.append(_zz[2] + "  " + str(o["idk"]) + (("  ·  " + _nz) if _nz else "") + "  ·  " + _zz[3])
        _lab2id = {_labels[i]: ids_sorted[i] for i in range(len(objekti))}
        _id2lab = {ids_sorted[i]: _labels[i] for i in range(len(objekti))}
        if ("adm_pick" not in st.session_state) or (st.session_state.adm_pick not in _labels):
            st.session_state.adm_pick = _labels[0]

        def _next_unrev():
            _cur = _lab2id.get(st.session_state.adm_pick, ids_sorted[0])
            _idx = ids_sorted.index(_cur) if _cur in ids_sorted else -1
            for _k in range(1, len(ids_sorted) + 1):
                _cand = ids_sorted[(_idx + _k) % len(ids_sorted)]
                if _cand not in reviewed:
                    st.session_state.adm_pick = _id2lab[_cand]
                    return

        _nc1, _nc2 = st.columns([3, 1])
        with _nc1:
            st.selectbox("Izaberi / pretraži objekat (ukucaj ID ili naziv)", _labels, key="adm_pick")
        with _nc2:
            st.markdown("<div style='height:28px;'></div>", unsafe_allow_html=True)
            st.button("Sledeći nepregledan →", on_click=_next_unrev, use_container_width=True, key="adm_next")

        sel_id = _lab2id[st.session_state.adm_pick]
        o = obj_by_id[sel_id]
        z = _zona_disp(o["nivo"])
        v = obrada_map.get(sel_id, {"reakcije": [], "trebovali_tip": ""})
        TREB_OPT = ["— nije trebovano", "Po našem sistemu", "Po njihovom sistemu (ne po našem)"]
        TREB_CODE = {"— nije trebovano": "", "Po našem sistemu": "nas", "Po njihovom sistemu (ne po našem)": "njihov"}
        _key_treb = "treb_" + str(sel_id)
        _loaded_tip = v.get("trebovali_tip", "") or ""
        if _key_treb in st.session_state:
            _tip_now = TREB_CODE.get(st.session_state[_key_treb], "")
        else:
            _tip_now = _loaded_tip
        _njihov_active = (_tip_now == "njihov")
        _revb = '<span class="revy">✓ Pregledano</span>' if sel_id in reviewed else '<span class="revn">Nepregledano</span>'
        _kinfo = komfull.get(int(sel_id), {}) or {}
        _knaziv = _kinfo.get("naziv", "")
        _naz_html = ('<span style="font-weight:600;color:#2a2f3a;">' + _h_escape(_knaziv) + '</span>') if _knaziv else '<span class="mut">— naziv naknadno</span>'
        st.markdown('<div class="adm-dh"><span class="id">' + str(sel_id) + '</span>'
                    '<span class="zona ' + z[0] + '"><span class="zd"></span>' + z[3] + '</span>'
                    + _naz_html +
                    '<span style="margin-left:auto;">' + _revb + '</span></div>', unsafe_allow_html=True)
        _kbits = []
        if _kinfo.get("mesto"):
            _kbits.append("📍 " + _h_escape(_kinfo["mesto"]))
        if _kinfo.get("telefon"):
            _kbits.append("📞 " + _h_escape(_kinfo["telefon"]))
        if _kinfo.get("email"):
            _kbits.append("✉️ " + _h_escape(_kinfo["email"]))
        if _kbits:
            st.markdown('<div style="margin:-8px 0 12px;color:#6b7280;font-size:13px;">'
                        + "&nbsp;&nbsp;·&nbsp;&nbsp;".join(_kbits) + '</div>', unsafe_allow_html=True)

        # --- Presek (01. u mesecu izveštaja) + učitavanje istorije iz admina ---
        import datetime as _dtp
        try:
            _yy, _mm = str(mesec_key).split("-")[:2]
            _cutoff = _dtp.date(int(_yy), int(_mm), 1)
        except Exception:
            _cutoff = None
        _hk = "hist_" + str(sistem) + "_" + str(sel_id)
        _naziv_kom = _knaziv

        _dc1, _dc2 = st.columns([1.7, 1])
        with _dc1:
            st.markdown('<div class="adm-lbl">Porudžbina i lager · upiši Njihovu por.</div>', unsafe_allow_html=True)
            _arts = sorted(o["lst"], key=lambda x: (int(x["lager"]), -int(x["kol"])))
            _njm = v.get("njihova") or {}

            _hist_cache = st.session_state.get(_hk)
            _treb_loaded = bool(_hist_cache and not _hist_cache.get("err"))
            _treb_map = _treb_posle_preseka(_hist_cache.get("lst") or [], _cutoff) if _treb_loaded else {}
            _treb_total = int(sum(_treb_map.values()))

            def _sd(lg):
                lg = int(lg)
                return "🔴" if lg == 0 else ("🟡" if lg <= 2 else "🟢")

            _dodatna_map = {}
            _rows_adf = []
            for a in _arts:
                _ida = int(a["ida"]); _lg = int(a["lager"]); _kol = int(a["kol"])
                _por = int(_treb_map.get(_ida, 0))
                _realni = _lg + _por
                _dod = max(_kol - _por, 0)
                _dodatna_map[_ida] = _dod
                _rows_adf.append({
                    " ": _sd(_realni),
                    "Artikal": str(a["naziv"]),
                    "Predikcija": int(a.get("pred", 0)),
                    "Lager (izv.)": _lg,
                    "Posle 01.": _por,
                    "Realni lager": _realni,
                    "Naša por.": _kol,
                    "Dodatna por.": _dod,
                    "Njihova por.": int(_njm.get(str(_ida), 0)),
                })
            _adf = pd.DataFrame(_rows_adf)
            if _treb_loaded and _treb_total > 0:
                st.markdown('<div style="background:#fff4e5;border:1px solid #f0b429;border-radius:8px;'
                            'padding:8px 12px;margin:2px 0 10px;color:#8a5a00;font-size:13px;font-weight:600;">'
                            '⚠️ Već trebovano ' + str(_treb_total) + ' kom posle 01. — porudžbina je umanjena '
                            '(zelena kolona Dodatna por.).</div>', unsafe_allow_html=True)
            _colcfg = {
                " ": st.column_config.TextColumn(" ", width="small"),
                "Predikcija": st.column_config.NumberColumn("Predikcija (mesec)", help="Predviđena prodaja za mesec dana"),
                "Lager (izv.)": st.column_config.NumberColumn("Lager (izveštaj)", help="Lager iz izveštaja — presek na 01. u mesecu"),
                "Posle 01.": st.column_config.NumberColumn("Posle 01.", help="Koliko je već trebovano iz admina posle 01. u mesecu"),
                "Realni lager": st.column_config.NumberColumn("Realni lager", help="Lager (izveštaj) + poručeno posle 01."),
                "Naša por.": st.column_config.NumberColumn(_por_lbl, help="Preporučena porudžbina (za zadati broj meseci)"),
                "Dodatna por.": st.column_config.NumberColumn("Dodatna por.", help="Naša por. minus već poručeno posle 01. — ovo se šalje u admin"),
                "Njihova por.": st.column_config.NumberColumn("Njihova por.", help="Koliko su stvarno poručili", min_value=0, step=1),
            }
            if _njihov_active and not _zakljucan:
                _order = [" ", "Artikal", "Predikcija", "Lager (izv.)", "Posle 01.",
                          "Realni lager", "Naša por.", "Dodatna por.", "Njihova por."]
                _edited = st.data_editor(_adf[_order], hide_index=True, use_container_width=True,
                    disabled=[c for c in _order if c != "Njihova por."],
                    column_config=_colcfg, key="ed_" + str(sel_id))
                _njihova_new = {}
                for _i, _a in enumerate(_arts):
                    try:
                        _njihova_new[str(int(_a["ida"]))] = int(_edited.iloc[_i]["Njihova por."])
                    except Exception:
                        _njihova_new[str(int(_a["ida"]))] = 0
            else:
                _order = [" ", "Artikal", "Predikcija", "Lager (izv.)", "Posle 01.",
                          "Realni lager", "Naša por.", "Dodatna por."]
                if _njihov_active:  # zaključan njihov — prikaži i njihovu kolonu (samo pregled)
                    _order.append("Njihova por.")
                _sty = _adf[_order].style.set_properties(subset=["Dodatna por."], **{
                    "background-color": "#dcfce7", "color": "#14532d", "font-weight": "700"})
                st.dataframe(_sty, hide_index=True, use_container_width=True, column_config=_colcfg)
                _njihova_new = {str(int(a["ida"])): int(_njm.get(str(int(a["ida"])), 0)) for a in _arts}

            # Naše količine = preporučena porudžbina umanjena za već trebovano posle 01. (min 0)
            _nase_rows = [(sel_id, int(a["ida"]), int(_dodatna_map.get(int(a["ida"]), 0)))
                          for a in _arts if int(_dodatna_map.get(int(a["ida"]), 0)) > 0]
            _njih_rows = [(sel_id, int(a["ida"]), int(_njihova_new.get(str(int(a["ida"])), 0)))
                          for a in _arts if int(_njihova_new.get(str(int(a["ida"])), 0)) > 0]
            st.markdown('<div class="adm-lbl" style="margin-top:16px;">Ubaci u admin</div>', unsafe_allow_html=True)
            if _treb_loaded:
                if _treb_total > 0:
                    st.caption("✅ Posle 01. je već trebovano " + str(_treb_total)
                               + " kom — Naše količine se šalju umanjeno (kolona Dodatna por.).")
                else:
                    st.caption("✅ Nema trebovanja posle 01. — šalju se pune količine.")
            elif _naziv_kom:
                st.caption("ℹ️ Klikni Ažuriraj iz admina (gore) da se količine umanje za već poručeno posle 01.")
            else:
                st.caption("Za proveru trebovanja posle 01. učitaj šifarnik komitenata (potreban je naziv).")
            _nase_items = [{"idArticle": _a, "quantity": _q} for (_k, _a, _q) in _nase_rows]
            _njih_items = [{"idArticle": _a, "quantity": _q} for (_k, _a, _q) in _njih_rows]

            def _push_admin(_tag, _items):
                _sk = "admsent_" + str(sistem) + "_" + str(sel_id) + "_" + _tag
                _prev = st.session_state.get(_sk)
                if _prev and _prev.get("ok"):
                    return  # već uspešno kreirano — ne šaljemo ponovo (bez duplih porudžbina)
                with st.spinner("Šaljem u admin..."):
                    _ok, _msg = posalji_u_admin(sel_id, _items)
                st.session_state[_sk] = {"ok": _ok, "msg": _msg}

            _xa, _xb, _xsp = st.columns([1.3, 1.3, 1])
            with _xa:
                _clk_nas = st.button("📦 Naše → admin", key="axn_" + str(sel_id),
                                     use_container_width=True,
                                     disabled=(len(_nase_items) == 0 or _zakljucan))
            with _xb:
                _clk_nj = st.button("📦 Njihove → admin", key="axj_" + str(sel_id),
                                    use_container_width=True,
                                    disabled=(len(_njih_items) == 0 or _zakljucan))
            if _clk_nas:
                _push_admin("nas", _nase_items)
            if _clk_nj:
                _push_admin("njihov", _njih_items)
            for _tag, _lbl in (("nas", "Naše"), ("njihov", "Njihove")):
                _sk = "admsent_" + str(sistem) + "_" + str(sel_id) + "_" + _tag
                _res = st.session_state.get(_sk)
                if _res:
                    (st.success if _res["ok"] else st.error)(_lbl + " → admin: " + _res["msg"])
                    if st.button("↺ Pošalji ponovo (" + _lbl + ")", key="axr_" + _tag + "_" + str(sel_id)):
                        st.session_state.pop(_sk, None)
                        st.rerun()

            st.markdown('<div class="adm-lbl" style="margin-top:18px;">🧾 Prethodne porudžbine (admin · ~3 meseca)</div>', unsafe_allow_html=True)
            if not _naziv_kom:
                st.caption("Nema naziva za ovaj objekat — učitaj šifarnik komitenata u Objavi izveštaja (ili poveži iz admina).")
                if st.button("🔗 Poveži šifre komitenata iz admina", key="bk_" + str(sel_id)):
                    with st.spinner("Čitam listu komitenata iz admina..."):
                        _bn, _be = admin_build_komitenti()
                    if _bn == 0:
                        st.error(_be or "Nije uspelo.")
                    else:
                        st.session_state["_komfull"] = sb_komitenti_full()
                        st.success("Povezano " + str(_bn) + " komitenata.")
                        st.rerun()
            else:
                _hist = st.session_state.get(_hk)
                if _hist is None:
                    st.caption("Klikni Ažuriraj podatke iz admina (gore) — prethodne porudžbine se prikažu automatski.")
                elif _hist.get("err"):
                    st.error(_hist["err"])
                elif not _hist.get("lst"):
                    st.caption("Nema porudžbina za ovaj objekat posle 01. u mesecu.")
                else:
                    st.caption("Pronađeno " + str(len(_hist["lst"])) + " porudžbina — klikni na datum za sadržaj.")
                    for _o in _hist["lst"]:
                        with st.expander("📅 " + (_o["datum"] or "?") + "   ·   " + (_o["status"] or "") + "   ·   " + (_o["cena"] or "")):
                            if _o.get("stavke"):
                                import pandas as _pdh
                                _hd = _pdh.DataFrame([{"Artikal": _s["naziv"], "Kol.": _s["kol"], "Cena": _s["cena"]}
                                                      for _s in _o["stavke"]])
                                st.dataframe(_hd, hide_index=True, use_container_width=True)
                            else:
                                st.caption("Nema stavki (ili nisu učitane).")
        with _dc2:
            if "Ubačena porudžbina" in (v.get("reakcije") or []):
                st.info("📦 Porudžbina je ubačena za ceo sistem (grupno).")
            st.markdown('<div class="adm-lbl">Reakcija</div>', unsafe_allow_html=True)
            _loaded = list(v.get("reakcije", []))
            _r1 = st.checkbox("Pozvala sam", value=("Pozvala sam" in _loaded), key="r1_" + str(sel_id))
            _r2 = st.checkbox("Poslala sam mejl", value=("Poslala sam mejl" in _loaded), key="r2_" + str(sel_id))
            _r3 = st.checkbox("Obavestila direktorku", value=("Obavestila direktorku" in _loaded), key="r3_" + str(sel_id))
            react = []
            if _r1: react.append("Pozvala sam")
            if _r2: react.append("Poslala sam mejl")
            if _r3: react.append("Obavestila direktorku")
            _can = len(react) > 0
            st.markdown('<div class="adm-lbl" style="margin-top:12px;">Trebovanje nakon reakcije</div>', unsafe_allow_html=True)
            _tip_idx = ["", "nas", "njihov"].index(_loaded_tip) if _loaded_tip in ["", "nas", "njihov"] else 0
            _treb_lbl = st.radio("Trebovanje", TREB_OPT, index=_tip_idx, disabled=not _can,
                                 key=_key_treb, label_visibility="collapsed")
            _tip_val = TREB_CODE.get(_treb_lbl, "") if _can else ""
            if not _can:
                st.caption("🔒 Otključava se kad izabereš bar jednu reakciju.")
            st.markdown('<div class="adm-lbl" style="margin-top:12px;">Napomena (interno)</div>', unsafe_allow_html=True)
            _nap = st.text_area("Napomena", value=(v.get("napomena", "") or ""), key="nap_" + str(sel_id),
                                height=72, label_visibility="collapsed",
                                placeholder="npr. zvati posle 15h, tražiti vlasnika...")
            if st.button("💾 Sačuvaj status", key="savest_" + str(sel_id), type="primary", disabled=_zakljucan):
                if not _can:
                    st.error("Izaberi bar jednu reakciju — ne može da se sačuva samo napomena.")
                elif ("Obavestila direktorku" in react) and not (_nap or "").strip():
                    st.error("Za prosleđivanje direktorki upiši napomenu — zašto prosleđuješ (obavezno).")
                elif _tip_val == "njihov" and sum(int(x) for x in _njihova_new.values()) == 0:
                    st.error("Za opciju Po njihovom sistemu upiši koliko su poručili (Njihova por.) pre čuvanja.")
                else:
                    try:
                        sb_save_obrada(mesec_key, sistem, sel_id, react, _tip_val, _njihova_new, _nap)
                        st.success("Sačuvano ✓")
                        st.rerun()
                    except Exception as _e:
                        st.error("Greška: " + str(_e))


def _direktor_blok_iz_prodaje(sistem, sales):
    """Napravi direktorski blok (prodaja_trend, poređenja, po grupama) iz tabele prodaje
    (Izveštaj prodaje), tako da ne treba ponovna objava sistema. Vrati dict ili None."""
    if not sales:
        return None
    po = sales.get("po_sistemu") or {}
    _sn = str(sistem).strip().upper()
    _key = None
    for k in po.keys():
        if str(k).strip().upper() == _sn:
            _key = k
            break
    if _key is None:
        return None
    nazivi = sales.get("nazivi") or []
    total = po[_key].get("total") or []
    grupe = po[_key].get("grupe") or {}
    _n = min(len(nazivi), len(total))
    trend = [{"mesec": nazivi[i], "kom": int(total[i])} for i in range(_n)]
    if not trend:
        return None
    d = {"prodaja_trend": trend, "prodaja_tekuci": trend[-1]}
    comp = {}
    if len(trend) >= 2:
        comp["prosli_mesec"] = trend[-2]
        _prev6 = trend[-7:-1] if len(trend) >= 7 else trend[:-1]
        if _prev6:
            comp["prosek_6m"] = {"kom": int(round(sum(t["kom"] for t in _prev6) / len(_prev6)))}
    _last = str(trend[-1]["mesec"]).split()
    if len(_last) == 2 and _last[1].isdigit():
        _lani = _last[0] + " " + str(int(_last[1]) - 1)
        for t in trend:
            if t["mesec"] == _lani:
                comp["isti_mesec_lani"] = t
                break
    d["poredjenja"] = comp
    _pg = []
    for grp, vals in grupe.items():
        if vals:
            _pg.append({"grupa": grp, "kom": int(vals[-1])})
    _pg.sort(key=lambda x: -x["kom"])
    d["po_grupama"] = _pg
    # Mesečne vrednosti po grupi (za složeni stub-grafikon) + nazivi meseca
    d["nazivi"] = [nazivi[i] for i in range(_n)]
    _gm = {}
    for grp, vals in grupe.items():
        vv = vals or []
        _gm[str(grp)] = [int(vv[i]) if i < len(vv) else 0 for i in range(_n)]
    d["grupe_mesecno"] = _gm
    return d


def prikazi_direktore():
    st.set_page_config(page_title="VAPE — Direktori", page_icon="📈",
                       layout="wide", initial_sidebar_state="collapsed")
    st.markdown("""<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');
    section[data-testid="stSidebar"]{display:none !important;}
    header[data-testid="stHeader"]{display:none !important;}
    .stApp{background:#f4f1fb !important;font-family:'Inter',sans-serif;}
    .block-container{max-width:1180px !important;padding-top:26px !important;}
    [class*="st-key-dir_odjava"] button{font-size:11px !important;padding:4px 9px !important;border-radius:7px !important;
        min-height:0 !important;background:#fff !important;border:1px solid #e5e7eb !important;color:#6b7280 !important;font-weight:600 !important;}
    </style>""", unsafe_allow_html=True)

    def _fmt(n):
        try:
            return f"{int(round(float(n))):,}".replace(",", ".")
        except Exception:
            return str(n)

    _h1, _h2 = st.columns([7.5, 1.15])
    with _h1:
        st.markdown('<div style="display:flex;align-items:center;gap:12px;padding:4px 0 2px;">'
                    '<div style="width:38px;height:38px;border-radius:10px;background:linear-gradient(135deg,#a855f7,#ec4899);"></div>'
                    '<div><span style="font-size:19px;font-weight:800;color:#1f2430;">VAPE Porudžbine</span>'
                    '<span style="font-size:13px;color:#8b8fa0;margin-left:6px;">· Direktori</span></div></div>',
                    unsafe_allow_html=True)
    with _h2:
        st.markdown("<div style='height:8px;'></div>", unsafe_allow_html=True)
        if st.button("Odjava", key="dir_odjava", use_container_width=True):
            for _k in ("authenticated", "role"):
                st.session_state.pop(_k, None)
            st.rerun()

    if not sb_dostupan():
        st.error("Veza sa bazom nije podešena. Javi se analitičaru.")
        return

    _pub = sb_meseci()
    _mk = {m["key"] for m in _pub}
    # Uključi i tekući mesec (ako je počeo) + mesece za koje su postavljeni rokovi,
    # da direktor može da izabere i pre objave (i vidi poruku o roku).
    _tdy0 = datetime.date.today()
    _mk.add(str(_tdy0.year) + "-" + ("0" + str(_tdy0.month))[-2:])
    try:
        for _rmk in sb_rokovi_all().keys():
            _mk.add(_rmk)
    except Exception:
        pass
    _mes_keys = sorted(_mk, reverse=True)
    if not _mes_keys:
        st.info("Još nema objavljenih izveštaja.")
        return

    _mlbls = [mesec_label(k) for k in _mes_keys]
    # Podrazumevani mesec za prikaze = prethodni mesec (u avgustu se gleda jul); ako ga nema, najnoviji.
    _pv_y0 = _tdy0.year; _pv_m0 = _tdy0.month - 1
    if _pv_m0 <= 0:
        _pv_m0 += 12; _pv_y0 -= 1
    _prev_mk = str(_pv_y0) + "-" + ("0" + str(_pv_m0))[-2:]
    _prev_idx = _mes_keys.index(_prev_mk) if _prev_mk in _mes_keys else 0
    _view = st.session_state.get("dir_view", "dash")

    # ---- Render izveštaja po sistemu (isti prikaz za pune i delimične podatke) ----
    def _render_sistem_report(dd, puno):
        if puno:
            tek = dd.get("prodaja_tekuci", {}) or {}
            tek_kom = int(tek.get("kom", 0))
            comp = dd.get("poredjenja", {}) or {}

            def _delta_html(base):
                if base and base.get("kom"):
                    p = (tek_kom - int(base["kom"])) / int(base["kom"]) * 100.0
                    arrow = "▲" if p >= 0 else "▼"
                    col = "#16a34a" if p >= 0 else "#dc2626"
                    return ('<span style="color:' + col + ';">' + arrow + " "
                            + (("%.1f" % abs(p)).replace(".", ",")) + "%</span>")
                return '<span style="color:#c4c7cf;">— nema podataka</span>'

            _kpi = '<div style="display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin-bottom:22px;">'
            for (v, k) in [
                (_fmt(tek_kom) + " <span style='font-size:13px;color:#9aa0ad;'>kom</span>", "Prodaja — " + str(tek.get("mesec", ""))),
                (_delta_html(comp.get("prosli_mesec")), "vs prošli mesec"),
                (_delta_html(comp.get("isti_mesec_lani")), "vs isti mesec lani"),
                (_delta_html(comp.get("prosek_6m")), "vs 6-mesečni prosek"),
            ]:
                _kpi += ('<div style="background:#fff;border:1px solid #efeaf7;border-radius:15px;padding:16px 18px;'
                         'box-shadow:0 2px 12px rgba(80,40,140,.06);">'
                         '<div style="font-size:22px;font-weight:800;color:#1f2430;">' + v + '</div>'
                         '<div style="font-size:11.5px;color:#8b8fa0;margin-top:5px;font-weight:600;">' + k + '</div></div>')
            _kpi += '</div>'
            st.markdown(_kpi, unsafe_allow_html=True)

            # 2. Kombinovani grafikon: prodaja po mesecima i grupama (složeni stub)
            _naz = dd.get("nazivi") or [str(t.get("mesec", "")) for t in dd.get("prodaja_trend", [])]
            _gm = dd.get("grupe_mesecno") or {}
            if _naz and _gm:
                _render_stacked_chart(_naz, _gm)
            else:
                trend = dd.get("prodaja_trend", [])
                _maxk = max([t["kom"] for t in trend] or [1]) or 1
                _tr = ('<div style="background:#fff;border:1px solid #efeaf7;border-radius:16px;padding:20px 22px;margin-bottom:20px;'
                       'box-shadow:0 2px 12px rgba(80,40,140,.05);">'
                       '<div style="font-size:15px;font-weight:800;margin-bottom:16px;">Prodaja — trend po mesecima (kom)</div>')
                for t in trend:
                    _w = int(int(t["kom"]) / _maxk * 100)
                    _tr += ('<div style="display:grid;grid-template-columns:96px 1fr 96px;align-items:center;gap:12px;margin-bottom:9px;">'
                            '<span style="font-size:12.5px;color:#4b5563;font-weight:600;">' + str(t["mesec"]) + '</span>'
                            '<div style="height:12px;background:#f0ebf9;border-radius:20px;overflow:hidden;">'
                            '<div style="height:100%;width:' + str(_w) + '%;background:linear-gradient(90deg,#a855f7,#ec4899);border-radius:20px;"></div></div>'
                            '<span style="font-size:12.5px;text-align:right;font-weight:700;color:#4b5563;">' + _fmt(t["kom"]) + '</span></div>')
                _tr += '</div>'
                st.markdown(_tr, unsafe_allow_html=True)

        # Predlog porudžbine po grupama — samo za parcijalni prikaz (kad nema mesečnih grupa)
        grupe = dd.get("po_grupama", [])
        if grupe and not (dd.get("nazivi") and dd.get("grupe_mesecno")):
            _gt = "Prodaja po grupama" if puno else "Predlog porudžbine po grupama"
            _gs = ("Udeo u ukupnoj prodaji sistema (tekući mesec)." if puno
                   else "Udeo u predloženoj porudžbini po grupama (tekući mesec).")
            _gtot = sum(int(g["kom"]) for g in grupe) or 1
            _gmax = max(int(g["kom"]) for g in grupe) or 1
            _gh = ('<div style="background:#fff;border:1px solid #efeaf7;border-radius:16px;padding:20px 22px;margin-bottom:20px;'
                   'box-shadow:0 2px 12px rgba(80,40,140,.05);">'
                   '<div style="font-size:15px;font-weight:800;margin-bottom:4px;">' + _gt + '</div>'
                   '<div style="font-size:12.5px;color:#9aa0ad;margin-bottom:16px;">' + _gs + '</div>')
            for g in grupe:
                _kom = int(g["kom"]); _udeo = int(round(_kom / _gtot * 100)); _w = int(_kom / _gmax * 100)
                _gh += ('<div style="display:grid;grid-template-columns:150px 1fr 120px;align-items:center;gap:12px;margin-bottom:12px;">'
                        '<span style="font-size:13px;font-weight:600;color:#374151;">' + _h_escape(str(g["grupa"])) + '</span>'
                        '<div style="height:11px;background:#f0ebf9;border-radius:20px;overflow:hidden;">'
                        '<div style="height:100%;width:' + str(_w) + '%;background:linear-gradient(90deg,#7c3aed,#c084fc);border-radius:20px;"></div></div>'
                        '<span style="font-size:12.5px;text-align:right;font-weight:700;color:#4b5563;">' + _fmt(_kom) + ' · ' + str(_udeo) + '%</span></div>')
            _gh += '</div>'
            st.markdown(_gh, unsafe_allow_html=True)

        # 3. Prosečna prodaja po objektu (line)
        _ppo = dd.get("prosek_po_objektu") or []
        if _ppo:
            _render_prosek_line(_ppo)

        # 3b. Predlog porudžbine za sistem + pokrivenost
        _por = dd.get("porudzbina") or {}
        if _por and (_por.get("ukupno") or _por.get("po_grupi")):
            _render_porudzbina(_por)

        # 3c. Bestseleri i najslabiji artikli
        _arg = dd.get("artikli_rang") or {}
        if _arg and (_arg.get("best") or _arg.get("slab")):
            _render_artikli_rang(_arg)

        # 3d. Uspešnost akcije
        _ak = dd.get("akcija") or {}
        if _ak and _ak.get("artikli"):
            _render_akcija(_ak)

        # 4. Out of stock — po količinama (poslednji mesec)
        _okd = dd.get("oos_kom") or {}
        if _okd and (_okd.get("po_artiklu") or _okd.get("izgubljeno_kom") or _okd.get("objekata_na_0")):
            _render_oos_kom(_okd)
        else:
            oos = dd.get("oos", {}) or {}
            if oos:
                _oh = ('<div style="background:#fff;border:1px solid #efeaf7;border-radius:16px;padding:20px 22px;margin-bottom:14px;'
                       'box-shadow:0 2px 12px rgba(80,40,140,.05);">'
                       '<div style="font-size:15px;font-weight:800;margin-bottom:4px;">Out of stock — izgubljena prodaja</div>'
                       '<div style="font-size:12.5px;color:#9aa0ad;margin-bottom:16px;">Na osnovu dnevnog lagera: koliko komada je izgubljeno jer je artikal bio na nuli.</div>'
                       '<div style="display:grid;grid-template-columns:repeat(2,1fr);gap:14px;">'
                       '<div style="background:#fef2f2;border:1px solid #fecaca;border-radius:14px;padding:16px;text-align:center;">'
                       '<div style="font-size:24px;font-weight:800;color:#dc2626;">~' + _fmt(oos.get("izgubljeno_kom", 0)) + '</div>'
                       '<div style="font-size:11.5px;color:#9b6b6b;margin-top:3px;font-weight:600;">Izgubljeno (kom)</div></div>'
                       '<div style="background:#fef2f2;border:1px solid #fecaca;border-radius:14px;padding:16px;text-align:center;">'
                       '<div style="font-size:24px;font-weight:800;color:#dc2626;">' + _fmt(oos.get("kombinacija_na_0", 0)) + '</div>'
                       '<div style="font-size:11.5px;color:#9b6b6b;margin-top:3px;font-weight:600;">Kombinacija objekat × artikal na 0</div></div>'
                       '</div></div>')
                st.markdown(_oh, unsafe_allow_html=True)
                _pa = dd.get("oos_po_artiklu", [])
                if _pa:
                    _df = pd.DataFrame([{"Artikal": r["artikal"], "U koliko objekata na 0": r["objekata"],
                                         "Izgubljeno (kom)": r["izgubljeno"]} for r in _pa])
                    st.dataframe(_df, hide_index=True, use_container_width=True)

        # 5. Profitabilnost (identično kao u analitici)
        _pf_blok = dd.get("profit") or {}
        if _pf_blok and (_pf_blok.get("total_bruto") is not None):
            _render_profit_blok(_pf_blok)

    def _render_stacked_chart(nazivi, grupe_mesecno):
        # boje po grupi (stabilan raspored)
        _pal = ["#7c3aed", "#ec4899", "#f59e0b", "#0ea5e9", "#10b981", "#a855f7", "#f43f5e", "#14b8a6"]
        _keys = sorted(grupe_mesecno.keys(), key=lambda k: -sum(grupe_mesecno[k]))
        _boje = {k: _pal[i % len(_pal)] for i, k in enumerate(_keys)}
        _n = len(nazivi)
        _tot = [sum(int(grupe_mesecno[k][i]) if i < len(grupe_mesecno[k]) else 0 for k in _keys) for i in range(_n)]
        _maxt = max(_tot or [1]) or 1
        _cols = ""
        for i in range(_n):
            _hpx = int(_tot[i] / _maxt * 210)
            _segs = ""
            for k in _keys:
                _v = int(grupe_mesecno[k][i]) if i < len(grupe_mesecno[k]) else 0
                if _v > 0 and _tot[i] > 0:
                    _segs = ('<div style="width:100%;height:' + ("%.2f" % (_v / _tot[i] * 100)) + '%;background:' + _boje[k] + ';"></div>') + _segs
            _cols += ('<div style="flex:1;display:flex;flex-direction:column;justify-content:flex-end;align-items:center;height:100%;min-width:0;">'
                      '<div title="' + _h_escape(str(nazivi[i])) + ': ' + _fmt(_tot[i]) + ' kom" '
                      'style="width:72%;height:' + str(_hpx) + 'px;border-radius:4px 4px 0 0;overflow:hidden;display:flex;flex-direction:column;justify-content:flex-end;box-shadow:0 1px 3px rgba(0,0,0,.06);">'
                      + _segs + '</div>'
                      '<div style="font-size:9px;color:#9aa0ad;margin-top:6px;font-weight:600;white-space:nowrap;transform:rotate(-30deg);transform-origin:center;">' + _h_escape(str(nazivi[i])) + '</div></div>')
        _leg = ""
        for k in _keys:
            _leg += ('<div style="display:flex;align-items:center;gap:7px;font-size:12.5px;color:#374151;font-weight:600;">'
                     '<span style="width:13px;height:13px;border-radius:4px;background:' + _boje[k] + ';"></span>' + _h_escape(str(k)) + '</div>')
        _html = ('<div style="background:#fff;border:1px solid #efeaf7;border-radius:16px;padding:20px 22px 14px;margin-bottom:20px;'
                 'box-shadow:0 2px 12px rgba(80,40,140,.05);">'
                 '<div style="font-size:15px;font-weight:800;margin-bottom:4px;">Prodaja po mesecima i grupama (kom)</div>'
                 '<div style="font-size:12.5px;color:#9aa0ad;margin-bottom:16px;">Visina stuba je ukupna prodaja meseca; boje pokazuju koliko je koja grupa prodala.</div>'
                 '<div style="display:flex;align-items:flex-end;gap:5px;height:250px;padding:8px 2px 0;">' + _cols + '</div>'
                 '<div style="display:flex;gap:18px;flex-wrap:wrap;margin-top:20px;">' + _leg + '</div></div>')
        st.markdown(_html, unsafe_allow_html=True)

    def _render_prosek_line(ppo):
        _lm = [str(p.get("mesec", "")) for p in ppo]
        _lv = [float(p.get("prosek", 0) or 0) for p in ppo]
        _n = len(_lv)
        if _n == 0:
            return
        W = 1080; H = 240; pl = 46; pr = 20; pt = 24; pb = 40
        pw = W - pl - pr; ph = H - pt - pb
        _mx = (max(_lv) or 1) * 1.15

        def _px(i):
            return pl + (i / (_n - 1) * pw if _n > 1 else pw / 2)

        def _py(v):
            return pt + ph - (v / _mx * ph if _mx else 0)

        _g = ""
        for f in (0, 0.25, 0.5, 0.75, 1):
            _y = pt + ph - f * ph
            _g += ('<line x1="' + str(pl) + '" y1="' + ("%.1f" % _y) + '" x2="' + str(W - pr) + '" y2="' + ("%.1f" % _y) + '" stroke="#f0ebf9"/>'
                   '<text x="' + str(pl - 8) + '" y="' + ("%.1f" % (_y + 4)) + '" font-size="10" fill="#b9bdc9" text-anchor="end">' + ("%.1f" % (_mx * f)).replace(".", ",") + '</text>')
        _pts = " ".join(("%.1f,%.1f" % (_px(i), _py(_lv[i]))) for i in range(_n))
        _g += '<polygon points="' + str(pl) + "," + ("%.1f" % (pt + ph)) + " " + _pts + " " + str(W - pr) + "," + ("%.1f" % (pt + ph)) + '" fill="#a855f7" fill-opacity="0.08"/>'
        _g += '<polyline points="' + _pts + '" fill="none" stroke="#7c3aed" stroke-width="2.5"/>'
        for i in range(_n):
            _x = _px(i); _y = _py(_lv[i])
            _g += ('<circle cx="' + ("%.1f" % _x) + '" cy="' + ("%.1f" % _y) + '" r="5" fill="#7c3aed" stroke="#fff" stroke-width="2"/>'
                   '<text x="' + ("%.1f" % _x) + '" y="' + ("%.1f" % (_y - 12)) + '" font-size="11" font-weight="700" fill="#6d28d9" text-anchor="middle">' + ("%.1f" % _lv[i]).replace(".", ",") + '</text>'
                   '<text x="' + ("%.1f" % _x) + '" y="' + str(H - 10) + '" font-size="10" fill="#9aa0ad" text-anchor="middle">' + _h_escape(_lm[i]) + '</text>')
        _svg = '<svg viewBox="0 0 ' + str(W) + ' ' + str(H) + '" style="width:100%;height:240px;">' + _g + '</svg>'
        _html = ('<div style="background:#fff;border:1px solid #efeaf7;border-radius:16px;padding:20px 22px;margin-bottom:20px;'
                 'box-shadow:0 2px 12px rgba(80,40,140,.05);">'
                 '<div style="font-size:15px;font-weight:800;margin-bottom:4px;">Prosečna prodaja po objektu (kom / objektu)</div>'
                 '<div style="font-size:12.5px;color:#9aa0ad;margin-bottom:10px;">Ukupna prodaja meseca podeljena brojem aktivnih objekata tog meseca.</div>'
                 + _svg + '</div>')
        st.markdown(_html, unsafe_allow_html=True)

    def _render_oos_kom(ok):
        _mes = str(ok.get("mesec", ""))
        _oh = ('<div style="background:#fff;border:1px solid #efeaf7;border-radius:16px;padding:20px 22px;margin-bottom:14px;'
               'box-shadow:0 2px 12px rgba(80,40,140,.05);">'
               '<div style="font-size:15px;font-weight:800;margin-bottom:4px;">Out of stock — po količinama · ' + _h_escape(_mes) + '</div>'
               '<div style="font-size:12.5px;color:#9aa0ad;margin-bottom:16px;">Koliko je komada izgubljeno u prethodnom mesecu jer je artikal bio na nuli.</div>'
               '<div style="display:grid;grid-template-columns:repeat(2,1fr);gap:14px;">'
               '<div style="background:#fef2f2;border:1px solid #fecaca;border-radius:14px;padding:16px;text-align:center;">'
               '<div style="font-size:24px;font-weight:800;color:#dc2626;">~' + _fmt(ok.get("izgubljeno_kom", 0)) + '</div>'
               '<div style="font-size:11.5px;color:#9b6b6b;margin-top:3px;font-weight:600;">Izgubljeno (kom) · ' + _h_escape(_mes) + '</div></div>'
               '<div style="background:#eff6ff;border:1px solid #bfdbfe;border-radius:14px;padding:16px;text-align:center;">'
               '<div style="font-size:24px;font-weight:800;color:#2563eb;">' + _fmt(ok.get("objekata_na_0", 0)) + '</div>'
               '<div style="font-size:11.5px;color:#5b7bb0;margin-top:3px;font-weight:600;">U koliko objekata je lager na 0</div></div>'
               '</div></div>')
        st.markdown(_oh, unsafe_allow_html=True)
        _pa = ok.get("po_artiklu", [])
        if _pa:
            _df = pd.DataFrame([{"Artikal": r["artikal"], "U koliko objekata na 0": r["objekata"],
                                 "Izgubljeno (kom) · " + _mes: r["izgubljeno"]} for r in _pa])
            st.dataframe(_df, hide_index=True, use_container_width=True, height=min(60 + 35 * len(_pa), 520))

    def _render_porudzbina(pr):
        _uk = int(pr.get("ukupno", 0)); _ob = int(pr.get("objekata", 0)); _dani = pr.get("dani_avg")
        _h = ('<div style="background:#fff;border:1px solid #efeaf7;border-radius:16px;padding:20px 22px;margin-bottom:20px;'
              'box-shadow:0 2px 12px rgba(80,40,140,.05);">'
              '<div style="font-size:15px;font-weight:800;margin-bottom:4px;">Predlog porudžbine za sistem</div>'
              '<div style="font-size:12.5px;color:#9aa0ad;margin-bottom:16px;">Koliko sistem treba da poruči (preporuka analitike) i za koliko dana lager traje.</div>'
              '<div style="display:grid;grid-template-columns:repeat(3,1fr);gap:14px;margin-bottom:' + ('16px' if pr.get("po_grupi") else '0') + ';">')
        _cards = [("#7c3aed", _fmt(_uk) + " <span style='font-size:13px;color:#9aa0ad;'>kom</span>", "Ukupno za poručivanje"),
                  ("#0ea5e9", _fmt(_ob), "Objekata poručuje")]
        if _dani is not None:
            _cards.append(("#f59e0b", _fmt(_dani) + " <span style='font-size:13px;color:#9aa0ad;'>dana</span>", "Prosečna pokrivenost lagera"))
        for (col, v, lab) in _cards:
            _h += ('<div style="background:#faf9fd;border:1px solid #efeaf7;border-left:4px solid ' + col + ';border-radius:14px;padding:15px 17px;">'
                   '<div style="font-size:20px;font-weight:800;color:' + col + ';">' + v + '</div>'
                   '<div style="font-size:11px;color:#9aa0ad;margin-top:4px;font-weight:600;">' + lab + '</div></div>')
        _h += '</div>'
        _pg = pr.get("po_grupi", [])
        if _pg:
            _gmax = max([int(g["kom"]) for g in _pg] or [1]) or 1
            _h += '<div style="font-size:12.5px;font-weight:700;color:#374151;margin:4px 0 10px;">Po grupama:</div>'
            for g in _pg:
                _k = int(g["kom"]); _w = int(_k / _gmax * 100)
                _h += ('<div style="display:grid;grid-template-columns:150px 1fr 90px;align-items:center;gap:12px;margin-bottom:9px;">'
                       '<span style="font-size:13px;font-weight:600;color:#374151;">' + _h_escape(str(g["grupa"])) + '</span>'
                       '<div style="height:11px;background:#f0ebf9;border-radius:20px;overflow:hidden;">'
                       '<div style="height:100%;width:' + str(_w) + '%;background:linear-gradient(90deg,#7c3aed,#c084fc);border-radius:20px;"></div></div>'
                       '<span style="font-size:12.5px;text-align:right;font-weight:700;color:#4b5563;">' + _fmt(_k) + ' kom</span></div>')
        _h += '</div>'
        st.markdown(_h, unsafe_allow_html=True)

    def _render_artikli_rang(ar):
        _best = ar.get("best", []); _slab = ar.get("slab", [])
        _c1, _c2 = st.columns(2)
        with _c1:
            _h = _card_open("🔝 Bestseleri", "Najprodavaniji artikli u sistemu (ceo period).")
            _mx = max([int(x["prodato"]) for x in _best] or [1]) or 1
            for x in _best:
                _p = int(x["prodato"]); _w = int(_p / _mx * 100)
                _h += ('<div style="margin-bottom:10px;">'
                       '<div style="display:flex;justify-content:space-between;font-size:12.5px;color:#374151;margin-bottom:3px;">'
                       '<span style="font-weight:600;">' + _h_escape(str(x["artikal"])[:38]) + '</span>'
                       '<span style="font-weight:700;color:#16a34a;">' + _fmt(_p) + ' kom</span></div>'
                       '<div style="height:9px;background:#f0fdf4;border-radius:20px;overflow:hidden;">'
                       '<div style="height:100%;width:' + str(_w) + '%;background:linear-gradient(90deg,#16a34a,#4ade80);border-radius:20px;"></div></div></div>')
            st.markdown(_h + "</div>", unsafe_allow_html=True)
        with _c2:
            _h = _card_open("🐌 Najslabiji artikli", "Najmanje prodaju — kandidati za smanjenje zaliha.")
            _mx = max([int(x["prodato"]) for x in _best] or [1]) or 1
            for x in _slab:
                _p = int(x["prodato"]); _w = int(_p / _mx * 100)
                _h += ('<div style="margin-bottom:10px;">'
                       '<div style="display:flex;justify-content:space-between;font-size:12.5px;color:#374151;margin-bottom:3px;">'
                       '<span style="font-weight:600;">' + _h_escape(str(x["artikal"])[:38]) + '</span>'
                       '<span style="font-weight:700;color:#dc2626;">' + _fmt(_p) + ' kom</span></div>'
                       '<div style="height:9px;background:#fef2f2;border-radius:20px;overflow:hidden;">'
                       '<div style="height:100%;width:' + str(max(_w, 2)) + '%;background:linear-gradient(90deg,#f59e0b,#fca5a5);border-radius:20px;"></div></div></div>')
            st.markdown(_h + "</div>", unsafe_allow_html=True)

    def _render_akcija(ak):
        _ua = int(ak.get("ukupno_akcija", 0)); _ur = int(ak.get("ukupno_redovna", 0)); _rz = int(ak.get("razlika", 0))
        _h = ('<div style="background:#fff;border:1px solid #efeaf7;border-radius:16px;padding:20px 22px;margin-bottom:14px;'
              'box-shadow:0 2px 12px rgba(80,40,140,.05);">'
              '<div style="font-size:15px;font-weight:800;margin-bottom:4px;">Uspešnost akcije</div>'
              '<div style="font-size:12.5px;color:#9aa0ad;margin-bottom:16px;">Koliko je akcija donela naspram da se prodavalo po redovnoj ceni, i obrt po artiklu.</div>'
              '<div style="display:grid;grid-template-columns:repeat(3,1fr);gap:14px;">')
        for (col, v, lab) in [("#10b981", _fmt(_ua) + " RSD", "Profit ostvaren na akciji"),
                              ("#7c3aed", _fmt(_ur) + " RSD", "Da je bila redovna cena"),
                              ("#ec4899", ("-" if _rz > 0 else "+") + _fmt(abs(_rz)) + " RSD", "Razlika (koliko je akcija „koštala“)")]:
            _h += ('<div style="background:#faf9fd;border:1px solid #efeaf7;border-left:4px solid ' + col + ';border-radius:14px;padding:15px 17px;">'
                   '<div style="font-size:18px;font-weight:800;color:' + col + ';">' + v + '</div>'
                   '<div style="font-size:11px;color:#9aa0ad;margin-top:4px;font-weight:600;">' + lab + '</div></div>')
        _h += '</div></div>'
        st.markdown(_h, unsafe_allow_html=True)
        _rows = []
        for a in ak.get("artikli", []):
            _rows.append({"Artikal": str(a["naziv"]), "Grupa": str(a.get("grupa", "")),
                          "Prodato (kom)": int(a["prodato"]), "Obrt (x)": round(float(a["obrt"]), 1),
                          "Popust %": round(float(a["popust"]), 1),
                          "Profit akcija (RSD)": int(a["profit_akcija"]),
                          "Cena akcije (RSD)": int(a["cena_akcije"]),
                          "Dani pokrivanja": int(a["dani"])})
        if _rows:
            st.dataframe(pd.DataFrame(_rows), hide_index=True, use_container_width=True,
                         height=min(60 + 35 * len(_rows), 480))

    def _card_open(naslov, podnaslov=""):
        _h = ('<div style="background:#fff;border:1px solid #efeaf7;border-radius:16px;padding:20px 22px;margin-bottom:20px;'
              'box-shadow:0 2px 12px rgba(80,40,140,.05);">'
              '<div style="font-size:15px;font-weight:800;margin-bottom:' + ("4px" if podnaslov else "16px") + ';">' + naslov + '</div>')
        if podnaslov:
            _h += '<div style="font-size:12.5px;color:#9aa0ad;margin-bottom:16px;">' + podnaslov + '</div>'
        return _h

    def _rsd(v):
        try:
            return _fmt(int(round(v))) + " RSD"
        except Exception:
            return str(v) + " RSD"

    def _bar_row(lb, val, maxv, boja):
        _w = int(abs(val) / maxv * 100) if maxv else 0
        _w = min(_w, 100)
        return ('<div style="display:grid;grid-template-columns:70px 1fr 120px;align-items:center;gap:10px;margin-bottom:7px;">'
                '<span style="font-size:11.5px;color:#888;text-align:right;">' + str(lb) + '</span>'
                '<div style="height:16px;background:#f5f0ff;border-radius:4px;overflow:hidden;">'
                '<div style="height:100%;width:' + str(_w) + '%;background:' + boja + ';border-radius:4px;"></div></div>'
                '<span style="font-size:11.5px;font-weight:700;color:#555;">' + _rsd(val) + '</span></div>')

    def _render_profit_blok(pf):
        if not pf:
            return
        n_mes = int(pf.get("n_mes", 1)) or 1
        st.markdown("<div style='margin:26px 0 2px;height:1px;background:linear-gradient(90deg,transparent,#e6def5,transparent);'></div>"
                    "<div style='font-size:16px;font-weight:800;margin:16px 0 4px;'>💰 Profitabilnost</div>"
                    "<div style='font-size:12.5px;color:#9aa0ad;margin-bottom:16px;'>Period: <b>" + _h_escape(str(pf.get("period", ""))) + "</b> · "
                    + str(pf.get("n_obj", 0)) + " objekata · " + str(n_mes) + " meseci</div>", unsafe_allow_html=True)
        # 4 KPI kartice u dinarima
        _kards = [
            ("Trošak marketinga", pf.get("total_trosak", 0), "#a855f7", ""),
            ("Bruto profit", pf.get("total_bruto", 0), "#10b981", ""),
            ("Neto profit", pf.get("total_neto", 0), "#7c3aed" if pf.get("total_neto", 0) > 0 else "#ec4899", ""),
            ("OOS izgubljen", pf.get("total_oos", 0), "#ec4899", "-"),
        ]
        _kh = '<div style="display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin-bottom:20px;">'
        for (lab, tot, col, pre) in _kards:
            _mes = int(round(tot / n_mes))
            _kh += ('<div style="background:#fff;border:1px solid #efeaf7;border-left:4px solid ' + col + ';border-radius:14px;padding:15px 17px;'
                    'box-shadow:0 2px 12px rgba(80,40,140,.06);">'
                    '<div style="font-size:10.5px;color:#9aa0ad;font-weight:700;letter-spacing:.4px;text-transform:uppercase;margin-bottom:6px;">' + lab + '</div>'
                    '<div style="font-size:19px;font-weight:800;color:' + col + ';">' + pre + _rsd(tot) + '</div>'
                    '<div style="font-size:11px;color:#aab;margin-top:3px;">' + pre + _rsd(_mes) + ' / mesec</div></div>')
        _kh += '</div>'
        st.markdown(_kh, unsafe_allow_html=True)
        # Mesečni trend bruto / neto
        _bm = pf.get("bruto_po_mes", []); _nm = pf.get("neto_po_mes", [])
        if _bm or _nm:
            _cb, _cn = st.columns(2)
            with _cb:
                _mx = max([abs(v) for _, v in _bm] or [1]) or 1
                _html = _card_open("📈 Mesečni trend bruto profita")
                for lb, v in _bm:
                    _html += _bar_row(lb, v, _mx, "#a855f7")
                st.markdown(_html + "</div>", unsafe_allow_html=True)
            with _cn:
                _mx = max([abs(v) for _, v in _nm] or [1]) or 1
                _html = _card_open("📉 Mesečni trend neto profita")
                for lb, v in _nm:
                    _html += _bar_row(lb, v, _mx, "#7c3aed" if v >= 0 else "#ec4899")
                st.markdown(_html + "</div>", unsafe_allow_html=True)
        # Profitabilnost po objektima — donut + procena uštede
        _uk = int(pf.get("obj_ukupno", 0))
        if _uk > 0:
            _pr = int(pf.get("obj_profit", 0)); _on = int(pf.get("obj_oos_neg", 0)); _pn = int(pf.get("obj_pravi_neg", 0))
            import math as _m
            _cx, _cy, _rO, _rI = 90, 90, 78, 50
            def _arc(cx, cy, r, sd, ed):
                s = _m.radians(sd - 90); e = _m.radians(ed - 90)
                lg = 1 if (ed - sd) > 180 else 0
                return (cx + r * _m.cos(s), cy + r * _m.sin(s), cx + r * _m.cos(e), cy + r * _m.sin(e), lg)
            _segs = [(_pr, "#10b981"), (_on, "#f59e0b"), (_pn, "#ec4899")]
            _tot = sum(s[0] for s in _segs) or 1
            _nonzero = [s for s in _segs if s[0] > 0]
            _paths = ""
            if len(_nonzero) == 1:
                # jedan segment = pun prsten (SVG luk od 360° se ne iscrtava)
                _col = _nonzero[0][1]
                _paths = ('<circle cx="' + str(_cx) + '" cy="' + str(_cy) + '" r="' + str(_rO) + '" fill="' + _col + '"/>'
                          '<circle cx="' + str(_cx) + '" cy="' + str(_cy) + '" r="' + str(_rI) + '" fill="#ffffff"/>')
            else:
                _ang = 0.0
                for _val, _col in _segs:
                    if _val <= 0:
                        continue
                    _sw = _val / _tot * 360.0
                    x1, y1, x2, y2, lg = _arc(_cx, _cy, _rO, _ang, _ang + _sw)
                    xi2, yi2, xi1, yi1, _ = _arc(_cx, _cy, _rI, _ang, _ang + _sw)
                    _paths += ('<path d="M' + ("%.2f" % x1) + ',' + ("%.2f" % y1) + ' A' + str(_rO) + ',' + str(_rO) + ' 0 ' + str(lg) + ' 1 '
                               + ("%.2f" % x2) + ',' + ("%.2f" % y2) + ' L' + ("%.2f" % xi2) + ',' + ("%.2f" % yi2)
                               + ' A' + str(_rI) + ',' + str(_rI) + ' 0 ' + str(lg) + ' 0 ' + ("%.2f" % xi1) + ',' + ("%.2f" % yi1)
                               + ' Z" fill="' + _col + '"/>')
                    _ang += _sw
            _svg = ('<svg width="180" height="180" viewBox="0 0 180 180" xmlns="http://www.w3.org/2000/svg">' + _paths
                    + '<text x="90" y="86" text-anchor="middle" font-size="26" font-weight="800" fill="#1f2430">' + str(_uk) + '</text>'
                    + '<text x="90" y="104" text-anchor="middle" font-size="10" fill="#9aa0ad">objekata</text></svg>')
            _leg = ('<div style="display:flex;flex-direction:column;gap:10px;">'
                    '<div style="display:flex;align-items:center;gap:8px;"><span style="width:12px;height:12px;border-radius:3px;background:#10b981;"></span>'
                    '<span style="font-size:13px;color:#374151;">Profitabilni objekti: <b>' + str(_pr) + '</b></span></div>'
                    '<div style="display:flex;align-items:center;gap:8px;"><span style="width:12px;height:12px;border-radius:3px;background:#f59e0b;"></span>'
                    '<span style="font-size:13px;color:#374151;">Neprofitabilni zbog OOS: <b>' + str(_on) + '</b></span></div>'
                    '<div style="display:flex;align-items:center;gap:8px;"><span style="width:12px;height:12px;border-radius:3px;background:#ec4899;"></span>'
                    '<span style="font-size:13px;color:#374151;">Pravi neprofitabilni: <b>' + str(_pn) + '</b></span></div>'
                    '<div style="margin-top:8px;font-size:12.5px;color:#6b7280;line-height:1.5;">Procena uštede ako se pravi neprofitabilni ugase: '
                    '<b style="color:#16a34a;">' + _rsd(pf.get("usteda_ukupno", 0)) + '</b> za period.</div></div>')
            _ph = (_card_open("🏪 Profitabilnost po objektima")
                   + '<div style="display:grid;grid-template-columns:200px 1fr;align-items:center;gap:18px;">'
                   + '<div style="text-align:center;">' + _svg + '</div>' + _leg + '</div></div>')
            st.markdown(_ph, unsafe_allow_html=True)
            # tabela objekata
            _objs = pf.get("objekti", [])
            if _objs:
                _km = sb_komitenti_map()
                _rows = []
                for o in _objs:
                    _nz = _km.get(int(o["id"]), "") or ("ID " + str(o["id"]))
                    _rows.append({"Objekat": _nz, "Neto profit (RSD)": o["neto"],
                                  "Bruto (RSD)": o["bruto"], "Trošak (RSD)": o["trosak"],
                                  "Izgubljeno OOS (RSD)": o["oos"], "Potencijal (RSD)": o["potencijal"]})
                st.markdown("<div style='font-size:13px;font-weight:700;color:#374151;margin:4px 0 8px;'>Svi objekti (od najlošijeg neto profita):</div>", unsafe_allow_html=True)
                st.dataframe(pd.DataFrame(_rows), hide_index=True, use_container_width=True, height=340)
        # OOS — izgubljena zarada (u dinarima)
        _oa = pf.get("oos_artikli", [])
        _total_oos = int(pf.get("total_oos", 0))
        if _total_oos > 0 or _oa:
            _oh = _card_open("🔴 Out of stock — izgubljena zarada", "Koliko dinara je izgubljeno jer je artikal bio na nuli.")
            _mesv = int(round(_total_oos / n_mes))
            _oh += '<div style="display:grid;grid-template-columns:repeat(3,1fr);gap:14px;margin-bottom:6px;">'
            for (lab, val) in [("Izgubljen profit · " + str(n_mes) + " mes.", _rsd(_total_oos)),
                               ("Prosečno mesečno", _rsd(_mesv)),
                               ("Kombinacija na 0 danas", _fmt(pf.get("oos_0_danas", 0)))]:
                _oh += ('<div style="background:#fef2f2;border:1px solid #fecaca;border-radius:14px;padding:15px;text-align:center;">'
                        '<div style="font-size:18px;font-weight:800;color:#dc2626;">' + str(val) + '</div>'
                        '<div style="font-size:11px;color:#9b6b6b;margin-top:3px;font-weight:600;">' + lab + '</div></div>')
            _oh += '</div></div>'
            st.markdown(_oh, unsafe_allow_html=True)
            if _oa:
                _df = pd.DataFrame([{"Artikal": r["naziv"], "U koliko objekata": r["objekata"],
                                     "OOS meseci": r["meseci"], "Izgubljeni profit (RSD)": r["rsd"]} for r in _oa])
                st.dataframe(_df, hide_index=True, use_container_width=True, height=320)

    def _partial_iz_stavki(stavke):
        out = {}
        g = {}
        for s in stavke:
            gg = str(s.get("grupa", "") or "—")
            g[gg] = g.get(gg, 0) + int(s.get("kol", 0) or 0)
        out["po_grupama"] = [{"grupa": k, "kom": v} for k, v in
                             sorted(g.items(), key=lambda kv: kv[1], reverse=True) if v > 0]
        _oi = [s for s in stavke if int(s.get("lager", 0) or 0) == 0 and int(s.get("pred", 0) or 0) > 0]
        out["oos"] = {"kombinacija_na_0": len(_oi), "izgubljeno_kom": sum(int(s.get("pred", 0) or 0) for s in _oi)}
        _da = {}
        for s in _oi:
            nz = str(s.get("naziv", ""))
            e = _da.setdefault(nz, {"obj": set(), "izg": 0})
            e["obj"].add(int(s["idk"])); e["izg"] += int(s.get("pred", 0) or 0)
        _top = sorted(_da.items(), key=lambda kv: kv[1]["izg"], reverse=True)[:10]
        out["oos_po_artiklu"] = [{"artikal": k, "objekata": len(v["obj"]), "izgubljeno": v["izg"]} for k, v in _top]
        return out

    # ---------- TABLA (dve kartice) ----------
    if _view == "dash":
        st.markdown('<div style="color:#6b7280;font-size:14px;margin:6px 0 20px;">Dobrodošli 👋 &nbsp;'
                    'Pregled izveštaja i efikasnosti administracije.</div>', unsafe_allow_html=True)
        st.markdown('<div style="font-size:12px;text-transform:uppercase;letter-spacing:.6px;color:#9aa0ad;'
                    'font-weight:700;margin-bottom:12px;">Kartice</div>', unsafe_allow_html=True)
        _cc1, _cc2, _cc3 = st.columns(3)
        with _cc1:
            with st.container(border=True):
                st.markdown('<div style="font-size:32px;">📊</div>'
                            '<div style="font-size:16px;font-weight:800;margin:6px 0 4px;">Izveštaj efikasnosti administracije</div>'
                            '<div style="font-size:13px;color:#8b8fa0;line-height:1.5;margin-bottom:12px;">PDF izveštaj administracije i upozorenja koja je administracija prosledila direktoru.</div>',
                            unsafe_allow_html=True)
                if st.button("Otvori →", key="dir_open_efik", use_container_width=True):
                    st.session_state["dir_view"] = "efikasnost"; st.rerun()
        with _cc2:
            with st.container(border=True):
                st.markdown('<div style="font-size:32px;">📈</div>'
                            '<div style="font-size:16px;font-weight:800;margin:6px 0 4px;">Detaljan izveštaj po sistemima</div>'
                            '<div style="font-size:13px;color:#8b8fa0;line-height:1.5;margin-bottom:12px;">Prodaja, grupe, out-of-stock i preuzimanje analitike — po sistemu i mesecu.</div>',
                            unsafe_allow_html=True)
                if st.button("Otvori →", key="dir_open_sist", use_container_width=True):
                    st.session_state["dir_view"] = "sistemi"; st.rerun()
        with _cc3:
            with st.container(border=True):
                st.markdown('<div style="font-size:32px;">💹</div>'
                            '<div style="font-size:16px;font-weight:800;margin:6px 0 4px;">Izveštaj prodaje</div>'
                            '<div style="font-size:13px;color:#8b8fa0;line-height:1.5;margin-bottom:12px;">Kompletan dashboard: prodaja, uspešnost akcije, profitabilnost, zalihe — plus izvoz u Excel.</div>',
                            unsafe_allow_html=True)
                if st.button("Otvori →", key="dir_open_prod", use_container_width=True):
                    st.session_state["dir_view"] = "prodaja"; st.rerun()
        _cd1, _cd2, _cd3 = st.columns(3)
        with _cd1:
            with st.container(border=True):
                st.markdown('<div style="font-size:32px;">🧊</div>'
                            '<div style="font-size:16px;font-weight:800;margin:6px 0 4px;">Izveštaj SYX</div>'
                            '<div style="font-size:13px;color:#8b8fa0;line-height:1.5;margin-bottom:12px;">Word izveštaji za SYX nikotinske vrećice, izlistani po mesecima za preuzimanje.</div>',
                            unsafe_allow_html=True)
                if st.button("Otvori →", key="dir_open_syx", use_container_width=True):
                    st.session_state["dir_view"] = "syx"; st.rerun()
        with _cd2:
            with st.container(border=True):
                st.markdown('<div style="font-size:32px;">💳</div>'
                            '<div style="font-size:16px;font-weight:800;margin:6px 0 4px;">Izveštaj potraživanja</div>'
                            '<div style="font-size:13px;color:#8b8fa0;line-height:1.5;margin-bottom:12px;">Potraživanja po mesecu koja je administracija dopunila — pregled i izvoz u Excel.</div>',
                            unsafe_allow_html=True)
                if st.button("Otvori →", key="dir_open_potraz", use_container_width=True):
                    st.session_state["dir_view"] = "potraz"; st.rerun()
        with _cd3:
            with st.container(border=True):
                st.markdown('<div style="font-size:32px;">📅</div>'
                            '<div style="font-size:16px;font-weight:800;margin:6px 0 4px;">Rokovi</div>'
                            '<div style="font-size:13px;color:#8b8fa0;line-height:1.5;margin-bottom:12px;">Postavi rokove po mesecu: administracija, izveštaj po sistemu i osvežavanje izveštaja prodaje.</div>',
                            unsafe_allow_html=True)
                if st.button("Otvori →", key="dir_open_rok", use_container_width=True):
                    st.session_state["dir_view"] = "rokovi"; st.rerun()
        return

    if st.button("← Nazad na kartice", key="dir_back"):
        st.session_state["dir_view"] = "dash"; st.rerun()

    # ---------- KARTICA: IZVEŠTAJ POTRAŽIVANJA ----------
    if _view == "potraz":
        potraz_director_ui()
        return

    # ---------- KARTICA 4: ROKOVI ----------
    if _view == "rokovi":
        st.markdown('<div style="font-size:20px;font-weight:800;margin:6px 0 6px;">📅 Rokovi</div>', unsafe_allow_html=True)
        st.caption("Postavi rokove po mesecu. Administracija radi do svog roka pa se zaključava; izveštaj efikasnosti "
                   "vidiš tek kad rok prođe. Za izveštaj po sistemu direktori vide poruku o roku dok se ne objavi. "
                   "Izveštaj prodaje se uvek vidi, a rok i napomena stoje u uglu.")

        # meseci: par unazad + par unapred + objavljeni
        _today = datetime.date.today()
        _opts = set(_mes_keys)
        _yy, _mm = _today.year, _today.month - 2
        while _mm <= 0:
            _mm += 12; _yy -= 1
        for _ in range(9):
            _opts.add(str(_yy) + "-" + ("0" + str(_mm))[-2:])
            _mm += 1
            if _mm > 12:
                _mm = 1; _yy += 1
        _opts = sorted(_opts, reverse=True)
        _rsel = st.selectbox("Mesec", _opts, format_func=mesec_label, key="rok_mes_sel")
        _post = sb_rokovi_get(_rsel)

        def _dflt(v):
            try:
                return datetime.date.fromisoformat(str(v)[:10])
            except Exception:
                return _today

        _r1, _r2, _r3 = st.columns(3)
        with _r1:
            _da = st.date_input("Rok — Izveštaj administracije", value=_dflt(_post.get("rok_admin")),
                                key="rok_admin_in", format="DD.MM.YYYY")
        with _r2:
            _ds = st.date_input("Rok — Izveštaj po sistemu", value=_dflt(_post.get("rok_sistemi")),
                                key="rok_sis_in", format="DD.MM.YYYY")
        with _r3:
            _dp = st.date_input("Rok — Osvežavanje izveštaja prodaje", value=_dflt(_post.get("rok_prodaja")),
                                key="rok_prod_in", format="DD.MM.YYYY")
        _r4, _r5, _r6 = st.columns(3)
        with _r4:
            _dsx = st.date_input("Rok — Izveštaj SYX", value=_dflt(_post.get("rok_syx")),
                                 key="rok_syx_in", format="DD.MM.YYYY")
        with _r5:
            _dpz = st.date_input("Rok — Izveštaj potraživanja", value=_dflt(_post.get("rok_potraz")),
                                 key="rok_potraz_in", format="DD.MM.YYYY")
        with _r6:
            st.markdown("<div style='height:8px;'></div>", unsafe_allow_html=True)
        _nap = st.text_area("Napomena (za izveštaj prodaje — šta osvežiti, na šta obratiti pažnju)",
                            value=_post.get("napomena") or "", key="rok_nap_in", height=80)
        if st.button("💾 Sačuvaj rokove", key="rok_save", type="primary"):
            try:
                sb_rokovi_set(_rsel, _da.isoformat(), _ds.isoformat(), _dp.isoformat(), _nap,
                              rok_syx=_dsx.isoformat(), rok_potraz=_dpz.isoformat())
                st.success("Rokovi za " + mesec_label(_rsel) + " sačuvani.")
                st.rerun()
            except Exception as _e:
                st.error("Greška pri čuvanju: " + str(_e))

        _all = sb_rokovi_all()
        if _all:
            st.markdown("<div style='margin:18px 0 6px;font-size:12px;text-transform:uppercase;letter-spacing:.6px;"
                        "color:#9aa0ad;font-weight:700;'>Postavljeni rokovi</div>", unsafe_allow_html=True)
            _rows = []
            for _mk in sorted(_all.keys(), reverse=True):
                _v = _all[_mk]
                _rows.append({"Mesec": mesec_label(_mk),
                              "Administracija": _rok_fmt(_v.get("rok_admin")),
                              "Po sistemu": _rok_fmt(_v.get("rok_sistemi")),
                              "Izveštaj prodaje": _rok_fmt(_v.get("rok_prodaja")),
                              "SYX": _rok_fmt(_v.get("rok_syx")),
                              "Potraživanja": _rok_fmt(_v.get("rok_potraz")),
                              "Napomena": (_v.get("napomena") or "")[:50]})
            st.dataframe(pd.DataFrame(_rows), hide_index=True, use_container_width=True)
        return

    # ---------- KARTICA: IZVEŠTAJ SYX ----------
    if _view == "syx":
        st.markdown('<div style="font-size:20px;font-weight:800;margin:6px 0 6px;">🧊 Izveštaj SYX (nikotinske vrećice)</div>', unsafe_allow_html=True)
        st.caption("Word izveštaji za SYX, po mesecima. Analitičar ih ubacuje u delu Objava izveštaja.")
        _syl = sb_syx_list()
        if not _syl:
            st.info("Još nema objavljenih SYX izveštaja.")
            return
        import base64 as _b64y
        for _r in _syl:
            _mk = _r.get("mesec", "")
            with st.container(border=True):
                _sc1, _sc2 = st.columns([3, 1])
                with _sc1:
                    st.markdown("<div style='font-size:15px;font-weight:800;'>" + _h_escape(mesec_label(_mk)) + "</div>"
                                "<div style='font-size:12.5px;color:#9aa0ad;margin-top:2px;'>" + _h_escape(str(_r.get("filename", ""))) + "</div>",
                                unsafe_allow_html=True)
                with _sc2:
                    _doc = sb_syx_get(_mk)
                    if _doc and _doc.get("docx_b64"):
                        try:
                            _fn = str(_r.get("filename") or ("Izvestaj_SYX_" + _mk + ".docx"))
                            st.download_button("⬇️ Preuzmi", _b64y.b64decode(_doc["docx_b64"]),
                                file_name=_fn,
                                mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                                key="syx_dl_" + _mk, use_container_width=True)
                        except Exception:
                            st.caption("Greška pri učitavanju dokumenta.")
        return

    # ---------- KARTICA 1: EFIKASNOST ADMINISTRACIJE ----------
    if _view == "efikasnost":
        st.markdown('<div style="font-size:20px;font-weight:800;margin:6px 0 14px;">📊 Izveštaj efikasnosti administracije</div>', unsafe_allow_html=True)
        _sel_lbl = st.selectbox("Mesec", _mlbls, index=_prev_idx, key="dir_efik_mes")
        mesec_key = _mes_keys[_mlbls.index(_sel_lbl)]

        # Zaključavanje: izveštaj efikasnosti se vidi TEK kad prođe rok administracije.
        _rok_a = sb_rokovi_get(mesec_key).get("rok_admin")
        if _rok_a and not _rok_je_prosao(_rok_a):
            st.info("Izveštaj efikasnosti za " + _sel_lbl + " biće dostupan posle roka administracije: "
                    + _rok_fmt(_rok_a) + ". Do tada administracija još obrađuje podatke.")
            return

        _pc1, _pc2 = st.columns([1, 2])
        with _pc1:
            if st.button("📄 Napravi PDF izveštaj", key="dir_pdf_make", use_container_width=True):
                try:
                    st.session_state["dir_pdf"] = napravi_pdf_izvestaj(mesec_key, _sel_lbl)
                    st.session_state["dir_pdf_for"] = mesec_key
                except Exception as _e:
                    st.session_state["dir_pdf"] = None
                    st.error("Greška pri pravljenju PDF-a: " + str(_e))
        with _pc2:
            if st.session_state.get("dir_pdf") and st.session_state.get("dir_pdf_for") == mesec_key:
                st.download_button("⬇️ Preuzmi PDF", st.session_state["dir_pdf"],
                    file_name="Izvestaj_administracije_" + mesec_key + ".pdf", mime="application/pdf",
                    key="dir_pdf_dl", use_container_width=True)

        st.markdown('<div style="margin:18px 0 4px;font-size:12px;text-transform:uppercase;letter-spacing:.6px;'
                    'color:#9aa0ad;font-weight:700;">⚠️ Upozorenja prosleđena od administracije</div>', unsafe_allow_html=True)
        st.caption("Objekti koje je administracija označila da su prosleđeni direktoru — problem, koliko treba/koliko je poručeno i šta je preduzeto.")

        _kf = st.session_state.get("_komfull_dir")
        if _kf is None:
            _kf = sb_komitenti_full(); st.session_state["_komfull_dir"] = _kf

        def _chip(on, txt):
            _bg = "#dcfce7;color:#166534" if on else "#f3f4f6;color:#9ca3af"
            return '<span style="background:' + _bg + ';padding:4px 11px;border-radius:20px;font-weight:600;font-size:12px;">' + txt + '</span>'

        _found = 0
        for _sis in sb_sisteme(mesec_key):
            _pod = sb_ucitaj(mesec_key, _sis)
            if not _pod or not _pod.get("stavke"):
                continue
            _po = {}
            for s in _pod["stavke"]:
                _po.setdefault(int(s["idk"]), []).append(s)
            _obr = sb_load_obrada(mesec_key, _sis)
            for _idk, _v in _obr.items():
                if "Obavestila direktorku" not in (_v.get("reakcije") or []):
                    continue
                _lst = _po.get(int(_idk))
                if not _lst:
                    continue
                _found += 1
                nivo, n_nula, izgub = hitnost_objekta(_lst)
                z = _zona_disp(nivo)
                _naz = (_kf.get(int(_idk), {}) or {}).get("naziv", "") or ("ID " + str(_idk))
                _treba = sum(int(a.get("kol", 0) or 0) for a in _lst)
                _njih = sum(int(x) for x in (_v.get("njihova") or {}).values())
                _reak = _v.get("reakcije") or []
                _nap = _v.get("napomena", "") or ""
                _bc = "#ef4444" if nivo == "crveno" else ("#f59e0b" if nivo == "zuto" else "#22c55e")
                _html = ('<div style="background:#fff;border:1px solid #efeaf7;border-left:4px solid ' + _bc + ';'
                         'border-radius:14px;padding:16px 18px;margin-bottom:12px;box-shadow:0 2px 12px rgba(80,40,140,.05);">'
                         '<div style="display:flex;align-items:center;gap:10px;flex-wrap:wrap;margin-bottom:8px;">'
                         '<span style="font-weight:700;font-size:14.5px;">' + _h_escape(_naz) + '</span>'
                         '<span style="font-size:12px;color:#9aa0ad;">· ' + _h_escape(_sis) + '</span>'
                         '<span style="margin-left:auto;font-size:12px;font-weight:700;color:' + _bc + ';">' + z[3] + '</span></div>'
                         '<div style="font-size:13px;color:#4b5563;margin-bottom:11px;">Problem: <b>' + str(n_nula)
                         + '</b> artikala na nuli · procenjeno izgubljeno <b>' + str(izgub) + '</b> kom/mesec.</div>'
                         '<div style="display:grid;grid-template-columns:repeat(2,1fr);gap:10px;margin-bottom:11px;">'
                         '<div style="background:#faf9fd;border-radius:10px;padding:8px 12px;">'
                         '<div style="font-size:10.5px;color:#9aa0ad;text-transform:uppercase;font-weight:700;">Treba da poruči</div>'
                         '<div style="font-size:16px;font-weight:800;color:#7c3aed;">' + _fmt(_treba) + ' kom</div></div>'
                         '<div style="background:#faf9fd;border-radius:10px;padding:8px 12px;">'
                         '<div style="font-size:10.5px;color:#9aa0ad;text-transform:uppercase;font-weight:700;">Poručio (njihovo)</div>'
                         '<div style="font-size:16px;font-weight:800;color:' + ("#16a34a" if _njih > 0 else "#dc2626") + ';">' + _fmt(_njih) + ' kom</div></div>'
                         '</div>'
                         '<div style="display:flex;gap:8px;flex-wrap:wrap;' + ('margin-bottom:9px;' if _nap else '') + '">'
                         + _chip("Pozvala sam" in _reak, "📞 Pozvano")
                         + _chip("Poslala sam mejl" in _reak, "✉️ Mejl")
                         + _chip(True, "👤 Prosleđeno direktoru")
                         + '</div>'
                         + (('<div style="background:#fffbeb;border:1px solid #fde68a;border-radius:9px;padding:8px 12px;'
                             'font-size:12.5px;color:#78500a;"><b style="color:#4b5563;">Napomena:</b> ' + _h_escape(_nap) + '</div>') if _nap else '')
                         + '</div>')
                st.markdown(_html, unsafe_allow_html=True)
        if _found == 0:
            st.caption("Nema upozorenja prosleđenih direktoru za ovaj mesec.")
        return

    # ---------- KARTICA 2: DETALJAN IZVEŠTAJ PO SISTEMIMA ----------
    if _view == "sistemi":
        st.markdown('<div style="font-size:20px;font-weight:800;margin:6px 0 12px;">📈 Detaljan izveštaj po sistemima</div>', unsafe_allow_html=True)
        _c1, _c2 = st.columns(2)
        with _c1:
            _sel_lbl = st.selectbox("Mesec", _mlbls, index=_prev_idx, key="dir_sis_mes")
        mesec_key = _mes_keys[_mlbls.index(_sel_lbl)]
        _sisteme = sb_sisteme(mesec_key)
        if not _sisteme:
            _rs = sb_rokovi_get(mesec_key).get("rok_sistemi")
            if _rs and not _rok_je_prosao(_rs):
                st.info("Izveštaj po sistemu za " + _sel_lbl + " još nije objavljen — biće objavljen najkasnije do "
                        + _rok_fmt(_rs) + " (rok za popunjavanje još nije istekao).")
            elif _rs:
                st.warning("Izveštaj po sistemu za " + _sel_lbl + " nije objavljen, a rok ("
                           + _rok_fmt(_rs) + ") je istekao.")
            else:
                st.info("Za " + _sel_lbl + " još nema objavljenih sistema (rok nije postavljen).")
            return
        with _c2:
            sistem = st.selectbox("Sistem", _sisteme, index=0, key="dir_sis_sis")

        st.markdown("<div style='margin:8px 0 14px;font-size:12px;text-transform:uppercase;letter-spacing:.6px;"
                    "color:#9aa0ad;font-weight:700;'>" + _h_escape(sistem) + " · " + _sel_lbl + "</div>",
                    unsafe_allow_html=True)

        podaci = sb_ucitaj(mesec_key, sistem)
        if not podaci:
            _rs = sb_rokovi_get(mesec_key).get("rok_sistemi")
            if _rs:
                st.info("Izveštaj za „" + _h_escape(str(sistem)) + "“ (" + _sel_lbl
                        + ") biće objavljen najkasnije do " + _rok_fmt(_rs) + ".")
            else:
                st.info("Izveštaj za ovaj sistem/mesec još nije objavljen.")
            return

        # Preuzimanje analitike (Excel, bez sheeta o modelu)
        _dl1, _dl2 = st.columns([1, 2])
        with _dl1:
            if st.button("⬇️ Pripremi analitiku (Excel)", key="dir_prep_xlsx", use_container_width=True):
                _b = sb_ucitaj_xlsx(mesec_key, sistem)
                if _b:
                    import base64 as _b64
                    try:
                        st.session_state["dir_xlsx_bytes"] = _b64.b64decode(_b)
                    except Exception:
                        st.session_state["dir_xlsx_bytes"] = None
                else:
                    st.session_state["dir_xlsx_bytes"] = None
                st.session_state["dir_xlsx_for"] = (mesec_key, sistem)
        with _dl2:
            if st.session_state.get("dir_xlsx_for") == (mesec_key, sistem):
                _xb = st.session_state.get("dir_xlsx_bytes")
                if _xb:
                    st.download_button("⬇️ Sačuvaj Analitika.xlsx", _xb,
                        file_name="Analitika_" + str(sistem).replace(" ", "_") + "_" + mesec_key + ".xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dir_dl_xlsx", use_container_width=True)
                else:
                    st.caption("Analitika (Excel) za ovaj sistem biće dostupna čim se sistem ponovo objavi.")
        st.markdown("<div style='height:8px;'></div>", unsafe_allow_html=True)

        d = podaci.get("direktor") or {}
        if not isinstance(d, dict):
            d = {}
        _pf = d.get("profit")  # profitabilnost iz sačuvane analitike (ako je sistem objavljen novom verzijom)

        # Prodaju/trend/grupe uzimamo prvenstveno iz tabele prodaje (18 meseci, lepši trend);
        # profitabilnost uvek iz sačuvane analitike ovog sistema.
        _izvp = sb_ucitaj_izvestaj_prodaje()
        try:
            _sales = json.loads(_izvp["prodaja_json"]) if (_izvp and _izvp.get("prodaja_json")) else {}
        except Exception:
            _sales = {}
        _dsales = _direktor_blok_iz_prodaje(sistem, _sales)

        # Iz analitike (objava sistema): prosek po objektu, OOS, porudžbina, artikli, akcija, profit
        _extra_keys = ["profit", "prosek_po_objektu", "oos_kom", "porudzbina", "artikli_rang", "akcija"]
        _has_analitika = any(d.get(k) for k in _extra_keys)

        def _pripoji(base):
            for _k in _extra_keys:
                if d.get(_k):
                    base[_k] = d.get(_k)
            return base

        if _dsales:
            _part = _partial_iz_stavki(podaci.get("stavke") or [])
            _dsales["oos"] = _part.get("oos")
            _dsales["oos_po_artiklu"] = _part.get("oos_po_artiklu")
            _pripoji(_dsales)
            st.caption("Prodaja i grupe su iz tabele prodaje; ostalo (porudžbina, artikli, akcija, OOS, profit) je iz analitike sistema.")
            _render_sistem_report(_dsales, True)
        elif d.get("prodaja_trend"):
            _render_sistem_report(_pripoji(d), True)
        else:
            _base = _pripoji(_partial_iz_stavki(podaci.get("stavke") or []))
            if not _has_analitika:
                st.info("Prodaja i trend se pojave kad objaviš Izveštaj prodaje (ako tabela prodaje sadrži ovaj sistem), "
                        "ili kad ponovo objaviš ovaj sistem. Ispod je što je već dostupno.")
            _render_sistem_report(_base, False)

        if not _has_analitika:
            st.markdown("<div style='margin-top:10px;padding:11px 14px;background:#fff7ed;border:1px solid #fed7aa;"
                        "border-radius:10px;font-size:12.5px;color:#9a5b1e;'>Detaljni delovi (prosek po objektu, predlog "
                        "porudžbine, bestseleri, uspešnost akcije, OOS po količinama, profitabilnost) se pojave čim ponovo "
                        "objaviš ovaj sistem novom verzijom aplikacije. Fajl je isti kao i do sad.</div>", unsafe_allow_html=True)
        return

    # ---------- KARTICA 3: IZVEŠTAJ PRODAJE (dashboard, pun ekran) ----------
    if _view == "prodaja":
        # Proširi na pun ekran (samo ovaj prikaz)
        st.markdown("<style>.block-container{max-width:100% !important;"
                    "padding-left:1rem !important;padding-right:1rem !important;padding-top:10px !important;}</style>",
                    unsafe_allow_html=True)
        _izv = sb_ucitaj_izvestaj_prodaje()
        if not _izv or not _izv.get("html"):
            st.info("Izveštaj prodaje još nije objavljen. Analitičar ga pravi u delu Objava izveštaja → Izveštaj prodaje.")
            return
        _mlbl = _izv.get("mesec_label") or ""
        _gen = _izv.get("generisano") or ""
        _tc1, _tc2 = st.columns([4, 1])
        with _tc1:
            st.caption("💹 Izveštaj prodaje · poslednji mesec " + str(_mlbl)
                       + ("  ·  generisano " + str(_gen) if _gen else ""))
        with _tc2:
            _xb64 = _izv.get("xlsx_b64")
            if _xb64:
                try:
                    import base64 as _b64d
                    st.download_button("⬇️ Izvezi u Excel", _b64d.b64decode(_xb64),
                        file_name="Izvestaj_prodaje.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dir_prod_xlsx", use_container_width=True)
                except Exception:
                    pass
        # Rok za osvežavanje + napomena (u uglu) — iz rokova tekućeg meseca
        _tt = datetime.date.today()
        _curmk = str(_tt.year) + "-" + ("0" + str(_tt.month))[-2:]
        _rp = sb_rokovi_get(_curmk)
        _rpd = _rp.get("rok_prodaja")
        _rpn = (_rp.get("napomena") or "").strip()
        if _rpd or _rpn:
            _corner = ('<div style="display:flex;justify-content:flex-end;margin:-4px 0 10px;">'
                       '<div style="max-width:520px;background:#fff7ed;border:1px solid #fed7aa;border-radius:12px;'
                       'padding:10px 14px;font-size:12.5px;color:#9a5b1e;">')
            if _rpd:
                _corner += '⏰ Rok za osvežavanje podataka: <b>' + _rok_fmt(_rpd) + '</b>'
            if _rpn:
                _corner += ('<div style="margin-top:5px;color:#7c5320;"><b>Napomena:</b> ' + _h_escape(_rpn) + '</div>')
            _corner += '</div></div>'
            st.markdown(_corner, unsafe_allow_html=True)
        import streamlit.components.v1 as _components
        _components.html(_izv["html"], height=2600, scrolling=True)
        return


# =====================================================================
# ROUTER: prijava -> uloga
# =====================================================================
if not check_password():
    st.stop()

if st.session_state.get("role") == "administracija":
    prikazi_administraciju()
    st.stop()

if st.session_state.get("role") == "direktori":
    prikazi_direktore()
    st.stop()

# --- od ovde nadole ide ANALITIČKI deo (pun pristup) ---

WMA_WEIGHTS = np.array([0.03, 0.07, 0.12, 0.28, 0.50])
HIST_WEIGHT = 0.03

class PredictionEngine:
    def __init__(self, file_bytes, excluded_ids, alpha, beta, min_lager, min_order, mesecni_trosak=0, analitika_meseci=None, min_per_artikal=None, meseci=1.0, max_per_artikal=None):
        self.file_bytes = file_bytes; self.excluded = excluded_ids
        self.alpha = alpha; self.beta = beta; self.min_lager = min_lager; self.min_order = min_order
        self.min_per_artikal = min_per_artikal
        self.max_per_artikal = max_per_artikal
        self.meseci = meseci if (meseci and meseci > 0) else 1.0
        self.mesecni_trosak = mesecni_trosak
        self.analitika_meseci = analitika_meseci
        self.logs = []; self.adjustments = []; self.has_history = False
        self.has_prices = False
    def log(self, msg): self.logs.append(msg)
    def run(self, progress_bar):
        progress_bar.progress(5, "Ucitavanje..."); self._load_sheets()
        progress_bar.progress(15, "Priprema..."); self._prepare_lookups()
        progress_bar.progress(25, "Povrat/korekcija..."); self._compute_povrat()
        progress_bar.progress(40, "Mesecni pregled..."); self._build_monthly()
        progress_bar.progress(55, "Predikcija..."); self._predict_all()
        progress_bar.progress(70, "Lager..."); self._merge_lager()
        progress_bar.progress(80, "Porudzbine..."); self._compute_orders()
        progress_bar.progress(85, "Analitika..."); self._apply_min_order(); self._compute_analytics()
        progress_bar.progress(100, "Gotovo!"); return self.df_result
    def _load_sheets(self):
        xls = pd.ExcelFile(io.BytesIO(self.file_bytes))
        sm = {s.strip().lower(): s for s in xls.sheet_names}
        def find(kws):
            for kw in kws:
                for nl, no in sm.items():
                    if kw in nl: return no
            return None
        s_prod=find(['prodaja']); s_start=find(['startni']); s_pov=find(['povrat'])
        s_tl=find(['trenutni']); s_hist=find(['pre sept','pre sep','istorij'])
        if not s_prod: raise ValueError("Nema sheeta 'prodaja'!")
        if not s_start: raise ValueError("Nema sheeta 'startni lager'!")
        self.prodaja = pd.read_excel(xls, sheet_name=s_prod); self.prodaja.columns=[c.strip() for c in self.prodaja.columns]
        self.prodaja = self.prodaja[[c for c in self.prodaja.columns if 'Unnamed' not in str(c)]]
        self.log(f"Prodaja: {len(self.prodaja)} redova")
        self.region_map = {}
        if 'Region' in self.prodaja.columns:
            self.region_map = self.prodaja.drop_duplicates('ID KOMITENTA').set_index('ID KOMITENTA')['Region'].to_dict()
            self.log(f"Region: {len(set(self.region_map.values()))} regiona")
        self.startni = pd.read_excel(xls, sheet_name=s_start); self.startni.columns=[c.strip() for c in self.startni.columns]
        self.log(f"Startni: {len(self.startni)} redova")
        price_cols = ['Redovna cena','Akcijska cena','Finalna cena','Nabavna vrednost','Profit']
        if all(c in self.prodaja.columns for c in price_cols):
            self.has_prices = True; self.log("Cene i profit: DA")
        self.povrat_df = pd.DataFrame()
        if s_pov:
            self.povrat_df = pd.read_excel(xls, sheet_name=s_pov); self.povrat_df.columns=[c.strip() for c in self.povrat_df.columns]
            self.log(f"Povrat: {len(self.povrat_df)} redova")
        self.trenutni = pd.DataFrame()
        if s_tl:
            self.trenutni = pd.read_excel(xls, sheet_name=s_tl); self.trenutni.columns=[c.strip() for c in self.trenutni.columns]
            self.log(f"Trenutni lager: {len(self.trenutni)} redova")
        self.hist_df = pd.DataFrame()
        self.has_history = False
        _meseci_u_prodaji = self.prodaja[['Godina','Mesec']].drop_duplicates().values.tolist()
        _ima_pre_sept = any((int(g) < 2025) or (int(g) == 2025 and int(m) < 9) for g, m in _meseci_u_prodaji)
        if _ima_pre_sept:
            self.log("Rezim: KOMPLETAN ISTORIJAT u prodaja sheetu — istorijski sheet se ignorise")
        elif s_hist:
            self.hist_df = pd.read_excel(xls, sheet_name=s_hist); self.hist_df.columns=[c.strip() for c in self.hist_df.columns]
            self.has_history = True; self.log(f"Istorija: {len(self.hist_df)} redova")
        self.meseci_order = sorted(self.prodaja[['Godina','Mesec']].drop_duplicates().values.tolist())
        mn={1:'Jan',2:'Feb',3:'Mar',4:'Apr',5:'Maj',6:'Jun',7:'Jul',8:'Avg',9:'Sep',10:'Okt',11:'Nov',12:'Dec'}
        self.mesec_labels = [f"{mn.get(int(m),'?')} {int(g)}" for g,m in self.meseci_order]
        lg,lm = self.meseci_order[-1]; nm=int(lm)+1; ng=int(lg)
        if nm>12: nm=1; ng+=1
        self.pred_label = f"{mn.get(nm,'?')} {ng}"
        # Porudzbina je za isti mesec kao i predikcija (poslednji mesec podataka + 1)
        self.order_label = self.pred_label
        self.log(f"Meseci: {', '.join(self.mesec_labels)}")
        self.num_komitenti = self.prodaja['ID KOMITENTA'].nunique()
        self.trosak_po_objektu = self.mesecni_trosak / max(self.num_komitenti, 1) if self.mesecni_trosak > 0 else 0
        if self.mesecni_trosak > 0:
            self.log(f"Ukupan trosak: {self.mesecni_trosak:,.0f} / {self.num_komitenti} objekata = {self.trosak_po_objektu:,.0f} po objektu za period")
    def _prepare_lookups(self):
        kp = self.prodaja[['ID KOMITENTA','id artikla','Naziv artikla','Grupa']].drop_duplicates()
        ks = self.startni[['ID KOMITENTA','id artikla','Naziv artikla','Grupa']].drop_duplicates()
        frames = [kp, ks]
        if self.has_history:
            hcols = self.hist_df.columns.tolist()
            col_map = {}
            for c in hcols:
                cl = c.lower()
                if 'komitent' in cl: col_map[c] = 'ID KOMITENTA'
                elif 'id' in cl and 'artikl' in cl: col_map[c] = 'id artikla'
                elif 'naziv' in cl and 'artikl' in cl: col_map[c] = 'Naziv artikla'
                elif 'grup' in cl: col_map[c] = 'Grupa'
            hdf = self.hist_df.rename(columns=col_map)
            for nc in ['ID KOMITENTA','id artikla','Naziv artikla','Grupa']:
                if nc not in hdf.columns: hdf[nc] = ''
            kh = hdf[['ID KOMITENTA','id artikla','Naziv artikla','Grupa']].drop_duplicates()
            frames.append(kh)
        self.all_keys = pd.concat(frames).drop_duplicates().sort_values(['ID KOMITENTA','id artikla']).reset_index(drop=True)
        self.startni_dict = {(r['ID KOMITENTA'],r['id artikla']): r['Kolicina'] for _,r in self.startni.iterrows()}
        self.has_promet = 'PROMET KA NJIMA' in self.prodaja.columns
        self.prodaja_dict = {}
        for _,r in self.prodaja.iterrows():
            key=(r['ID KOMITENTA'],r['id artikla'],r['Godina'],r['Mesec'])
            pm = r['PROMET KA NJIMA'] if self.has_promet else 0
            self.prodaja_dict[key] = (r.get('Prodata Kolicina',r.get('Kolicina',0)), r.get('Lager',0), pm if not pd.isna(pm) else 0)
        self.hist_dict={}; self.hist_total_dict={}; self.hist_months_per_art={}
        if self.has_history:
            ha = self.hist_df.groupby(['ID KOMITENTA','id artikla'])['Prodata Kolicina'].agg(['sum','mean']).reset_index()
            for _,r in ha.iterrows():
                self.hist_dict[(int(r['ID KOMITENTA']),int(r['id artikla']))] = float(r['mean'])
                self.hist_total_dict[(int(r['ID KOMITENTA']),int(r['id artikla']))] = int(r['sum'])
            for ida in self.hist_df['id artikla'].unique():
                sub=self.hist_df[self.hist_df['id artikla']==ida]
                self.hist_months_per_art[int(ida)]=sub[['Godina','Mesec']].drop_duplicates().shape[0]
            self.log(f"Istorijski prosek za {len(self.hist_dict)} kombinacija")
        self.recent_months_per_art={}
        for ida in self.prodaja['id artikla'].unique():
            sub=self.prodaja[self.prodaja['id artikla']==ida]
            self.recent_months_per_art[int(ida)]=sub[['Godina','Mesec']].drop_duplicates().shape[0]
        self.total_months_per_art={}
        all_arts=set([int(x) for x in self.prodaja['id artikla'].unique()])
        if self.has_history: all_arts|=set([int(x) for x in self.hist_df['id artikla'].unique()])
        for ida in all_arts:
            self.total_months_per_art[ida]=self.hist_months_per_art.get(ida,0)+self.recent_months_per_art.get(ida,0)
        self.povrat_total={}
        if len(self.povrat_df)>0:
            ic=[c for c in self.povrat_df.columns if 'id' in c.lower() and 'artikl' in c.lower()]
            mc=[c for c in self.povrat_df.columns if 'mesec' in c.lower()]
            gc=[c for c in self.povrat_df.columns if 'godin' in c.lower()]
            kc=[c for c in self.povrat_df.columns if 'koli' in c.lower()]
            if ic and mc and gc and kc:
                for _,r in self.povrat_df.iterrows():
                    key=(r[ic[0]],r[gc[0]],r[mc[0]]); self.povrat_total[key]=self.povrat_total.get(key,0)+r[kc[0]]
        self.trenutni_dict={}
        if len(self.trenutni)>0:
            ikc=[c for c in self.trenutni.columns if 'komitent' in c.lower()]
            iac=[c for c in self.trenutni.columns if 'artikl' in c.lower() and 'id' in c.lower()]
            lc=[c for c in self.trenutni.columns if 'lager' in c.lower()]
            if ikc and iac and lc:
                for _,r in self.trenutni.iterrows():
                    k,a=r[ikc[0]],r[iac[0]]
                    if pd.notna(k) and pd.notna(a): self.trenutni_dict[(int(k),int(a))]=int(r[lc[0]]) if pd.notna(r[lc[0]]) else 0
        self.profit_per_unit = {}
        self.price_info = {}
        if self.has_prices:
            for ida in self.prodaja['id artikla'].unique():
                sub = self.prodaja[self.prodaja['id artikla']==ida].iloc[0]
                red, akc, fin, nab = sub['Redovna cena'], sub['Akcijska cena'], sub['Finalna cena'], sub['Nabavna vrednost']
                ppu_fin = fin/1.2/1.2 - nab
                ppu_red = red/1.2/1.2 - nab
                self.profit_per_unit[int(ida)] = ppu_fin
                self.price_info[int(ida)] = {'redovna': red, 'akcijska': akc, 'finalna': fin, 'nabavna': nab, 'profit_akcija': ppu_fin, 'profit_redovna': ppu_red}
        self.log(f"Kombinacija: {len(self.all_keys)}")
    def _compute_povrat(self):
        self.final_povrat={}; self.final_korekcija={}
        if not self.has_promet or not self.povrat_total: return
        implied={}
        for _,k in self.all_keys.iterrows():
            idk,ida=k['ID KOMITENTA'],k['id artikla']; poc=self.startni_dict.get((idk,ida),0)
            for god,mes in self.meseci_order:
                pv,lv,tv=self.prodaja_dict.get((idk,ida,god,mes),(0,0,0)); lv=lv if not pd.isna(lv) else 0
                implied[(idk,ida,god,mes)]=poc+tv-pv-lv; poc=lv
        all_art=set(list(self.prodaja['id artikla'].unique())+list(self.startni['id artikla'].unique()))
        for god,mes in self.meseci_order:
            for ida in all_art:
                ap=self.povrat_total.get((ida,god,mes),0); pi={}; ni={}
                for _,k in self.all_keys[self.all_keys['id artikla']==ida].iterrows():
                    i2=k['ID KOMITENTA']; im=implied.get((i2,ida,god,mes),0)
                    if im>0: pi[i2]=im
                    elif im<0: ni[i2]=im
                tp=sum(pi.values())
                if ap>0 and tp>0:
                    raw={i:ap*(v/tp) for i,v in pi.items()}; fl={i:int(v) for i,v in raw.items()}
                    d=ap-sum(fl.values()); rem={i:raw[i]-fl[i] for i in raw}
                    for j,i in enumerate(sorted(rem,key=rem.get,reverse=True)):
                        if j<int(d): fl[i]+=1
                    for i,pv2 in fl.items():
                        self.final_povrat[(i,ida,god,mes)]=pv2; self.final_korekcija[(i,ida,god,mes)]=pi[i]-pv2
                elif ap==0:
                    for i,v in pi.items(): self.final_korekcija[(i,ida,god,mes)]=v
                for i,v in ni.items(): self.final_korekcija[(i,ida,god,mes)]=self.final_korekcija.get((i,ida,god,mes),0)+v
    def _build_monthly(self):
        rows=[]
        for _,k in self.all_keys.iterrows():
            idk,ida=k['ID KOMITENTA'],k['id artikla']; poc=self.startni_dict.get((idk,ida),0)
            row={'ID KOMITENTA':idk,'id artikla':ida,'Naziv artikla':k['Naziv artikla'],'Grupa':k['Grupa']}
            row['Total_JanAvg']=self.hist_total_dict.get((idk,ida),0)
            for i,(god,mes) in enumerate(self.meseci_order):
                lb=self.mesec_labels[i]; pv,lv,tv=self.prodaja_dict.get((idk,ida,god,mes),(0,0,0))
                lv=lv if not pd.isna(lv) else 0; tv=tv if not pd.isna(tv) else 0
                row[f'{lb}_Pocetno']=poc; row[f'{lb}_Promet']=tv; row[f'{lb}_Prodaja']=pv
                row[f'{lb}_Povrat']=self.final_povrat.get((idk,ida,god,mes),0)
                row[f'{lb}_Korekcija']=self.final_korekcija.get((idk,ida,god,mes),0); poc=lv
            rows.append(row)
        self.df_monthly=pd.DataFrame(rows)
    def _predict_all(self):
        analysis=[]
        for _,k in self.all_keys.iterrows():
            idk,ida=k['ID KOMITENTA'],k['id artikla']; poc=self.startni_dict.get((idk,ida),0)
            sales,oos,pocs,end_lagers,promets=[],[],[],[],[]
            for god,mes in self.meseci_order:
                pv,lv,tv=self.prodaja_dict.get((idk,ida,god,mes),(0,0,0))
                lv=lv if not pd.isna(lv) else 0; tv=tv if not pd.isna(tv) else 0
                sales.append(pv); oos.append(1 if poc==0 else 0); pocs.append(poc)
                end_lagers.append(lv); promets.append(tv); poc=lv
            ha=self.hist_dict.get((idk,ida),0)
            lager_danas=self.trenutni_dict.get((idk,ida),0)
            analysis.append({'idk':idk,'ida':ida,'sales':np.array(sales,dtype=float),'oos':np.array(oos),
                'poc':np.array(pocs,dtype=float),'ha':ha,'lager_danas':lager_danas,
                'end_lagers':np.array(end_lagers,dtype=float),'promets':np.array(promets,dtype=float)})
        preds={}
        for it in analysis:
            s,o,p=it['sales'],it['oos'],it['poc']; n=len(s); ha=it['ha']
            lager_danas=it['lager_danas']
            el=it['end_lagers']; tv=it['promets']
            constrained = np.zeros(n, dtype=bool)
            for m in range(n):
                if p[m]==0 and tv[m]==0: constrained[m] = True
                elif el[m]==0 and s[m]>0: constrained[m] = True
                elif p[m]==0 and tv[m]>0 and el[m]==0: constrained[m] = True
            normal_mask = ~constrained & (p > 0)
            normal_sales = s[normal_mask]
            normal_with_sales = normal_sales[normal_sales > 0]
            if len(normal_with_sales) > 0: an = normal_with_sales.mean()
            elif len(normal_sales) > 0: an = normal_sales.mean()
            else: an = 0
            if an > 0:
                adj = s.copy().astype(float)
                for m in range(n):
                    if constrained[m]:
                        if p[m]==0 and tv[m]==0: adj[m] = an
                        elif el[m]==0 and s[m]>0: adj[m] = max(an, s[m])
                        else: adj[m] = an
                    elif p[m]>0 and p[m]<an*0.5: adj[m] = 0.5*s[m] + 0.5*an
            elif ha>0: adj=np.full(n,ha)
            else: adj=s.copy().astype(float)
            if n>=2:
                lev=adj[0]; tr=(adj[-1]-adj[0])/max(n-1,1)
                for i in range(1,n):
                    nl=self.alpha*adj[i]+(1-self.alpha)*(lev+tr); nt=self.beta*(nl-lev)+(1-self.beta)*tr; lev,tr=nl,nt
                holt=lev+tr
            else: holt=adj[0]
            w=WMA_WEIGHTS[-n:] if n<=5 else WMA_WEIGHTS; w=w/w.sum()
            wma=np.dot(adj[-len(w):],w) if n>=3 else adj.mean()
            comb = 0.4 * min(holt, wma) + 0.6 * max(holt, wma)
            ma=adj.mean()
            if ma>0 and n>=3: comb*=(1+min((np.std(adj)/ma)*0.4,0.7))
            if ha>0 and comb>0: comb=(1-HIST_WEIGHT)*comb+HIST_WEIGHT*ha
            elif ha>0 and comb==0 and s.sum()==0: comb=ha*0.20
            has_recent_sales = (s[-2:].sum() > 0) if n >= 2 else (s.sum() > 0)
            if lager_danas <= 2 and has_recent_sales:
                stocked_sales = [s[i] for i in range(n) if p[i] > 0]
                avg_when_stocked = np.mean(stocked_sales) if stocked_sales else 0
                if avg_when_stocked > 0 and comb < avg_when_stocked: comb = avg_when_stocked
            if ma > 5 and comb < ma: comb = ma
            avg_5m_raw = float(adj[-5:].mean()) if n >= 5 else float(adj.mean())
            ht=self.hist_total_dict.get((it['idk'],it['ida']),0)
            rt=float(s.sum()); tm=self.total_months_per_art.get(it['ida'],n)
            full_avg=(ht+rt)/max(tm,1)
            if comb < full_avg and comb > 0:
                if n >= 5: declining = all(adj[i] <= adj[i-1] for i in range(n-4, n))
                elif n >= 3: declining = all(adj[i] <= adj[i-1] for i in range(1, n))
                else: declining = (n >= 2 and adj[-1] <= adj[-2])
                if not declining: comb = full_avg
            if comb <= 0:
                last5 = s[-5:] if n >= 5 else s
                if last5.sum() > 0:
                    comb = 1.0
                    if s[-1] > 1: comb = s[-1]
            preds[(it['idk'],it['ida'])]=(max(0,comb),full_avg,avg_5m_raw)
        items=[{'k':k,'p':v[0],'a':v[1],'avg5':v[2]} for k,v in preds.items()]; df_p=pd.DataFrame(items)
        df_p['pr']=df_p['p'].apply(lambda x: round(x))
        df_p['ar']=df_p['a'].apply(lambda x: round(x))
        self.pred_dict={r['k']:(int(r['pr']),int(r['ar']),int(r['pr']-r['ar']),r['avg5']) for _,r in df_p.iterrows()}
        self.log(f"Predikcija: {sum(v[0] for v in self.pred_dict.values())} kom")
    def _merge_lager(self):
        for _,k in self.all_keys.iterrows():
            idk,ida=k['ID KOMITENTA'],k['id artikla']; pred,avg,razl,avg5m=self.pred_dict.get((idk,ida),(0,0,0,0))
            lager=self.trenutni_dict.get((idk,ida),None)
            idx=self.df_monthly[(self.df_monthly['ID KOMITENTA']==idk)&(self.df_monthly['id artikla']==ida)].index
            if len(idx)>0:
                ix=idx[0]; self.df_monthly.loc[ix,'Predikcija']=pred; self.df_monthly.loc[ix,'Prosek']=avg; self.df_monthly.loc[ix,'Razlika']=razl
                self.df_monthly.loc[ix,'Avg5m']=avg5m
                if lager is not None: self.df_monthly.loc[ix,'Lager_danas']=lager
                else: self.df_monthly.loc[ix,'Lager_danas']=0
        for col in ['Predikcija','Prosek','Razlika','Lager_danas']:
            if col not in self.df_monthly.columns: self.df_monthly[col]=0
            self.df_monthly[col]=self.df_monthly[col].fillna(0).astype(int)
        if 'Avg5m' not in self.df_monthly.columns: self.df_monthly['Avg5m']=0
        self.df_monthly['Avg5m']=self.df_monthly['Avg5m'].fillna(0)
    def _compute_orders(self):
        self.df_result=self.df_monthly.copy()
        def p1(row):
            if row['ID KOMITENTA'] in self.excluded: return 0
            pred_o=int(round(int(row['Predikcija'])*self.meseci))
            return max(pred_o-int(row['Lager_danas']),0)
        def p2(row):
            if row['ID KOMITENTA'] in self.excluded: return 0
            pred=int(row['Predikcija']); pred_o=int(round(pred*self.meseci)); lager=int(row['Lager_danas']); prosek=int(row['Prosek'])
            osnova=max(pred_o-lager,0)
            if self.min_lager is not None and lager < self.min_lager and pred > 0:
                dopuna = max(self.min_lager - lager, osnova)
            else:
                dopuna = osnova
            return dopuna
        self.df_result['Porudzbina_1']=self.df_result.apply(p1,axis=1).astype(int)
        self.df_result['Porudzbina_2']=self.df_result.apply(p2,axis=1).astype(int)
        last_label = self.mesec_labels[-1]
        def extra_buffer(prodaja_poslednji):
            if prodaja_poslednji <= 0: return 0
            elif prodaja_poslednji <= 5: return 2
            elif prodaja_poslednji <= 10: return 3
            elif prodaja_poslednji <= 15: return 4
            else: return 5
        def finalna_provera(row):
            if row['ID KOMITENTA'] in self.excluded: return int(row['Porudzbina_2'])
            p2_val = int(row['Porudzbina_2'])
            lager = int(row['Lager_danas'])
            prodaja_poslednji = int(row.get(f'{last_label}_Prodaja', 0))
            if (p2_val + lager) <= prodaja_poslednji:
                dodatak = extra_buffer(prodaja_poslednji)
                return p2_val + dodatak
            return p2_val
        self.df_result['Porudzbina_2'] = self.df_result.apply(finalna_provera, axis=1).astype(int)
        n_korigovano = (self.df_result['Porudzbina_2'] > self.df_result.apply(p2, axis=1)).sum()
        self.log(f"Finalna provera P2: {n_korigovano} kombinacija korigovano (porudzbina+lager <= prodaja poslednjeg meseca)")
        if self.min_per_artikal is not None and self.min_per_artikal > 1:
            mask_p2 = (
                (self.df_result['Porudzbina_2'] > 0) &
                (self.df_result['Porudzbina_2'] < self.min_per_artikal) &
                (~self.df_result['ID KOMITENTA'].isin(self.excluded))
            )
            n_podignuto_p2 = int(mask_p2.sum())
            self.df_result.loc[mask_p2, 'Porudzbina_2'] = self.min_per_artikal
            mask_p1 = (
                (self.df_result['Porudzbina_1'] > 0) &
                (self.df_result['Porudzbina_1'] < self.min_per_artikal) &
                (~self.df_result['ID KOMITENTA'].isin(self.excluded))
            )
            n_podignuto_p1 = int(mask_p1.sum())
            self.df_result.loc[mask_p1, 'Porudzbina_1'] = self.min_per_artikal
            if n_podignuto_p2 > 0 or n_podignuto_p1 > 0:
                self.log(f"Min po artiklu ({self.min_per_artikal} kom): P1={n_podignuto_p1}, P2={n_podignuto_p2} stavki podignuto na minimum")
        # Maksimum po komadu (po stavci): ograniči porudžbinu po artiklu na zadati maksimum
        if getattr(self, "max_per_artikal", None) is not None and self.max_per_artikal > 0:
            mx = int(self.max_per_artikal)
            cap2 = self.df_result['Porudzbina_2'] > mx
            cap1 = self.df_result['Porudzbina_1'] > mx
            n_cap2 = int(cap2.sum()); n_cap1 = int(cap1.sum())
            self.df_result.loc[cap2, 'Porudzbina_2'] = mx
            self.df_result.loc[cap1, 'Porudzbina_1'] = mx
            if n_cap2 > 0 or n_cap1 > 0:
                self.log(f"Max po artiklu ({mx} kom): P1={n_cap1}, P2={n_cap2} stavki ograničeno na maksimum")
    def _apply_min_order(self):
        self.adjustments = []
        if self.min_order is None or self.min_order <= 0: return
        grp = self.df_result.groupby('ID KOMITENTA')['Porudzbina_2'].sum()
        ima_nesto = grp[grp > 0]
        granica = self.min_order / 2
        premali = ima_nesto[ima_nesto < granica].index
        dopuni = ima_nesto[(ima_nesto >= granica) & (ima_nesto < self.min_order)].index
        mask_gasi = self.df_result['ID KOMITENTA'].isin(premali)
        n_gasi = len(premali)
        self.df_result.loc[mask_gasi, 'Porudzbina_2'] = 0
        n_dopuni = 0
        for komt_id in dopuni:
            mask_obj = (self.df_result['ID KOMITENTA'] == komt_id) & (self.df_result['Porudzbina_2'] > 0)
            ukupno = int(self.df_result.loc[self.df_result['ID KOMITENTA'] == komt_id, 'Porudzbina_2'].sum())
            nedostaje = self.min_order - ukupno
            if nedostaje <= 0 or not mask_obj.any(): continue
            idx_max = self.df_result.loc[mask_obj, 'Porudzbina_2'].idxmax()
            self.df_result.at[idx_max, 'Porudzbina_2'] += nedostaje
            n_dopuni += 1
        if n_gasi > 0:
            self.log(f"Min order ({self.min_order} kom): {n_gasi} objekata imalo premalo komada ukupno — postavljeno na 0")
        if n_dopuni > 0:
            self.log(f"Min order ({self.min_order} kom): {n_dopuni} objekata dopunjeno do minimuma {self.min_order} kom")
    def _compute_analytics(self):
        if not self.has_prices:
            self.df_oos = pd.DataFrame()
            self.df_profit_obj = pd.DataFrame()
            self.df_promo = pd.DataFrame()
            self.analitika_labels = []
            return
        df = self.df_result; ml = self.mesec_labels
        if self.analitika_meseci and len(self.analitika_meseci) > 0:
            a_meseci = self.analitika_meseci
        else:
            a_meseci = self.meseci_order
        a_indices = []
        for i, (g, m) in enumerate(self.meseci_order):
            for ag, am in a_meseci:
                if int(g) == int(ag) and int(m) == int(am):
                    a_indices.append(i); break
        if not a_indices:
            a_indices = list(range(len(self.meseci_order)))
        a_labels = [ml[i] for i in a_indices]
        a_meseci_order = [self.meseci_order[i] for i in a_indices]
        n_a = len(a_indices)
        self.analitika_labels = a_labels
        self.log(f"Analitika period: {', '.join(a_labels)} ({n_a} meseci)")
        a_set = set((int(g), int(m)) for g, m in a_meseci_order)
        prodaja_a = self.prodaja[self.prodaja.apply(lambda r: (int(r['Godina']), int(r['Mesec'])) in a_set, axis=1)]
        ppu_mesec = {}
        if self.has_prices:
            for (ida_v, god_v, mes_v), grp in self.prodaja.groupby(['id artikla','Godina','Mesec']):
                kol = grp['Prodata Kolicina'].sum()
                if kol > 0:
                    ppu_mesec[(int(ida_v), int(god_v), int(mes_v))] = grp['Profit'].sum() / kol
                else:
                    r0 = grp.iloc[0]
                    ppu_mesec[(int(ida_v), int(god_v), int(mes_v))] = r0['Finalna cena'] / 1.2 / 1.2 - r0['Nabavna vrednost']
        def get_ppu(ida_v, god_v, mes_v):
            key = (int(ida_v), int(god_v), int(mes_v))
            if key in ppu_mesec:
                return ppu_mesec[key]
            art_keys = sorted([k for k in ppu_mesec if k[0] == int(ida_v)], key=lambda x: (x[1], x[2]))
            if not art_keys:
                return self.profit_per_unit.get(int(ida_v), 0)
            target = int(god_v) * 12 + int(mes_v)
            best = min(art_keys, key=lambda x: abs(x[1] * 12 + x[2] - target))
            return ppu_mesec[best]
        oos_rows = []
        for _, k in self.all_keys.iterrows():
            idk, ida = k['ID KOMITENTA'], k['id artikla']
            poc = self.startni_dict.get((idk, ida), 0)
            month_sales = []
            month_poc = []
            month_ulaz = []
            month_kraj = []
            month_gm = []
            for i, (god, mes) in enumerate(self.meseci_order):
                lb = ml[i]
                pv_arr = df[(df['ID KOMITENTA']==idk)&(df['id artikla']==ida)][f'{lb}_Prodaja'].values
                pv = int(pv_arr[0]) if len(pv_arr) > 0 else 0
                tv_arr = df[(df['ID KOMITENTA']==idk)&(df['id artikla']==ida)][f'{lb}_Promet'].values
                tv = int(tv_arr[0]) if len(tv_arr) > 0 else 0
                lv_col = self.prodaja_dict.get((idk, ida, god, mes), (0, 0, 0))
                kraj = lv_col[1] if not pd.isna(lv_col[1]) else 0
                if i in a_indices:
                    month_sales.append(pv)
                    month_poc.append(poc)
                    month_ulaz.append(tv)
                    month_kraj.append(kraj)
                    month_gm.append((god, mes))
                poc = kraj
            month_constrained_type = []
            for j in range(len(month_sales)):
                p = month_poc[j]
                u = month_ulaz[j]
                kr = month_kraj[j]
                pr = month_sales[j]
                if p == 0 and u == 0:
                    month_constrained_type.append(1)
                elif p == 0 and u > 0 and kr == 0:
                    month_constrained_type.append(3)
                elif kr == 0 and pr > 0:
                    month_constrained_type.append(2)
                else:
                    month_constrained_type.append(0)
            normal_sales = [month_sales[j] for j in range(len(month_sales))
                            if month_constrained_type[j] == 0 and month_sales[j] > 0]
            avg_stocked = np.mean(normal_sales) if normal_sales else 0
            month_oos_kom = []
            month_oos_flag = []
            total_lost_kom = 0
            for j in range(len(month_sales)):
                t = month_constrained_type[j]
                pr = month_sales[j]
                if avg_stocked == 0:
                    izgub_kom = 0
                elif t == 1:
                    izgub_kom = avg_stocked
                elif t == 2 or t == 3:
                    izgub_kom = max(0, avg_stocked - pr)
                else:
                    izgub_kom = 0
                month_oos_kom.append(izgub_kom)
                month_oos_flag.append(1 if izgub_kom >= 0.5 else 0)
                total_lost_kom += izgub_kom
            oos_count = sum(month_oos_flag)
            if total_lost_kom > 0 and avg_stocked > 0:
                row = {
                    'ID KOMITENTA': idk, 'id artikla': ida,
                    'Naziv artikla': k['Naziv artikla'], 'Grupa': k['Grupa'],
                    'Prosek_kad_ima': round(avg_stocked, 1),
                    'Lager_danas': self.trenutni_dict.get((idk, ida), 0)
                }
                total_lost_rsd = 0
                for j in range(len(month_sales)):
                    god_j, mes_j = month_gm[j]
                    lb_j = a_labels[j]
                    if month_oos_kom[j] > 0:
                        ppu_j = get_ppu(ida, god_j, mes_j)
                        izgub_rsd = round(month_oos_kom[j] * ppu_j, 0)
                        row[f'OOS_{lb_j}'] = round(month_oos_kom[j], 1)
                        row[f'Izgub_{lb_j}'] = izgub_rsd
                        total_lost_rsd += izgub_rsd
                    else:
                        row[f'OOS_{lb_j}'] = 0
                        row[f'Izgub_{lb_j}'] = 0
                row['OOS_meseci'] = oos_count
                row['Izgubljeni_profit'] = round(total_lost_rsd, 0)
                oos_rows.append(row)
        self.df_oos = pd.DataFrame(oos_rows)
        if len(self.df_oos) > 0:
            self.df_oos = self.df_oos.sort_values('Izgubljeni_profit', ascending=False)
            self.log(f"OOS analiza (nova logika - kom izgubljeno): {len(self.df_oos)} kombinacija, izgubljeno {self.df_oos['Izgubljeni_profit'].sum():,.0f} RSD")
        trosak_mes_po_obj = self.trosak_po_objektu / max(n_a, 1) if self.trosak_po_objektu > 0 else 0
        profit_rows = []
        for idk in self.prodaja['ID KOMITENTA'].unique():
            sub = prodaja_a[prodaja_a['ID KOMITENTA'] == idk]
            total_prod = int(sub['Prodata Kolicina'].sum())
            total_profit = sub['Profit'].sum()
            n_art = self.all_keys[self.all_keys['ID KOMITENTA'] == idk]['id artikla'].nunique()
            mes_data = {}
            for _, r in sub.iterrows():
                key = f"{int(r['Godina'])}/{int(r['Mesec'])}"
                mes_data[key] = mes_data.get(key, 0) + r['Profit']
            mes_data_neto = {k: v - trosak_mes_po_obj for k, v in mes_data.items()}
            oos_sub = self.df_oos[self.df_oos['ID KOMITENTA'] == idk] if len(self.df_oos) > 0 else pd.DataFrame()
            lost = oos_sub['Izgubljeni_profit'].sum() if len(oos_sub) > 0 else 0
            trosak_total = self.trosak_po_objektu
            neto = total_profit - trosak_total
            row_dict = {
                'ID KOMITENTA': int(idk), 'Artikala': n_art,
                'Prodato_kom': total_prod, 'Bruto_profit': round(total_profit, 0),
                'Trosak_mkt': round(trosak_total, 0),
                'Neto_profit': round(neto, 0),
                'Izgubljeno_OOS': round(lost, 0),
                'Potencijalni_profit': round(neto + lost, 0),
            }
            for j in range(n_a):
                key_j = f"{int(a_meseci_order[j][0])}/{int(a_meseci_order[j][1])}"
                row_dict[f'Neto_{a_labels[j]}'] = round(mes_data_neto.get(key_j, -trosak_mes_po_obj), 0)
                row_dict[f'Bruto_{a_labels[j]}'] = round(mes_data.get(key_j, 0), 0)
            profit_rows.append(row_dict)
        self.trosak_mes_po_obj = trosak_mes_po_obj
        self.df_profit_obj = pd.DataFrame(profit_rows).sort_values('Neto_profit', ascending=True)
        promo_rows = []
        for ida in self.prodaja['id artikla'].unique():
            pi = self.price_info.get(int(ida), {})
            if not pi: continue
            sub = prodaja_a[prodaja_a['id artikla'] == ida]
            total_prod = int(sub['Prodata Kolicina'].sum())
            if total_prod == 0: continue
            profit_akcija = sub['Profit'].sum()
            profit_redovna = pi['profit_redovna'] * total_prod
            razlika = profit_redovna - profit_akcija
            prihod_akcija = (sub['Finalna cena'] * sub['Prodata Kolicina']).sum()
            prihod_redovna = (sub['Redovna cena'] * sub['Prodata Kolicina']).sum()
            first_a_idx = a_indices[0]
            if first_a_idx == 0:
                start_lager = self.startni[self.startni['id artikla']==ida]['Kolicina'].sum() if 'Kolicina' in self.startni.columns else 0
            else:
                prev_god, prev_mes = self.meseci_order[first_a_idx - 1]
                prev_sub = self.prodaja[(self.prodaja['id artikla']==ida) & (self.prodaja['Godina']==prev_god) & (self.prodaja['Mesec']==prev_mes)]
                start_lager = prev_sub['Lager'].sum() if len(prev_sub) > 0 else 0
                start_lager = start_lager if not pd.isna(start_lager) else 0
            lageri = [start_lager]
            for god, mes in a_meseci_order:
                msub = self.prodaja[(self.prodaja['id artikla']==ida) & (self.prodaja['Godina']==god) & (self.prodaja['Mesec']==mes)]
                lager_kraj = msub['Lager'].sum() if len(msub) > 0 else 0
                lageri.append(lager_kraj if not pd.isna(lager_kraj) else 0)
            avg_lager = np.mean(lageri)
            obrt = total_prod / avg_lager if avg_lager > 0 else 0
            dani_pokrivanja = (avg_lager / (total_prod / (n_a * 30))) if total_prod > 0 else 999
            n_obj_aktiv = sub[sub['Prodata Kolicina']>0]['ID KOMITENTA'].nunique()
            n_obj_total = sub['ID KOMITENTA'].nunique()
            prod_po_obj = total_prod / n_obj_aktiv if n_obj_aktiv > 0 else 0
            mes_prod = {}
            for _, r in sub.iterrows():
                key = f"{int(r['Godina'])}/{int(r['Mesec'])}"
                mes_prod[key] = mes_prod.get(key, 0) + int(r['Prodata Kolicina'])
            promo_rows.append({
                'id artikla': int(ida),
                'Naziv': sub.iloc[0]['Naziv artikla'],
                'Grupa': sub.iloc[0]['Grupa'],
                'Redovna': pi['redovna'], 'Akcijska': pi['akcijska'],
                'Popust_%': round((1 - pi['akcijska']/pi['redovna'])*100, 1),
                'Prodato_kom': total_prod,
                'Prihod_akcija': round(prihod_akcija, 0),
                'Prihod_redovna': round(prihod_redovna, 0),
                'Profit_akcija': round(profit_akcija, 0),
                'Profit_da_je_redovna': round(profit_redovna, 0),
                'Cena_akcije': round(razlika, 0),
                'Avg_lager': round(avg_lager, 0),
                'Obrt_x': round(obrt, 1),
                'Dani_pokrivanja': round(dani_pokrivanja, 0),
                'Obj_aktivnih': n_obj_aktiv,
                'Obj_ukupno': n_obj_total,
                'Prod_po_obj': round(prod_po_obj, 1),
                **{f'Prod_{a_labels[j]}': mes_prod.get(f"{int(a_meseci_order[j][0])}/{int(a_meseci_order[j][1])}", 0) for j in range(n_a)}
            })
        self.df_promo = pd.DataFrame(promo_rows).sort_values('Obrt_x', ascending=False)

def create_excel(engine, ukljuci_model=True):
    df=engine.df_result; ml=engine.mesec_labels; wb=Workbook()
    hf=PatternFill('solid',fgColor='2F5496'); hfn=Font(bold=True,color='FFFFFF',name='Arial',size=10)
    sfnt=Font(bold=True,name='Arial',size=9); dfn=Font(name='Arial',size=9)
    tb=Border(left=Side('thin','B4C6E7'),right=Side('thin','B4C6E7'),top=Side('thin','B4C6E7'),bottom=Side('thin','B4C6E7'))
    ca=Alignment(horizontal='center',vertical='center'); caw=Alignment(horizontal='center',vertical='center',wrap_text=True)
    sf_poc=PatternFill('solid',fgColor='D6E4F0'); sf_prom=PatternFill('solid',fgColor='C6EFCE')
    sf_prod=PatternFill('solid',fgColor='FFF2CC'); sf_pov=PatternFill('solid',fgColor='FCE4EC')
    sf_kor=PatternFill('solid',fgColor='E8E8E8'); sf_pred=PatternFill('solid',fgColor='D5A6E6')
    sf_avg=PatternFill('solid',fgColor='B4D7E8'); sf_razl=PatternFill('solid',fgColor='FFD699')
    sf_lager=PatternFill('solid',fgColor='DAEEF3'); sf_p1=PatternFill('solid',fgColor='92D050')
    sf_p2=PatternFill('solid',fgColor='00B050'); pred_hdr=PatternFill('solid',fgColor='7030A0')
    ord_hdr=PatternFill('solid',fgColor='375623'); sf_hist=PatternFill('solid',fgColor='E2D5F1')
    nf_money='#,##0'
    SC=5; sub_h=['Pocetno stanje','Promet (ulaz)','Prodaja','Povrat','Korekcija']
    sub_f=[sf_poc,sf_prom,sf_prod,sf_pov,sf_kor]; col_suf=['_Pocetno','_Promet','_Prodaja','_Povrat','_Korekcija']
    ws1=wb.active; ws1.title="Pregled po objektima"
    for c,t in enumerate(['ID Komitenta','ID Artikla','Naziv Artikla','Grupa'],1):
        cell=ws1.cell(1,c,t); cell.font=hfn; cell.fill=hf; cell.alignment=ca; cell.border=tb
        ws1.merge_cells(start_row=1,end_row=2,start_column=c,end_column=c)
    hist_col=5; month_start=5
    if engine.has_history:
        cell=ws1.cell(1,hist_col,'Jan-Avg 2025'); cell.font=hfn; cell.fill=PatternFill('solid',fgColor='6B3FA0')
        cell.alignment=ca; cell.border=tb
        ws1.merge_cells(start_row=1,end_row=1,start_column=hist_col,end_column=hist_col)
        c2=ws1.cell(2,hist_col,'Total prodaja'); c2.font=sfnt; c2.fill=sf_hist; c2.alignment=caw; c2.border=tb
        month_start=6
    for i,label in enumerate(ml):
        sc=month_start+i*SC
        ws1.merge_cells(start_row=1,end_row=1,start_column=sc,end_column=sc+SC-1)
        cell=ws1.cell(1,sc,label); cell.font=hfn; cell.fill=hf; cell.alignment=ca
        for cc in range(sc,sc+SC): ws1.cell(1,cc).border=tb; ws1.cell(1,cc).fill=hf
        for j,(sh,sfill) in enumerate(zip(sub_h,sub_f)):
            cell=ws1.cell(2,sc+j,sh); cell.font=sfnt; cell.fill=sfill; cell.border=tb; cell.alignment=caw
    ps=month_start+len(ml)*SC
    ws1.merge_cells(start_row=1,end_row=1,start_column=ps,end_column=ps+2)
    cell=ws1.cell(1,ps,f'{engine.pred_label} - PREDIKCIJA'); cell.font=hfn; cell.fill=pred_hdr; cell.alignment=ca
    for cc in range(ps,ps+3): ws1.cell(1,cc).border=tb; ws1.cell(1,cc).fill=pred_hdr
    for j,(sh,sfill) in enumerate(zip(['Predikcija','Prosek (svi mes.)','Razlika'],[sf_pred,sf_avg,sf_razl])):
        cell=ws1.cell(2,ps+j,sh); cell.font=sfnt; cell.fill=sfill; cell.border=tb; cell.alignment=caw
    os_c=ps+3
    ws1.merge_cells(start_row=1,end_row=1,start_column=os_c,end_column=os_c+2)
    cell=ws1.cell(1,os_c,f'PORUDZBINA - {engine.order_label}'); cell.font=hfn; cell.fill=ord_hdr; cell.alignment=ca
    for cc in range(os_c,os_c+3): ws1.cell(1,cc).border=tb; ws1.cell(1,cc).fill=ord_hdr
    ll="Lager danas"
    if len(engine.trenutni)>0:
        dc=[c for c in engine.trenutni.columns if 'dan' in c.lower()]
        if dc:
            try: d=pd.to_datetime(engine.trenutni[dc[0]].iloc[0]); ll=f"Lager na dan\n{d.strftime('%d.%m.%Y')}"
            except: pass
    for j,(sh,sfill) in enumerate(zip([ll,'Porudzbina\n(osnovna)',f'Porudzbina\n(min. {engine.min_lager} na stanju)'],[sf_lager,sf_p1,sf_p2])):
        cell=ws1.cell(2,os_c+j,sh); cell.font=sfnt; cell.fill=sfill; cell.border=tb; cell.alignment=caw
    for idx,row in df.iterrows():
        r=idx+3
        for c2,col in enumerate(['ID KOMITENTA','id artikla','Naziv artikla','Grupa'],1):
            ws1.cell(r,c2,row[col]).font=dfn; ws1.cell(r,c2).border=tb
        if engine.has_history:
            v=int(row.get('Total_JanAvg',0)); cell=ws1.cell(r,hist_col,v); cell.font=dfn
            cell.alignment=ca; cell.border=tb
            if v>0: cell.fill=PatternFill('solid',fgColor='F3EAFA')
        for i,label in enumerate(ml):
            cb=month_start+i*SC
            for j,suf in enumerate(col_suf):
                cn=f'{label}{suf}'; v=row.get(cn,0)
                cell=ws1.cell(r,cb+j,int(v) if not pd.isna(v) else 0); cell.font=dfn; cell.alignment=ca; cell.border=tb
        for j,cn in enumerate(['Predikcija','Prosek','Razlika']):
            v=int(row.get(cn,0)); cell=ws1.cell(r,ps+j,v); cell.alignment=ca; cell.border=tb
            if cn=='Razlika':
                if v>0: cell.font=Font(name='Arial',size=9,color='006100',bold=True)
                elif v<0: cell.font=Font(name='Arial',size=9,color='9C0006',bold=True)
                else: cell.font=dfn
            else: cell.font=dfn
        for j,cn in enumerate(['Lager_danas','Porudzbina_1','Porudzbina_2']):
            v=int(row.get(cn,0)); cell=ws1.cell(r,os_c+j,v); cell.alignment=ca; cell.border=tb
            if cn!='Lager_danas' and v>0: cell.font=Font(name='Arial',size=9,bold=True,color='375623')
            else: cell.font=dfn
    ws1.column_dimensions['A'].width=14; ws1.column_dimensions['B'].width=11; ws1.column_dimensions['C'].width=50; ws1.column_dimensions['D'].width=12
    if engine.has_history: ws1.column_dimensions[get_column_letter(hist_col)].width=14
    for i in range(len(ml)):
        for j in range(SC): ws1.column_dimensions[get_column_letter(month_start+i*SC+j)].width=14
    for j in range(3): ws1.column_dimensions[get_column_letter(ps+j)].width=14
    for j in range(3): ws1.column_dimensions[get_column_letter(os_c+j)].width=18
    ws1.freeze_panes=f'{get_column_letter(month_start)}3'
    ws1.auto_filter.ref=f"A2:{get_column_letter(ws1.max_column)}{ws1.max_row}"
    ws2=wb.create_sheet("Totali po mesecima")
    for c,h in enumerate(['Mesec','Promet (ulaz)','Prodaja','Stvarni povrat','Korekcija','Neto (Promet-Povrat)'],1):
        cell=ws2.cell(1,c,h); cell.font=hfn; cell.fill=hf; cell.alignment=caw; cell.border=tb
    ro=2
    if engine.has_history:
        ws2.cell(ro,1,'Jan-Avg 2025 (UKUPNO)').font=Font(bold=True,name='Arial',size=10,color='6B3FA0')
        ws2.cell(ro,1).alignment=ca; ws2.cell(ro,1).border=tb
        cell=ws2.cell(ro,3,int(df['Total_JanAvg'].sum())); cell.font=Font(bold=True,name='Arial',size=10,color='6B3FA0')
        cell.fill=sf_hist; cell.alignment=ca; cell.border=tb; cell.number_format=nf_money
        for c in [2,4,5,6]: ws2.cell(ro,c,'-').font=dfn; ws2.cell(ro,c).alignment=ca; ws2.cell(ro,c).border=tb
        ro+=2
    for ri,label in enumerate(ml,ro):
        ws2.cell(ri,1,label).font=Font(bold=True,name='Arial',size=10); ws2.cell(ri,1).alignment=ca; ws2.cell(ri,1).border=tb
        vals=[int(df[f'{label}_Promet'].sum()),int(df[f'{label}_Prodaja'].sum()),int(df[f'{label}_Povrat'].sum()),int(df[f'{label}_Korekcija'].sum())]
        vals.append(vals[0]-vals[2])
        fills=[sf_prom,sf_prod,sf_pov,sf_kor,sf_poc]
        for c2,(v,f) in enumerate(zip(vals,fills),2):
            cell=ws2.cell(ri,c2,v); cell.font=dfn; cell.fill=f; cell.alignment=ca; cell.border=tb; cell.number_format=nf_money
    fr=ro+len(ml)+1
    ws2.cell(fr,1,f'PORUDZBINA {engine.order_label.upper()}').font=Font(bold=True,name='Arial',size=11,color='375623'); ws2.cell(fr,1).border=tb
    ir=[(f'Predikcija {engine.pred_label}',int(df['Predikcija'].sum()),sf_pred),('Prosek (svi meseci)',int(df['Prosek'].sum()),sf_avg),
        ('Trenutni lager',int(df['Lager_danas'].sum()),sf_lager),
        ('Porudzbina (osnovna)',int(df[~df['ID KOMITENTA'].isin(engine.excluded)]['Porudzbina_1'].sum()),sf_p1),
        (f'Porudzbina (min. {engine.min_lager})',int(df[~df['ID KOMITENTA'].isin(engine.excluded)]['Porudzbina_2'].sum()),sf_p2)]
    for i,(label,val,fill) in enumerate(ir,fr+1):
        ws2.cell(i,1,label).font=Font(bold=True,name='Arial',size=10); ws2.cell(i,1).alignment=ca; ws2.cell(i,1).border=tb
        cell=ws2.cell(i,2,val); cell.font=Font(bold=True,name='Arial',size=11); cell.fill=fill; cell.alignment=ca; cell.border=tb; cell.number_format=nf_money
    ws2.column_dimensions['A'].width=32; ws2.column_dimensions['B'].width=18
    for c in 'CDEF': ws2.column_dimensions[c].width=18
    if engine.has_prices and len(engine.df_oos) > 0:
        ws_oos = wb.create_sheet("OOS Izgubljeni profit")
        oos_hdr = PatternFill('solid', fgColor='C00000')
        oos_fill = PatternFill('solid', fgColor='FCE4EC')
        a_labels_oos = engine.analitika_labels if engine.analitika_labels else engine.mesec_labels
        fixed_h = ['ID Komitenta','ID Artikla','Naziv','Grupa','Prosek kad ima','Lager danas']
        mes_h = []
        for lb in a_labels_oos: mes_h += [f'OOS {lb} (kom)', f'Izgub {lb} (RSD)']
        all_h = fixed_h + mes_h + ['OOS meseci ukupno','Izgubljeni profit (RSD)']
        for c, h in enumerate(all_h, 1):
            cell = ws_oos.cell(1, c, h)
            cell.font=Font(bold=True,color='FFFFFF',name='Arial',size=9)
            cell.fill=oos_hdr; cell.alignment=caw; cell.border=tb
        for idx, (_, row) in enumerate(engine.df_oos.iterrows(), 2):
            vals = [row['ID KOMITENTA'], row['id artikla'], row['Naziv artikla'], row['Grupa'],
                    row.get('Prosek_kad_ima',0), row.get('Lager_danas',0)]
            for lb in a_labels_oos:
                vals.append(row.get(f'OOS_{lb}', 0))
                vals.append(row.get(f'Izgub_{lb}', 0))
            vals += [row.get('OOS_meseci',0), row.get('Izgubljeni_profit',0)]
            for c, v in enumerate(vals, 1):
                cell = ws_oos.cell(idx, c, v); cell.font=dfn; cell.border=tb; cell.alignment=ca
                col_name = all_h[c-1]
                if col_name.startswith('OOS ') and isinstance(v, (int, float)) and v > 0:
                    cell.fill = oos_fill; cell.font = Font(name='Arial',size=9,bold=True,color='C00000')
                if col_name.startswith('Izgub ') or col_name == 'Izgubljeni profit (RSD)':
                    cell.number_format = nf_money
                if col_name == 'Lager danas' and v == 0:
                    cell.fill = oos_fill; cell.font = Font(name='Arial',size=9,bold=True,color='C00000')
        ws_oos.column_dimensions['A'].width=13; ws_oos.column_dimensions['B'].width=10
        ws_oos.column_dimensions['C'].width=45; ws_oos.column_dimensions['D'].width=12
        ws_oos.column_dimensions['E'].width=14; ws_oos.column_dimensions['F'].width=12
        for i in range(len(a_labels_oos)*2):
            ws_oos.column_dimensions[get_column_letter(7+i)].width=13
        last_col = 7 + len(a_labels_oos)*2
        ws_oos.column_dimensions[get_column_letter(last_col)].width=14
        ws_oos.column_dimensions[get_column_letter(last_col+1)].width=18
        ws_oos.freeze_panes='E2'
        ws_oos.auto_filter.ref=f"A1:{get_column_letter(len(all_h))}{len(engine.df_oos)+1}"
    if engine.has_prices and len(engine.df_profit_obj) > 0:
        ws_prof = wb.create_sheet("Profitabilnost objekata")
        prof_hdr = PatternFill('solid', fgColor='1F4E79')
        bad_fill = PatternFill('solid', fgColor='FCE4EC')
        good_fill = PatternFill('solid', fgColor='E2EFDA')
        headers = ['ID Komitenta','Artikala','Prodato kom','Bruto profit (RSD)','Trosak mkt (RSD)','Neto profit (RSD)','Izgubljeno OOS (RSD)','Potencijal (RSD)']
        for lb in (engine.analitika_labels if engine.analitika_labels else ml): headers.append(f'Neto {lb}')
        for c, h in enumerate(headers, 1):
            cell = ws_prof.cell(1, c, h); cell.font=Font(bold=True,color='FFFFFF',name='Arial',size=9); cell.fill=prof_hdr; cell.alignment=caw; cell.border=tb
        for idx, (_, row) in enumerate(engine.df_profit_obj.iterrows(), 2):
            vals = [row['ID KOMITENTA'], row['Artikala'], row['Prodato_kom'], row['Bruto_profit'],
                    row['Trosak_mkt'], row['Neto_profit'], row['Izgubljeno_OOS'], row['Potencijalni_profit']]
            for lb in (engine.analitika_labels if engine.analitika_labels else ml): vals.append(row.get(f'Neto_{lb}', 0))
            for c, v in enumerate(vals, 1):
                cell = ws_prof.cell(idx, c, v); cell.font=dfn; cell.border=tb; cell.alignment=ca
                if c >= 4: cell.number_format=nf_money
                if c == 6:
                    if v <= 0: cell.fill = bad_fill; cell.font = Font(name='Arial', size=9, bold=True, color='C00000')
                    elif v > 0: cell.fill = good_fill
                if c >= 9:
                    if v < 0: cell.font = Font(name='Arial', size=9, color='C00000')
                    elif v > 0: cell.font = Font(name='Arial', size=9, color='006100')
        for cl in 'AB': ws_prof.column_dimensions[cl].width=13
        ws_prof.column_dimensions['C'].width=12
        for cl in 'DEFGH': ws_prof.column_dimensions[cl].width=18
        a_ml = engine.analitika_labels if engine.analitika_labels else ml
        for i in range(len(a_ml)): ws_prof.column_dimensions[get_column_letter(9+i)].width=14
        ws_prof.freeze_panes='B2'
        ws_prof.auto_filter.ref=f"A1:{get_column_letter(len(headers))}{len(engine.df_profit_obj)+1}"
    if engine.has_prices and len(engine.df_promo) > 0:
        ws_akc = wb.create_sheet("Analiza akcije")
        akc_hdr = PatternFill('solid', fgColor='BF8F00')
        good_obrt = PatternFill('solid', fgColor='E2EFDA')
        bad_obrt = PatternFill('solid', fgColor='FCE4EC')
        headers = ['ID Artikla','Naziv','Grupa','Redovna\ncena','Akcijska\ncena','Popust\n%',
                   'Prodato\nkom','Prihod\nakcija (RSD)','Prihod da je\nredovna (RSD)',
                   'Profit\nakcija (RSD)','Profit da je\nredovna (RSD)','Cena akcije\n(RSD)',
                   'Prosecni\nlager','Obrt\n(x)','Dani\npokrivanja',
                   'Aktivnih\nobjekata','Ukupno\nobjekata','Prod.\npo objektu']
        for lb in (engine.analitika_labels if engine.analitika_labels else ml): headers.append(f'Prod.\n{lb}')
        for c, h in enumerate(headers, 1):
            cell = ws_akc.cell(1, c, h); cell.font=Font(bold=True,color='FFFFFF',name='Arial',size=9); cell.fill=akc_hdr; cell.alignment=caw; cell.border=tb
        for idx, (_, row) in enumerate(engine.df_promo.iterrows(), 2):
            vals = [row['id artikla'], row['Naziv'], row['Grupa'], row['Redovna'], row['Akcijska'],
                    row['Popust_%'], row['Prodato_kom'],
                    row['Prihod_akcija'], row['Prihod_redovna'],
                    row['Profit_akcija'], row['Profit_da_je_redovna'], row['Cena_akcije'],
                    row['Avg_lager'], row['Obrt_x'], row['Dani_pokrivanja'],
                    row['Obj_aktivnih'], row['Obj_ukupno'], row['Prod_po_obj']]
            for lb in (engine.analitika_labels if engine.analitika_labels else ml): vals.append(row.get(f'Prod_{lb}', 0))
            for c, v in enumerate(vals, 1):
                cell = ws_akc.cell(idx, c, v); cell.font=dfn; cell.border=tb; cell.alignment=ca
                if c in [4,5,8,9,10,11,12]: cell.number_format=nf_money
                if c == 14:
                    if v >= 2.0: cell.fill = good_obrt; cell.font = Font(name='Arial',size=9,bold=True,color='006100')
                    elif v < 1.0: cell.fill = bad_obrt; cell.font = Font(name='Arial',size=9,bold=True,color='C00000')
                if c == 15 and v > 120: cell.fill = bad_obrt
        ws_akc.column_dimensions['A'].width=10; ws_akc.column_dimensions['B'].width=45; ws_akc.column_dimensions['C'].width=12
        for cl in 'DEFG': ws_akc.column_dimensions[cl].width=12
        for cl in 'HIJKL': ws_akc.column_dimensions[cl].width=16
        for cl in 'MNOPQR': ws_akc.column_dimensions[cl].width=13
        a_ml2 = engine.analitika_labels if engine.analitika_labels else ml
        for i in range(len(a_ml2)): ws_akc.column_dimensions[get_column_letter(19+i)].width=11
        ws_akc.auto_filter.ref=f"A1:{get_column_letter(len(headers))}{len(engine.df_promo)+1}"
    ws3=wb.create_sheet("O modelu"); ws3.column_dimensions['A'].width=100
    info=["OPIS MODELA PREDIKCIJE I PORUDZBINE","",f"=== PREDIKCIJA ZA {engine.pred_label.upper()} ===","",
        "Model predvidja POTENCIJAL PRODAJE.","",
        f"  1. Constrained sales korekcija:",
        f"     - Kraj meseca lager=0 i prodaja>0: rasprodato, potraznja veca — zameni prosekom normalnih meseci",
        f"     - Pocetno=0 i promet=0: cist OOS — zameni prosekom normalnih meseci",
        f"     - Pocetno=0 i promet>0 i kraj=0: dobili i rasprodali — zameni prosekom",
        f"     - Normalni meseci = ostalo robe na kraju (lager>0)",
        f"  2. Holt DES (alpha={engine.alpha}, beta={engine.beta}) + WMA (50/28/12/7/3%)",
        "  3. Kombinacija: 60% veci + 40% manji od Holt/WMA",
        "  4. Varijansa boost (faktor 0.4, max 70%)",
        "  5. Niska zaliha (0-2): predikcija minimum prosek kad je na stanju",
        "  6. Prodaja 5+ mesecno: predikcija minimum prosek",
        "  7. Donje ogranicenje: predikcija < prosek samo ako poslednjih 5 meseci pada ili stagnira (<=)",
        "  8. Sigurnosna mreza: predikcija=0 samo ako nista prodato u poslednjih 5 meseci; ako poslednji mesec >1 onda min taj broj",
        "  9. Zaokruzivanje: round (predikcija i prosek)",
        ]
    if engine.has_history: info+=[f"  10. Istorijski podaci: {HIST_WEIGHT*100:.0f}% tezina"]
    info+=["",f"=== PORUDZBINA ZA {engine.order_label.upper()} ===","",
        f"P1 (osnovna): max(Pred-Lager, 0)",
        f"P2 (sa dopunom): Za lager<=2: dopuna do max(predikcija, prosek, min porudzbina={engine.min_order}); Za lager>2: dopuna do min {engine.min_lager}",
        f"P2 finalna provera: ako (P2+lager) <= prodaja_poslednjeg_meseca, dodaje se buffer (1-5 kom: +2, 6-10: +3, 11-15: +4, 16+: +5)",
        f"Min kom po artiklu (po stavci): {engine.min_per_artikal if engine.min_per_artikal else 'nije zadat'} — ako je porudzbina > 0 ali manja od minimuma, podize se na minimum. Nule ostaju 0.",
        f"Iskljuceni: {', '.join(str(x) for x in sorted(engine.excluded))}"]
    if engine.has_prices:
        info+=["",f"=== ANALITIKA ===","",
            f"Profit formula: (Finalna cena / 1.2 / 1.2 - Nabavna) x Kolicina",
            f"OOS izgubljeni profit (NOVA LOGIKA - 3 uslova):",
            f"  USLOV 1 - Cist OOS (poc=0 i ulaz=0): izgubljeno_kom = avg_stocked",
            f"  USLOV 2 - Rasprodato (kraj=0 i prodaja>0): izgubljeno_kom = max(0, avg_stocked - prodaja)",
            f"  USLOV 3 - Dobili i rasprodali (poc=0, ulaz>0, kraj=0): izgubljeno_kom = max(0, avg_stocked - prodaja)",
            f"  Ako mesec nije constrained (kraj > 0): OOS = 0 (imali su robe i nisu rasprodali)",
            f"  avg_stocked = prosek prodaje u mesecima koji NISU constrained",
            f"  Izgubljeni profit = izgubljeno_kom x profit/kom po mesecu",
            f"Ukupan trosak marketinga: {engine.mesecni_trosak:,.0f} RSD / {engine.num_komitenti} objekata = {engine.trosak_po_objektu:,.0f} RSD po objektu za period",
            f"Mesecni trosak po objektu: {engine.trosak_po_objektu / max(len(engine.analitika_labels), 1):,.0f} RSD",
            f"Neto po mesecu = Bruto profit meseca - mesecni trosak po objektu"]
    info+=[f"","Generisano: {datetime.datetime.now().strftime('%d.%m.%Y. u %H:%M')}"]
    for i,line in enumerate(info,1):
        cell=ws3.cell(i,1,line)
        if i==1: cell.font=Font(bold=True,name='Arial',size=14,color='375623')
        elif '===' in line: cell.font=Font(bold=True,name='Arial',size=12,color='7030A0')
        else: cell.font=Font(name='Arial',size=10)
    if not ukljuci_model:
        try:
            if "O modelu" in wb.sheetnames:
                del wb["O modelu"]
        except Exception:
            pass
    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf

DEFAULT_EXCLUDED = "1023, 1027, 1034, 1043, 1057, 1060, 1061, 1076, 1315, 1347, 1349, 1359"
st.set_page_config(page_title="VAPE Analitika", page_icon="\U0001f4a8", layout="wide", initial_sidebar_state="collapsed")
st.markdown("""<style>
section[data-testid="stSidebar"] { display: none !important; }
header[data-testid="stHeader"] { display: none !important; }
#MainMenu { visibility: hidden !important; }
footer { visibility: hidden !important; }
.main .block-container,
div[data-testid="block-container"],
div[data-testid="stMainBlockContainer"] {
    padding: 12px 16px 0 16px !important;
    max-width: 100% !important;
}
</style>""", unsafe_allow_html=True)
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Poppins:wght@300;400;500;600;700&display=swap');
    .stApp {
        background: #f5f0ff !important;
        font-family: 'Poppins', sans-serif;
    }
    .main .stTextInput > div > div > input,
    .main .stNumberInput > div > div > input {
        background: white !important;
        border: 1px solid rgba(168,85,247,0.25) !important;
        color: #1a0533 !important;
        border-radius: 8px !important;
    }
    .main .stTextInput > div > div > input::placeholder,
    .main .stNumberInput > div > div > input::placeholder {
        color: #9ca3af !important;
    }
    .main .stTextInput > div > div > input:focus,
    .main .stNumberInput > div > div > input:focus {
        border-color: #a855f7 !important;
        box-shadow: 0 0 0 2px rgba(168,85,247,0.15) !important;
    }
    .metric-card {
        background: white;
        border-radius: 14px;
        padding: 16px 20px;
        box-shadow: 0 2px 12px rgba(124,58,237,0.07);
        border: 1px solid rgba(168,85,247,0.12);
        text-align: center;
    }
    .metric-value {
        font-size: 26px; font-weight: 700;
        background: linear-gradient(135deg, #7c3aed, #ec4899);
        -webkit-background-clip: text; -webkit-text-fill-color: transparent;
    }
    .metric-value-red { font-size: 26px; font-weight: 700; color: #dc2626; }
    .metric-value-green { font-size: 26px; font-weight: 700; color: #059669; }
    .metric-label { font-size: 11px; color: #888; margin-top: 4px; }
    .stButton > button {
        background: linear-gradient(135deg, #a855f7 0%, #ec4899 100%) !important;
        color: white !important;
        border: none !important;
        border-radius: 12px !important;
        padding: 14px 32px !important;
        font-weight: 700 !important;
        font-size: 15px !important;
        box-shadow: 0 4px 15px rgba(168,85,247,0.3) !important;
        transition: opacity 0.2s !important;
    }
    .stButton > button:hover { opacity: 0.88 !important; }
    .stDownloadButton > button {
        background: linear-gradient(135deg, #10b981 0%, #059669 100%) !important;
        color: white !important;
        border: none !important;
        border-radius: 12px !important;
        padding: 14px 32px !important;
        font-weight: 700 !important;
        box-shadow: 0 4px 15px rgba(16,185,129,0.25) !important;
    }
    .stMultiSelect [data-baseweb="tag"] {
        background: linear-gradient(135deg, #a855f7, #ec4899) !important;
        border: none !important;
        border-radius: 99px !important;
        color: white !important;
        font-weight: 600 !important;
        font-size: 12px !important;
    }
    .stMultiSelect [data-baseweb="tag"] span { color: white !important; }
    .stMultiSelect [data-baseweb="tag"] button { color: rgba(255,255,255,0.8) !important; }
    .stMultiSelect [data-baseweb="select"] > div {
        border: 1px solid rgba(168,85,247,0.3) !important;
        border-radius: 10px !important;
        background: white !important;
    }
    .stMultiSelect [data-baseweb="select"] > div:focus-within {
        border-color: #a855f7 !important;
        box-shadow: 0 0 0 2px rgba(168,85,247,0.15) !important;
    }
    .success-box {
        background: linear-gradient(135deg, rgba(16,185,129,0.08), rgba(5,150,105,0.04));
        border: 1px solid rgba(16,185,129,0.2);
        border-radius: 10px;
        padding: 12px 16px;
    }
    .warn-box {
        background: linear-gradient(135deg, rgba(220,38,38,0.07), rgba(220,38,38,0.02));
        border: 1px solid rgba(220,38,38,0.18);
        border-radius: 10px;
        padding: 10px 14px;
        margin: 6px 0;
    }
    .section-title {
        font-size: 17px; font-weight: 600; color: #4c1d95; margin: 16px 0 8px 0;
    }
</style>
""", unsafe_allow_html=True)
alpha = 0.4
beta = 0.2
min_lager = None
min_order = None
min_per_artikal = None
max_per_artikal = None
mesecni_trosak = 0
excluded_str = DEFAULT_EXCLUDED
excluded = set()
for part in excluded_str.replace('\n', ',').split(','):
    p = part.strip()
    if p.isdigit(): excluded.add(int(p))
def render_header(subtitle):
    st.markdown(f'''<div style="background:#12002a;border-radius:16px;padding:0 28px;height:60px;
        display:flex;align-items:center;justify-content:space-between;margin-bottom:24px;
        border-bottom:3px solid;border-image:linear-gradient(90deg,#a855f7,#ec4899) 1;
        box-shadow:0 4px 20px rgba(18,0,42,0.18);">
        <div style="display:flex;align-items:center;gap:12px;">
            <div style="width:30px;height:30px;background:linear-gradient(135deg,#a855f7,#ec4899);
                border-radius:8px;display:flex;align-items:center;justify-content:center;">
                <div style="width:11px;height:11px;background:white;border-radius:3px;"></div>
            </div>
            <span style="font-size:18px;font-weight:700;color:white;">VAPE</span>
            <span style="font-size:18px;font-weight:300;color:rgba(255,255,255,0.4);">Analitika</span>
            <span style="font-size:11px;color:rgba(255,255,255,0.25);margin-left:8px;">·</span>
            <span style="font-size:12px;color:rgba(255,255,255,0.35);">{subtitle}</span>
        </div>
        <div style="display:flex;gap:12px;align-items:center;">
            <div style="display:flex;gap:6px;align-items:center;">
                <div style="width:8px;height:8px;border-radius:50%;background:rgba(168,85,247,0.7);"></div>
                <div style="width:8px;height:8px;border-radius:50%;background:rgba(236,72,153,0.5);"></div>
                <div style="width:8px;height:8px;border-radius:50%;background:rgba(255,255,255,0.15);"></div>
            </div>
        </div>
    </div>''', unsafe_allow_html=True)
render_header("Predikcija prodaje · Profitabilnost · OOS analiza · Efekti akcije")
_co = st.columns([6, 1])
with _co[1]:
    if st.button("🔓 Odjava", key="ana_odjava"):
        for _k in ("authenticated", "role"):
            st.session_state.pop(_k, None)
        st.rerun()

st.markdown("""<style>
/* --- Tabovi kao pilule --- */
div[data-baseweb="tab-list"] { gap: 8px; border-bottom: none !important; }
button[data-baseweb="tab"] {
    background:#fff; border:1px solid rgba(168,85,247,0.22); border-radius:12px;
    padding:8px 20px; color:#6b21a8; font-weight:600; }
button[data-baseweb="tab"][aria-selected="true"] {
    background:linear-gradient(135deg,#a855f7,#ec4899); color:#fff; border-color:transparent;
    box-shadow:0 4px 14px rgba(168,85,247,0.3); }
div[data-baseweb="tab-highlight"], div[data-baseweb="tab-border"] { display:none !important; }
/* --- Kartice (bordered container) --- */
div[data-testid="stVerticalBlockBorderWrapper"] {
    background:#fff !important; border:1px solid rgba(168,85,247,0.14) !important;
    border-radius:14px !important; box-shadow:0 2px 12px rgba(124,58,237,0.07) !important;
    padding:14px 18px !important; margin-bottom:10px !important; }
/* --- File uploader dropzone --- */
[data-testid="stFileUploaderDropzone"] {
    background:#faf5ff !important; border:2px dashed rgba(168,85,247,0.35) !important;
    border-radius:12px !important; }
/* --- Objavi dugme zeleno --- */
.st-key-obj_btn button {
    background:linear-gradient(135deg,#10b981,#059669) !important;
    box-shadow:0 4px 15px rgba(16,185,129,0.25) !important; }
/* --- Naslovi kartica --- */
.obj-title { font-size:15px; font-weight:600; color:#4c1d95; margin:0 0 10px 0; display:flex; align-items:center; gap:8px; }
.obj-badge { background:linear-gradient(135deg,#a855f7,#ec4899); color:#fff; width:22px; height:22px;
    border-radius:50%; display:inline-flex; align-items:center; justify-content:center; font-size:12px; font-weight:700; }
.obj-badge.green { background:linear-gradient(135deg,#10b981,#059669); }
.chip-ok { display:inline-block; background:#eafaf0; color:#059669; border:1px solid rgba(16,185,129,.25);
    border-radius:99px; padding:5px 13px; font-size:12.5px; font-weight:600; margin:4px 5px 0 0; }
.chip-no { display:inline-block; background:#f3f4f6; color:#9ca3af; border:1px dashed #d1d5db;
    border-radius:99px; padding:5px 13px; font-size:12.5px; font-weight:600; margin:4px 5px 0 0; }
</style>""", unsafe_allow_html=True)

tab_obj, tab_ana = st.tabs(["📤 Objava izveštaja", "📊 Analitika"])

with tab_obj:
    with st.container(border=True):
        st.markdown('<div class="obj-title">📋 Objavljeno za koleginice</div>', unsafe_allow_html=True)
        if not sb_dostupan():
            st.info("Supabase nije podešen — dodaj SUPABASE_URL i SUPABASE_KEY u Streamlit Secrets da bi objava radila.")
        else:
            _pmeseci = sb_meseci()
            if not _pmeseci:
                st.caption("Još nijedan sistem nije objavljen.")
            else:
                _plabels = [m["label"] for m in _pmeseci]
                _pkeys = [m["key"] for m in _pmeseci]
                _psel = st.selectbox("Mesec", _plabels, index=0, key="obj_preg_mes")
                _pmk = _pkeys[_plabels.index(_psel)]
                _imaju = sb_sisteme(_pmk)
                _svi = sb_svi_sistemi()
                _chip = ""
                for _s in _imaju:
                    _chip += '<span class="chip-ok">\u2713 ' + str(_s) + '</span>'
                for _s in _svi:
                    if _s not in _imaju:
                        _chip += '<span class="chip-no">\u25cb ' + str(_s) + ' \u00b7 nije objavljen</span>'
                if not _chip:
                    _chip = '<span style="color:#9ca3af;font-size:13px;">Nema objavljenih sistema za ovaj mesec.</span>'
                st.markdown(_chip, unsafe_allow_html=True)

                # --- Brisanje objavljenih izve\u0161taja za ovaj mesec ---
                if _imaju:
                    with st.expander("Obri\u0161i objavljeni izve\u0161taj (" + _psel + ")"):
                        st.caption("Izaberi sistem(e) za ovaj mesec koje \u017Eeli\u0161 da obri\u0161e\u0161, pa potvrdi. "
                                   "Brisanje je trajno \u2014 posle mo\u017Ee\u0161 da objavi\u0161 nove.")
                        _del_sel = st.multiselect("Sistemi za brisanje", _imaju, key="del_sis_" + _pmk)
                        _del_ok = st.checkbox("Potvr\u0111ujem brisanje izabranih", key="del_ok_" + _pmk)
                        if st.button("Obri\u0161i izabrano", key="del_btn_" + _pmk,
                                     disabled=not (_del_sel and _del_ok), use_container_width=True):
                            _nbr = 0
                            for _s in _del_sel:
                                try:
                                    sb_obrisi(_pmk, _s); _nbr += 1
                                except Exception as _e:
                                    st.error("Gre\u0161ka pri brisanju \u201E" + str(_s) + "\u201C: " + str(_e))
                            if _nbr:
                                st.success("Obrisano: " + str(_nbr) + " izve\u0161taj(a). Sad mo\u017Ee\u0161 da objavi\u0161 nove.")
                                st.rerun()

    with st.container(border=True):
        st.markdown('<div class="obj-title">\U0001F5D3\uFE0F Plan objave (za koleginice)</div>', unsafe_allow_html=True)
        if not sb_dostupan():
            st.caption("Nedostupno dok Supabase nije podešen.")
        else:
            _today = datetime.date.today()
            _mopts = []
            _yy = _today.year; _mm = _today.month - 3  # uključi i prethodne mesece
            while _mm <= 0:
                _mm += 12; _yy -= 1
            for _i in range(9):
                _mopts.append(str(_yy) + "-" + ("0" + str(_mm))[-2:])
                _mm += 1
                if _mm > 12:
                    _mm = 1; _yy += 1
            # podrazumevano: prethodni mesec (u avgustu se radi izveštaj za jul)
            _pv_y = _today.year; _pv_m = _today.month - 1
            if _pv_m <= 0:
                _pv_m += 12; _pv_y -= 1
            _prev_key = str(_pv_y) + "-" + ("0" + str(_pv_m))[-2:]
            _def_idx = _mopts.index(_prev_key) if _prev_key in _mopts else 0
            _pc1, _pc2, _pc3 = st.columns([1.3, 1, 1])
            with _pc1:
                _pmes = st.selectbox("Mesec", _mopts, index=_def_idx, format_func=mesec_label, key="plan_mes")
            with _pc2:
                _pdat = st.date_input("Objaviću do", value=_today, key="plan_dat", format="DD.MM.YYYY")
            with _pc3:
                st.markdown("<div style='height:28px;'></div>", unsafe_allow_html=True)
                if st.button("\U0001F4BE Sačuvaj plan", key="plan_btn", use_container_width=True):
                    try:
                        sb_save_plan(_pmes, _pdat.strftime("%d.%m.%Y"))
                        st.success("Plan sačuvan: " + mesec_label(_pmes) + " \u2192 do " + _pdat.strftime("%d.%m.%Y"))
                    except Exception as _e:
                        st.error("Greška: " + str(_e))
            _curplan = sb_load_plan(_pmes)
            if _curplan:
                st.caption("Trenutni plan za " + mesec_label(_pmes) + ": do " + str(_curplan))

    with st.container(border=True):
        st.markdown('<div class="obj-title">📄 Izveštaj SYX (nikotinske vrećice)</div>', unsafe_allow_html=True)
        st.caption("Ubaci Word (.docx) izveštaj za SYX po mesecu. Kod direktora se pojavljuje u kartici Izveštaj SYX, izlistan po mesecima.")
        if not sb_dostupan():
            st.caption("Nedostupno dok Supabase nije podešen.")
        else:
            _sy_today = datetime.date.today()
            _sy_opts = set(m["key"] for m in sb_meseci())
            for _r in sb_syx_list():
                _sy_opts.add(_r.get("mesec"))
            _yy, _mm = _sy_today.year, _sy_today.month - 2
            while _mm <= 0:
                _mm += 12; _yy -= 1
            for _ in range(9):
                _sy_opts.add(str(_yy) + "-" + ("0" + str(_mm))[-2:])
                _mm += 1
                if _mm > 12:
                    _mm = 1; _yy += 1
            _sy_opts = sorted([o for o in _sy_opts if o], reverse=True)
            _sy_pvy = _sy_today.year; _sy_pvm = _sy_today.month - 1
            if _sy_pvm <= 0:
                _sy_pvm += 12; _sy_pvy -= 1
            _sy_prev = str(_sy_pvy) + "-" + ("0" + str(_sy_pvm))[-2:]
            _sy_idx = _sy_opts.index(_sy_prev) if _sy_prev in _sy_opts else 0
            _syc1, _syc2 = st.columns([1, 2])
            with _syc1:
                _sy_mes = st.selectbox("Mesec", _sy_opts, index=_sy_idx, format_func=mesec_label, key="syx_mes")
            with _syc2:
                _sy_file = st.file_uploader("Word dokument (.docx)", type=["docx"], key="syx_up")
            _rok_sy = sb_rokovi_get(_sy_mes).get("rok_syx")
            if _rok_sy:
                st.caption(("⏰ Rok za SYX (" + mesec_label(_sy_mes) + "): " + _rok_fmt(_rok_sy))
                           + ("  ·  rok istekao" if _rok_je_prosao(_rok_sy) else ""))
            if st.button("📤 Sačuvaj SYX izveštaj", key="syx_save", use_container_width=True,
                         disabled=(_sy_file is None)):
                try:
                    import base64 as _b64s
                    _bytes = _sy_file.getvalue()
                    _b64 = _b64s.b64encode(_bytes).decode("ascii")
                    sb_syx_set(_sy_mes, _sy_file.name, _b64)
                    st.success("SYX izveštaj za " + mesec_label(_sy_mes) + " sačuvan (" + _sy_file.name + ").")
                    st.rerun()
                except Exception as _e:
                    st.error("Greška pri čuvanju: " + str(_e))
            _sy_all = sb_syx_list()
            if _sy_all:
                st.markdown("<div style='margin-top:6px;font-size:12px;text-transform:uppercase;letter-spacing:.5px;"
                            "color:#9aa0ad;font-weight:700;'>Postavljeni SYX izveštaji</div>", unsafe_allow_html=True)
                for _r in _sy_all:
                    _rc1, _rc2, _rc3 = st.columns([2, 3, 1])
                    with _rc1:
                        st.markdown("**" + mesec_label(_r.get("mesec", "")) + "**")
                    with _rc2:
                        st.caption(str(_r.get("filename", "")))
                    with _rc3:
                        if st.button("Obriši", key="syx_del_" + str(_r.get("mesec")), use_container_width=True):
                            try:
                                sb_syx_obrisi(_r.get("mesec"))
                                st.rerun()
                            except Exception as _e:
                                st.error("Greška: " + str(_e))

    with st.container(border=True):
        st.markdown('<div class="obj-title">💳 Izveštaj potraživanja</div>', unsafe_allow_html=True)
        st.caption("Ubaci Excel izveštaja potraživanja po mesecu. Administracija ga dopunjava direktno u aplikaciji "
                   "(iznosi, statusi, komentari) i prosleđuje; direktor ga vidi i izvozi u identičan Excel.")
        if not sb_dostupan():
            st.caption("Nedostupno dok Supabase nije podešen.")
        else:
            _pz_today = datetime.date.today()
            _pz_opts = set(m["key"] for m in sb_meseci())
            for _r in sb_potraz_list():
                _pz_opts.add(_r.get("mesec"))
            _yy, _mm = _pz_today.year, _pz_today.month - 3
            while _mm <= 0:
                _mm += 12; _yy -= 1
            for _ in range(9):
                _pz_opts.add(str(_yy) + "-" + ("0" + str(_mm))[-2:])
                _mm += 1
                if _mm > 12:
                    _mm = 1; _yy += 1
            _pz_opts = sorted([o for o in _pz_opts if o], reverse=True)
            _pz_pvy = _pz_today.year; _pz_pvm = _pz_today.month - 1
            if _pz_pvm <= 0:
                _pz_pvm += 12; _pz_pvy -= 1
            _pz_prev = str(_pz_pvy) + "-" + ("0" + str(_pz_pvm))[-2:]
            _pz_idx = _pz_opts.index(_pz_prev) if _pz_prev in _pz_opts else 0
            _pzc1, _pzc2 = st.columns([1, 2])
            with _pzc1:
                _pz_mes = st.selectbox("Mesec", _pz_opts, index=_pz_idx, format_func=mesec_label, key="pz_mes")
            with _pzc2:
                _pz_file = st.file_uploader("Excel potraživanja (.xlsx)", type=["xlsx"], key="pz_up")
            _rok_pzc = sb_rokovi_get(_pz_mes).get("rok_potraz")
            if _rok_pzc:
                st.caption(("⏰ Rok za potraživanja (" + mesec_label(_pz_mes) + "): " + _rok_fmt(_rok_pzc))
                           + ("  ·  rok istekao" if _rok_je_prosao(_rok_pzc) else ""))
            if st.button("📤 Objavi za administraciju", key="pz_save", use_container_width=True,
                         disabled=(_pz_file is None)):
                try:
                    import base64 as _b64p
                    _pb = _pz_file.getvalue()
                    _struct = potraz_parse(_pb)
                    _pop = potraz_init_popuna(_struct)
                    sb_potraz_set(_pz_mes, _pz_file.name, _b64p.b64encode(_pb).decode("ascii"),
                                  json.dumps(_struct), json.dumps(_pop))
                    _nred = sum(len(s["redovi"]) for L in _struct["listovi"] for s in L["sekcije"])
                    st.success("Objavljeno za administraciju: " + mesec_label(_pz_mes) + " · "
                               + str(len(_struct["listovi"])) + " listova, " + str(_nred) + " stavki.")
                    st.rerun()
                except Exception as _e:
                    st.error("Greška pri obradi Excela: " + str(_e))
            _pz_all = sb_potraz_list()
            if _pz_all:
                st.markdown("<div style='margin-top:6px;font-size:12px;text-transform:uppercase;letter-spacing:.5px;"
                            "color:#9aa0ad;font-weight:700;'>Objavljeni izveštaji potraživanja</div>", unsafe_allow_html=True)
                for _r in _pz_all:
                    _rc1, _rc2, _rc3 = st.columns([2, 3, 1])
                    with _rc1:
                        st.markdown("**" + mesec_label(_r.get("mesec", "")) + "**"
                                    + ("  ✅ predato" if _r.get("predato") else "  ⏳ u obradi"))
                    with _rc2:
                        st.caption(str(_r.get("naziv", "")))
                    with _rc3:
                        if st.button("Obriši", key="pz_del_" + str(_r.get("mesec")), use_container_width=True):
                            try:
                                sb_potraz_obrisi(_r.get("mesec"))
                                st.rerun()
                            except Exception as _e:
                                st.error("Greška: " + str(_e))

    with st.container(border=True):
        st.markdown('<div class="obj-title">👥 Šifarnik komitenata (nazivi + kontakt)</div>', unsafe_allow_html=True)
        st.caption("Učitaj Excel sa kolonama: ID, Naziv, e-mail, Kontakt, Mesto, Adresa. Koristi se za nazive objekata i istoriju porudžbina. Ubaci pre objave da se pokupe i nove radnje.")
        if not sb_dostupan():
            st.caption("Nedostupno dok Supabase nije podešen.")
        else:
            _dc1, _dc2 = st.columns([2, 1])
            with _dc2:
                if st.button("🔍 Proveri šifarnik u bazi", key="kom_check", use_container_width=True):
                    _all = _sb_select_all("komitenti", "idk,naziv,email")
                    _sa_naz = sum(1 for r in _all if (r.get("naziv") or "").strip())
                    _sa_mail = sum(1 for r in _all if (r.get("email") or "").strip())
                    st.session_state["_kom_check"] = {"uk": len(_all), "naz": _sa_naz, "mail": _sa_mail}
            _chk = st.session_state.get("_kom_check")
            if _chk:
                with _dc1:
                    st.caption("U bazi: " + str(_chk["uk"]) + " komitenata · sa nazivom " + str(_chk["naz"])
                               + " · sa mejlom " + str(_chk["mail"]) + ".")
            up_k = st.file_uploader("Excel šifarnika komitenata (.xlsx)", type=['xlsx', 'xls'],
                                    key="kom_upl", label_visibility="collapsed")
            if up_k is not None:
                try:
                    _dfk = pd.read_excel(up_k, dtype=str)
                    _cmap = {}
                    for _c in _dfk.columns:
                        _cl = str(_c).strip().lower()
                        if _cl == "id" or _cl == "idk":
                            _cmap["idk"] = _c
                        elif _cl.startswith("naziv"):
                            _cmap["naziv"] = _c
                        elif ("mail" in _cl) or ("mejl" in _cl):
                            _cmap["email"] = _c
                        elif _cl.startswith("kontakt") or _cl.startswith("telefon") or _cl.startswith("tel"):
                            _cmap["telefon"] = _c
                        elif _cl.startswith("mesto") or _cl.startswith("grad"):
                            _cmap["mesto"] = _c
                        elif _cl.startswith("adresa"):
                            _cmap["adresa"] = _c
                    if "idk" not in _cmap or "naziv" not in _cmap:
                        st.error("Fajlu fale kolone. Potrebne su bar 'ID' i 'Naziv'. Pronađene kolone: " + ", ".join(str(c) for c in _dfk.columns))
                    else:
                        _krows = []
                        for _, _rr in _dfk.iterrows():
                            _idraw = _rr.get(_cmap["idk"])
                            if _idraw is None or str(_idraw).strip() == "" or str(_idraw).strip().lower() == "nan":
                                continue
                            try:
                                _idv = int(float(str(_idraw).strip()))
                            except Exception:
                                continue
                            def _cell(_key):
                                if _key not in _cmap:
                                    return ""
                                _val = _rr.get(_cmap[_key])
                                if _val is None or str(_val).strip().lower() == "nan":
                                    return ""
                                return str(_val).strip()
                            _krows.append({"idk": _idv,
                                           "naziv": _clean_komitent_naziv(_cell("naziv")),
                                           "email": _cell("email").replace(" ", ""),
                                           "telefon": _cell("telefon"),
                                           "mesto": _cell("mesto"),
                                           "adresa": _cell("adresa")})
                        st.caption("Pronađeno " + str(len(_krows)) + " komitenata u fajlu.")
                        if st.button("💾 Sačuvaj šifarnik komitenata", key="kom_save", use_container_width=True):
                            try:
                                _n = sb_komitenti_upsert_rows(_krows)
                                st.session_state["_komitenti_map"] = None
                                st.session_state["_komfull"] = None
                                if _n and _n > 0:
                                    st.success("Sačuvano ✓ U bazi sada ima " + str(_n) + " komitenata. "
                                               "Nazivi će se videti u administraciji (osveži F5).")
                                else:
                                    st.error("Ništa nije upisano u bazu. Pokreni SQL setup za tabelu 'komitenti' (vidi uputstvo).")
                            except Exception as _e:
                                st.error("Greška pri čuvanju: " + str(_e))
                except Exception as _e:
                    st.error("Ne mogu da pročitam fajl: " + str(_e))

    with st.container(border=True):
        st.markdown('<div class="obj-title">📊 Izveštaj prodaje (za direktore)</div>', unsafe_allow_html=True)
        st.caption("Ubaci dve tabele (tabela sistemi + tabela troškova). Pravi se dashboard koji direktori vide kao treću karticu. Čuva se samo poslednji.")
        if not sb_dostupan():
            st.caption("Nedostupno dok Supabase nije podešen.")
        else:
            _ip_c1, _ip_c2 = st.columns(2)
            with _ip_c1:
                _up_sis = st.file_uploader("Tabela sistemi (.xlsx)", type=['xlsx'], key="izp_sis")
            with _ip_c2:
                _up_tro = st.file_uploader("Tabela troškova (.xlsx)", type=['xlsx'], key="izp_tro")
            _ip_q1 = st.radio("Tip izveštaja", ["Potpun (sve 4 kartice)", "Nepotpun (Prodaja + Uspešnost akcije)"],
                              key="izp_potpun", horizontal=True)
            _potpun = _ip_q1.startswith("Potpun")
            _iskljuci = False
            if _potpun:
                _ip_q2 = st.radio("Poslednji mesec u profitabilnosti?", ["Uključi", "Isključi"],
                                  key="izp_iskljuci", horizontal=True)
                _iskljuci = (_ip_q2 == "Isključi")
            if _up_sis is not None and _up_tro is not None:
                if st.button("📊 Generiši i objavi izveštaj prodaje", key="izp_gen", use_container_width=True):
                    try:
                        import io as _io2, base64 as _b64i
                        import izvestaj_prodaje as _izp
                        with st.spinner("Generišem izveštaj prodaje (može par sekundi)..."):
                            _html, _xlsx, _mes, _spr = _izp.generisi_izvestaj_prodaje(
                                _io2.BytesIO(_up_sis.getvalue()), _io2.BytesIO(_up_tro.getvalue()),
                                potpun=_potpun, iskljuci_poslednji=_iskljuci)
                            _xb64 = _b64i.b64encode(_xlsx).decode("ascii") if _xlsx else ""
                            _pj = json.dumps(_spr, ensure_ascii=False) if _spr else ""
                            sb_objavi_izvestaj_prodaje(_html, _xb64, _mes, prodaja_json=_pj)
                        st.success("✅ Izveštaj prodaje objavljen (" + str(_mes) + "). Direktori ga vide u trećoj kartici.")
                    except ModuleNotFoundError:
                        st.error("Nedostaje fajl izvestaj_prodaje.py u projektu — dodaj ga na GitHub pored streamlit_app.py.")
                    except Exception as _e:
                        st.error("Greška: " + str(_e))
                        import traceback as _tb2
                        st.code(_tb2.format_exc())

    with st.container(border=True):
        st.markdown('<div class="obj-title"><span class="obj-badge">1</span> Učitaj Excel jednog sistema</div>', unsafe_allow_html=True)
        up_o = st.file_uploader("Excel fajl (.xlsx)", type=['xlsx', 'xls'], key="obj_upl", label_visibility="collapsed")

    if up_o is not None:
        _obytes = up_o.read()
        st.markdown(f'<div class="success-box">\u2705 Fajl <strong>{up_o.name}</strong> učitan ({len(_obytes)//1024} KB)</div>', unsafe_allow_html=True)
        _oy = None; _omn = None
        try:
            _x = pd.ExcelFile(io.BytesIO(_obytes))
            _sm = {s.strip().lower(): s for s in _x.sheet_names}
            _sp = None
            for _nl, _no in _sm.items():
                if 'prodaja' in _nl:
                    _sp = _no; break
            _pdf = pd.read_excel(_x, sheet_name=_sp)
            _pdf.columns = [c.strip() for c in _pdf.columns]
            _ms = sorted(_pdf[['Godina', 'Mesec']].drop_duplicates().values.tolist())
            _oy = int(_ms[-1][0]); _omn = int(_ms[-1][1]) + 1
            if _omn > 12:
                _omn = 1; _oy += 1
        except Exception:
            _oy = None; _omn = None
        _mk = (str(_oy) + "-" + ("0" + str(_omn))[-2:]) if _oy else None
        _mlbl = mesec_label(_mk) if _mk else "automatski"

        with st.container(border=True):
            st.markdown('<div class="obj-title"><span class="obj-badge">2</span> Parametri porudžbine</div>', unsafe_allow_html=True)
            _oc1, _oc2, _oc3 = st.columns(3)
            with _oc1:
                _o_mes = st.number_input("Broj meseci za porudžbinu", min_value=0.5, value=1.5, step=0.5, format="%.1f", key="obj_mes_num", help="Porudžbina pokriva ovoliko meseci predviđene prodaje (može 1.0, 1.5, 2.0...).")
                _o_ml = st.text_input("Min. lager po artiklu", value="", placeholder="prazno = bez ograničenja", key="obj_ml")
                _o_mo = st.text_input("Min. ukupna porudžbina po objektu", value="", placeholder="prazno = bez ograničenja", key="obj_mo")
            with _oc2:
                _o_mpa = st.text_input("Min. kom po artiklu (po stavci)", value="", placeholder="prazno = bez ograničenja", key="obj_mpa")
                _o_maxpa = st.text_input("Maksimum po komadu (po stavci)", value="", placeholder="prazno = bez ograničenja", key="obj_maxpa")
                _o_tr = st.number_input("Ukupan trosak mkt (RSD)", min_value=0, value=0, step=10000, key="obj_tr")
            with _oc3:
                _o_excl = st.text_area("Isključeni komitenti (ID, zarez)", value=DEFAULT_EXCLUDED, height=110, key="obj_excl")
        _o_min_lager = int(_o_ml) if _o_ml.strip().isdigit() else None
        _o_min_order = int(_o_mo) if _o_mo.strip().isdigit() else None
        _o_min_pa = int(_o_mpa) if _o_mpa.strip().isdigit() else None
        _o_max_pa = int(_o_maxpa) if _o_maxpa.strip().isdigit() else None
        _o_excluded = set()
        for _part in _o_excl.replace('\n', ',').split(','):
            _p = _part.strip()
            if _p.isdigit():
                _o_excluded.add(int(_p))

        with st.container(border=True):
            st.markdown('<div class="obj-title"><span class="obj-badge green">\u2713</span> Objavi za koleginice</div>', unsafe_allow_html=True)
            _od1, _od2 = st.columns(2)
            with _od1:
                _osist = st.text_input("Naziv sistema (kako će koleginice videti)", value=os.path.splitext(up_o.name)[0].strip(), key="obj_sist")
            with _od2:
                st.text_input("Mesec porudžbine (automatski)", value=_mlbl, disabled=True, key="obj_mes_disp")
            if not sb_dostupan():
                st.info("Objava nije moguća dok Supabase nije podešen.")
            elif st.button("📤 Objavi za koleginice", use_container_width=True, key="obj_btn"):
                if not _osist.strip():
                    st.error("Upiši naziv sistema.")
                else:
                    try:
                        _pb = st.progress(0, "Računam porudžbinu...")
                        _eng = PredictionEngine(_obytes, _o_excluded, alpha, beta, _o_min_lager, _o_min_order, _o_tr, None, _o_min_pa, meseci=float(_o_mes), max_per_artikal=_o_max_pa)
                        _res = _eng.run(_pb)
                        _pb.empty()
                        _lg2, _lm2 = _eng.meseci_order[-1]
                        _om2 = int(_lm2) + 1; _oy2 = int(_lg2)
                        while _om2 > 12:
                            _om2 -= 12; _oy2 += 1
                        _mk2 = str(_oy2) + "-" + ("0" + str(_om2))[-2:]
                        _mlbl2 = mesec_label(_mk2)
                        _stavke = stavke_iz_rezultata(_res, _eng)
                        _payload = {"mesec_label": _mlbl2, "meta": {"pred_label": _eng.pred_label, "order_label": _eng.order_label, "min_lager": _eng.min_lager, "meseci": round(float(_o_mes), 1), "generisano": datetime.datetime.now().strftime("%d.%m.%Y %H:%M"), "n_objekata": int(len({_s2['idk'] for _s2 in _stavke})), "ukupno_kom": int(sum(_s2['kol'] for _s2 in _stavke))}, "stavke": _stavke}
                        try:
                            _payload["direktor"] = direktor_blok(_eng, _res)
                        except Exception:
                            _payload["direktor"] = None
                        _xb64 = None
                        try:
                            import base64 as _b64
                            _xbuf = create_excel(_eng, ukljuci_model=False)
                            _xb64 = _b64.b64encode(_xbuf.getvalue()).decode("ascii")
                        except Exception:
                            _xb64 = None
                        sb_objavi(_mk2, _osist, _payload, xlsx_b64=_xb64)
                        st.success(f"\u2705 Objavljeno: {_osist} \u2014 {_mlbl2} \u00b7 {len(_stavke)} stavki, {_payload['meta']['n_objekata']} objekata. Osveži (F5) da se ažurira lista gore.")
                    except Exception as _e:
                        st.error(f"Greška pri objavi: {_e}")
                        import traceback as _tb
                        st.code(_tb.format_exc())
    else:
        st.caption("\u2191 Učitaj Excel fajl da bi objavio porudžbinu za koleginice.")

with tab_ana:
    with st.expander("⚙️ Parametri analize", expanded=False):
        pc1, pc2, pc3 = st.columns(3)
        with pc1:
            st.markdown("**📦 Porudžbina**")
            meseci_ana = st.number_input("Broj meseci za porudžbinu", min_value=0.5, value=1.5, step=0.5, format="%.1f",
                                         help="Porudžbina pokriva ovoliko meseci predviđene prodaje (može 1.0, 1.5, 2.0...).")
            _ml_str = st.text_input("Minimalni lager po artiklu", value="", placeholder="prazno = bez ograničenja")
            min_lager = int(_ml_str) if _ml_str.strip().isdigit() else None
            _mo_str = st.text_input("Min. ukupna porudžbina po objektu", value="", placeholder="prazno = bez ograničenja")
            min_order = int(_mo_str) if _mo_str.strip().isdigit() else None
            _mpa_str = st.text_input("Min. kom po artiklu (po stavci)", value="", placeholder="prazno = bez ograničenja",
                                      help="Ako je porudžbina za jedan artikal manja od ovog broja (ali > 0), podiže se na minimum. Nule ostaju nule.")
            min_per_artikal = int(_mpa_str) if _mpa_str.strip().isdigit() else None
            _maxpa_str = st.text_input("Maksimum po komadu (po stavci)", value="", placeholder="prazno = bez ograničenja",
                                       help="Gornja granica porudžbine po jednom artiklu (po stavci). Ako je predlog veći, spušta se na ovaj maksimum.")
            max_per_artikal = int(_maxpa_str) if _maxpa_str.strip().isdigit() else None
        with pc2:
            st.markdown("**💰 Troškovi**")
            mesecni_trosak = st.number_input(
                "Ukupan trosak mkt/ulistavanja (RSD)",
                min_value=0, value=0, step=10000)
        with pc3:
            st.markdown("**⛔ Isključeni komitenti**")
            excluded_str = st.text_area("ID-evi razdvojeni zarezom", value=DEFAULT_EXCLUDED, height=80)
    excluded = set()
    for part in excluded_str.replace('\n', ',').split(','):
        p = part.strip()
        if p.isdigit(): excluded.add(int(p))
    uploaded = st.file_uploader("Učitaj Excel fajl sa podacima", type=['xlsx','xls'])
    if uploaded:
        file_bytes = uploaded.read()
        st.markdown(f'<div class="success-box">✅ Fajl <strong>{uploaded.name}</strong> učitan ({len(file_bytes)//1024} KB)</div>', unsafe_allow_html=True)
        st.markdown("")
        try:
            _xls = pd.ExcelFile(io.BytesIO(file_bytes))
            _sm = {s.strip().lower(): s for s in _xls.sheet_names}
            _sp = None
            for kw in ['prodaja']:
                for nl, no in _sm.items():
                    if kw in nl: _sp = no; break
            if _sp:
                _prod = pd.read_excel(_xls, sheet_name=_sp); _prod.columns=[c.strip() for c in _prod.columns]
                _meseci = sorted(_prod[['Godina','Mesec']].drop_duplicates().values.tolist())
                _mn={1:'Jan',2:'Feb',3:'Mar',4:'Apr',5:'Maj',6:'Jun',7:'Jul',8:'Avg',9:'Sep',10:'Okt',11:'Nov',12:'Dec'}
                _labels = [f"{_mn.get(int(m),'?' )} {int(g)}" for g,m in _meseci]
                st.markdown('**📅 Period za analizu** (OOS, Profitabilnost, Akcija — ne utiče na predikciju):')
                selected_labels = st.multiselect("Odaberi mesece", _labels, default=_labels, help="Predikcija uvek koristi sve mesece. Ovaj filter se odnosi samo na analitiku.")
                if not selected_labels:
                    st.warning("⚠️ Mora biti odabran bar jedan mesec za analizu. Automatski je odabran poslednji mesec.")
                    selected_labels = [_labels[-1]] if _labels else []
                selected_meseci = [_meseci[i] for i, lb in enumerate(_labels) if lb in selected_labels]
            else:
                selected_labels = []; selected_meseci = []
        except:
            selected_labels = []; selected_meseci = []
        if st.button("🚀 POKRENI ANALIZU", use_container_width=True):
            progress_bar = st.progress(0)
            try:
                engine = PredictionEngine(file_bytes, excluded, alpha, beta, min_lager, min_order, mesecni_trosak, selected_meseci, min_per_artikal, meseci=float(meseci_ana), max_per_artikal=max_per_artikal)
                result = engine.run(progress_bar)
                st.session_state["last_engine"] = engine
                st.session_state["last_result"] = result
                st.session_state["last_filename"] = uploaded.name
                st.markdown("---")
                tp = int(result['Predikcija'].sum()); tl = int(result['Lager_danas'].sum())
                t1 = int(result[~result['ID KOMITENTA'].isin(excluded)]['Porudzbina_1'].sum())
                t2 = int(result[~result['ID KOMITENTA'].isin(excluded)]['Porudzbina_2'].sum())
                if engine.has_prices:
                    tab1, tab2 = st.tabs(["📦 Porudžbina", "💰 Profitabilnost objekata & OOS"])
                else:
                    tab1, = st.tabs(["📦 Porudžbina"])
                with tab1:
                    n_obj_salji = int(result[result['Porudzbina_2'] > 0]['ID KOMITENTA'].nunique())
                    tp_prosek = int(result['Prosek'].sum())
                    m1,m2,m3,m4,m5 = st.columns(5)
                    m1.markdown(f'<div class="metric-card"><div class="metric-value">{tp:,}</div><div class="metric-label">Predikcija (kom)</div></div>', unsafe_allow_html=True)
                    m2.markdown(f'<div class="metric-card"><div class="metric-value">{tp_prosek:,}</div><div class="metric-label">Prosek (kom)</div></div>', unsafe_allow_html=True)
                    m3.markdown(f'<div class="metric-card"><div class="metric-value-green">{t2:,}</div><div class="metric-label">Porudžbina (kom)</div></div>', unsafe_allow_html=True)
                    m4.markdown(f'<div class="metric-card"><div class="metric-value">{n_obj_salji:,}</div><div class="metric-label">Objekata prima robu</div></div>', unsafe_allow_html=True)
                    m5.markdown(f'<div class="metric-card"><div class="metric-value">{tl:,}</div><div class="metric-label">Lager danas</div></div>', unsafe_allow_html=True)
                    st.markdown("")
                    st.markdown("<div style='margin:24px 0 4px 0;'></div>", unsafe_allow_html=True)
                    ml = engine.mesec_labels
                    df_r = engine.df_result.copy()
                    kom_mes = {}
                    for lb in ml:
                        col_lb = f'{lb}_Prodaja'
                        if col_lb in df_r.columns:
                            grp = df_r.groupby('ID KOMITENTA')[col_lb].sum()
                            for kid, v in grp.items():
                                if kid not in kom_mes: kom_mes[kid] = {}
                                kom_mes[kid][lb] = int(v)
                    import numpy as _np2
                    def _is_rastuci(vals5, dozvoljeni_sum=1):
                        padovi = sum(1 for i in range(1, len(vals5)) if vals5[i] < vals5[i-1])
                        return padovi <= dozvoljeni_sum and vals5[-1] > vals5[0] and sum(vals5) >= 10
                    def _is_padajuci(vals5, dozvoljeni_sum=1):
                        rasti = sum(1 for i in range(1, len(vals5)) if vals5[i] > vals5[i-1])
                        return rasti <= dozvoljeni_sum and vals5[-1] < vals5[0] and sum(vals5) >= 10
                    def _rast_pct(vals5):
                        first = vals5[0] if vals5[0] > 0 else 1
                        return (vals5[-1] - vals5[0]) / first * 100
                    rastuci_list = []
                    padajuci_list = []
                    for kid, mes_vals in kom_mes.items():
                        vals_all = [mes_vals.get(lb, 0) for lb in ml]
                        vals5 = vals_all[-5:] if len(vals_all) >= 5 else vals_all
                        if len(vals5) < 3: continue
                        if _is_rastuci(vals5):
                            rastuci_list.append({
                                'ID': kid, 'Ukupno': sum(vals_all),
                                'Vals': vals_all, 'Vals5': vals5,
                                'Rast': _rast_pct(vals5),
                                'Zadnji': vals5[-1], 'Prvi': vals5[0],
                            })
                        elif _is_padajuci(vals5):
                            padajuci_list.append({
                                'ID': kid, 'Ukupno': sum(vals_all),
                                'Vals': vals_all, 'Vals5': vals5,
                                'Pad': _rast_pct(vals5),
                                'Zadnji': vals5[-1], 'Prvi': vals5[0],
                            })
                    rastuci_list = sorted(rastuci_list, key=lambda x: x['Rast'], reverse=True)[:10]
                    padajuci_list = sorted(padajuci_list, key=lambda x: x['Pad'])[:10]
                    def _render_trend_section(title, icon, color, items, is_rast):
                        label_color = "#10b981" if is_rast else "#ef4444"
                        label_bg = "#f0fdf4" if is_rast else "#fef2f2"
                        if not items:
                            components.html(f"""<!DOCTYPE html><html><body style="margin:0;padding:4px 0;font-family:'DM Sans',sans-serif;">
                            <div style="display:flex;align-items:center;gap:8px;margin-bottom:12px;">
                                <span style="font-size:17px;">{icon}</span>
                                <span style="font-size:13px;font-weight:700;color:#111;">{title}</span>
                            </div>
                            <div style="color:#aaa;font-size:13px;padding:12px 0;">Nema podataka za prikaz</div>
                            </body></html>""", height=80)
                            return
                        rows_html = ""
                        for r in items:
                            vals5 = r['Vals5']
                            mx = max(vals5) if max(vals5) > 0 else 1
                            bars = "".join(
                                f'<div style="flex:1;display:flex;flex-direction:column;justify-content:flex-end;gap:0;">'
                                f'<div style="height:{int(v/mx*28)}px;background:{"linear-gradient(180deg,#a855f7,#c084fc)" if is_rast else "linear-gradient(180deg,#ec4899,#f9a8d4)"};border-radius:2px 2px 0 0;min-height:2px;"></div></div>'
                                for v in vals5
                            )
                            sign = "+" if is_rast else ""
                            pct = r['Rast'] if is_rast else r['Pad']
                            rows_html += f"""<div style="display:flex;align-items:center;gap:10px;padding:8px 0;border-bottom:1px solid #f3f4f6;">
                                <div style="font-family:'DM Mono',monospace;font-size:14px;font-weight:500;color:#111;width:46px;flex-shrink:0;">{int(r["ID"])}</div>
                                <div style="display:flex;align-items:flex-end;gap:2px;height:32px;width:90px;flex-shrink:0;">{bars}</div>
                                <div style="flex:1;font-size:11px;color:#aaa;">{int(r["Ukupno"]):,} kom</div>
                                <div style="font-size:12px;font-weight:700;color:{label_color};white-space:nowrap;">{sign}{pct:.0f}% &nbsp;<span style="font-weight:400;color:#bbb;font-size:11px;">({int(r["Prvi"])}→{int(r["Zadnji"])})</span></div>
                            </div>"""
                        h_px = len(items) * 48 + 56
                        components.html(f"""<!DOCTYPE html><html>
                        <head><link href="https://fonts.googleapis.com/css2?family=DM+Mono:wght@400;500&family=DM+Sans:wght@400;600;700&display=swap" rel="stylesheet"></head>
                        <body style="margin:0;padding:4px 0;font-family:'DM Sans',sans-serif;background:white;">
                            <div style="display:flex;align-items:center;gap:8px;margin-bottom:14px;">
                                <span style="font-size:17px;">{icon}</span>
                                <span style="font-size:13px;font-weight:700;color:#111;">{title}</span>
                                <span style="font-size:10px;font-weight:700;color:{label_color};background:{label_bg};border-radius:20px;padding:2px 8px;">zadnjih 5 mes.</span>
                            </div>
                            <div style="font-size:9px;color:#ccc;display:flex;gap:10px;margin-bottom:4px;">
                                <span style="width:46px;"></span>
                                <span style="width:90px;text-align:center;text-transform:uppercase;letter-spacing:.5px;">trend</span>
                                <span style="flex:1;text-transform:uppercase;letter-spacing:.5px;">ukupno</span>
                                <span style="text-transform:uppercase;letter-spacing:.5px;">rast (prvi→zadnji)</span>
                            </div>
                            {rows_html}
                        </body></html>""", height=h_px)
                    def _render_oos_section(items, max_val):
                        if not items:
                            components.html('''<!DOCTYPE html><html><body style="margin:0;padding:4px 0;font-family:sans-serif;">
                            <div style="display:flex;align-items:center;gap:8px;margin-bottom:12px;">
                                <span style="font-size:17px;">🔴</span>
                                <span style="font-size:13px;font-weight:700;color:#111;">OOS — Lager 0, najveći potencijal</span>
                            </div>
                            <div style="color:#aaa;font-size:13px;">Nema OOS podataka</div>
                            </body></html>''', height=80)
                            return
                        rows_html = ""
                        for r in items:
                            pct = int(r['Izgubljeno'] / max_val * 100)
                            rows_html += f"""<div style="padding:9px 0;border-bottom:1px solid #f9f9f9;">
                                <div style="display:flex;align-items:center;gap:10px;margin-bottom:5px;">
                                    <div style="font-family:'DM Mono',monospace;font-size:14px;font-weight:500;color:#111;width:46px;flex-shrink:0;">{int(r["ID KOMITENTA"])}</div>
                                    <div style="font-size:10px;font-weight:700;color:#ec4899;background:#fdf2f8;border-radius:4px;padding:2px 7px;">{int(r["Artikala"])} artikala bez robe</div>
                                    <div style="margin-left:auto;font-family:'DM Mono',monospace;font-size:13px;font-weight:700;color:#7c3aed;">{int(r["Izgubljeno"]):,} RSD</div>
                                </div>
                                <div style="height:5px;background:#f5f0ff;border-radius:99px;overflow:hidden;">
                                    <div style="width:{pct}%;height:100%;background:linear-gradient(90deg,#a855f7,#ec4899);border-radius:99px;"></div>
                                </div>
                            </div>"""
                        h_px = len(items) * 54 + 56
                        components.html(f"""<!DOCTYPE html><html>
                        <head><link href="https://fonts.googleapis.com/css2?family=DM+Mono:wght@400;500&family=DM+Sans:wght@400;600;700&display=swap" rel="stylesheet"></head>
                        <body style="margin:0;padding:4px 0;font-family:'DM Sans',sans-serif;background:white;">
                            <div style="display:flex;align-items:center;gap:8px;margin-bottom:14px;">
                                <span style="font-size:17px;">🔴</span>
                                <span style="font-size:13px;font-weight:700;color:#111;">OOS — Lager 0, najveći potencijal</span>
                                <span style="font-size:10px;font-weight:700;color:#ec4899;background:#fdf2f8;border-radius:20px;padding:2px 8px;">top 10</span>
                            </div>
                            <div style="font-size:9px;color:#ccc;display:flex;gap:10px;margin-bottom:4px;align-items:center;">
                                <span style="width:46px;"></span>
                                <span style="flex:1;text-transform:uppercase;letter-spacing:.5px;"></span>
                                <span style="text-transform:uppercase;letter-spacing:.5px;">izgubljen profit</span>
                            </div>
                            {rows_html}
                        </body></html>""", height=h_px)
                    col_rast, col_pad = st.columns(2)
                    with col_rast:
                        _render_trend_section("Rastući trendovi", "📈", "#a855f7", rastuci_list, True)
                    with col_pad:
                        _render_trend_section("Padajući trendovi", "📉", "#ec4899", padajuci_list, False)
                    st.markdown("<div style='margin:20px 0 4px 0;'></div>", unsafe_allow_html=True)
                    if engine.has_prices and len(engine.df_oos) > 0:
                        oos_k = engine.df_oos.copy()
                        if 'Lager_danas' in oos_k.columns:
                            oos_k = oos_k[oos_k['Lager_danas'] == 0]
                        oos_top = oos_k.groupby('ID KOMITENTA').agg(
                            Izgubljeno=('Izgubljeni_profit','sum'),
                            Artikala=('id artikla','nunique')
                        ).reset_index().sort_values('Izgubljeno', ascending=False).head(10)
                        oos_items = oos_top.to_dict('records')
                        oos_max = int(oos_top['Izgubljeno'].max()) if len(oos_top) > 0 else 1
                    else:
                        oos_items = []; oos_max = 1
                    col_oos2, col_empty = st.columns(2)
                    with col_oos2:
                        _render_oos_section(oos_items, oos_max)
                if engine.has_prices:
                    with tab2:
                        period_str2 = ", ".join(engine.analitika_labels) if engine.analitika_labels else "svi meseci"
                        n_mes = len(engine.analitika_labels) if engine.analitika_labels else len(engine.mesec_labels)
                        n_obj = engine.num_komitenti
                        prof = engine.df_profit_obj.copy()
                        total_bruto = int(prof['Bruto_profit'].sum())
                        total_neto = int(prof['Neto_profit'].sum())
                        total_trosak = int(prof['Trosak_mkt'].sum())
                        total_oos_izgubljen = int(engine.df_oos['Izgubljeni_profit'].sum()) if len(engine.df_oos) > 0 else 0
                        mes_trosak = total_trosak / max(n_mes, 1)
                        mes_bruto = total_bruto / max(n_mes, 1)
                        mes_neto = total_neto / max(n_mes, 1)
                        mes_oos = total_oos_izgubljen / max(n_mes, 1)
                        st.caption(f"📅 Period analize: **{period_str2}** · {n_obj} objekata · {n_mes} meseci")
                        ka, kb, kc, kd = st.columns(4)
                        def _kard(col, label, total, mes, color, prefix=""):
                            col.markdown(f"""
                            <div style="background:white;border-radius:12px;padding:16px 18px;
                                border-left:4px solid {color};box-shadow:0 2px 8px rgba(0,0,0,0.07);height:100%;">
                                <div style="font-size:10px;color:#999;font-weight:600;letter-spacing:.5px;text-transform:uppercase;margin-bottom:6px;">{label}</div>
                                <div style="font-size:22px;font-weight:700;color:{color};">{prefix}{total:,.0f} RSD</div>
                                <div style="font-size:11px;color:#aaa;margin-top:3px;">{prefix}{mes:,.0f} RSD / mesec</div>
                            </div>""", unsafe_allow_html=True)
                        _kard(ka, f"Ukupan trosak · {n_mes} meseci", total_trosak, mes_trosak, "#a855f7")
                        _kard(kb, f"Bruto profit · {n_mes} meseci", total_bruto, mes_bruto, "#10b981")
                        _kard(kc, f"Neto profit · {n_mes} meseci", total_neto, mes_neto, "#7c3aed" if total_neto > 0 else "#ec4899")
                        _kard(kd, f"OOS izgubljen · {n_mes} meseci", total_oos_izgubljen, mes_oos, "#ec4899", prefix="-")
                        st.markdown("<div style='margin:20px 0 4px 0;'></div>", unsafe_allow_html=True)
                        a_labels_trend = engine.analitika_labels if engine.analitika_labels else engine.mesec_labels
                        a_meseci_trend = engine.analitika_meseci if (engine.analitika_meseci and len(engine.analitika_meseci) > 0) else engine.meseci_order
                        bruto_po_mes = []
                        neto_po_mes = []
                        for i, lb in enumerate(a_labels_trend):
                            col_bruto = f'Bruto_{lb}'
                            col_neto = f'Neto_{lb}'
                            bruto_val = prof[col_bruto].sum() if col_bruto in prof.columns else 0
                            neto_val = prof[col_neto].sum() if col_neto in prof.columns else 0
                            bruto_po_mes.append((lb, bruto_val))
                            neto_po_mes.append((lb, neto_val))
                        def _trend_recenica(podaci, naziv):
                            vals = [v for _, v in podaci]
                            if len(vals) < 2: return ""
                            prvi_lb, prvi_v = podaci[0]
                            posl_lb, posl_v = podaci[-1]
                            if prvi_v == 0: return ""
                            promena_pct = ((posl_v - prvi_v) / abs(prvi_v)) * 100
                            smer = "porastao" if promena_pct > 0 else "pao"
                            boja = "#10b981" if promena_pct > 0 else "#ec4899"
                            return f'<span style="color:{boja};font-weight:600;">{naziv} je {smer} za {abs(promena_pct):.0f}%</span> — od <b>{prvi_v:,.0f} RSD</b> ({prvi_lb}) do <b>{posl_v:,.0f} RSD</b> ({posl_lb}).'
                        def _bar_chart_html(podaci, max_val, color_pos, color_neg):
                            bars = ""
                            for lb, val in podaci:
                                pct = abs(val) / max_val * 100 if max_val > 0 else 0
                                pct = min(pct, 100)
                                color = color_pos if val >= 0 else color_neg
                                val_fmt = f"{val:,.0f} RSD"
                                bars += f"""
                                <div style="display:flex;align-items:center;margin-bottom:5px;gap:8px;">
                                    <div style="width:52px;font-size:11px;color:#888;text-align:right;flex-shrink:0;">{lb}</div>
                                    <div style="flex:1;background:#f5f0ff;border-radius:3px;height:18px;position:relative;">
                                        <div style="width:{pct:.1f}%;background:{color};height:100%;border-radius:3px;transition:width .3s;"></div>
                                    </div>
                                    <div style="width:110px;font-size:11px;color:#555;font-weight:600;flex-shrink:0;">{val_fmt}</div>
                                </div>"""
                            return f'<div style="padding:4px 0;">{bars}</div>'
                        max_bruto = max(abs(v) for _, v in bruto_po_mes) if bruto_po_mes else 1
                        max_neto = max(abs(v) for _, v in neto_po_mes) if neto_po_mes else 1
                        col_bruto, col_neto = st.columns(2)
                        with col_bruto:
                            st.markdown('<div class="section-title">📈 Mesečni trend bruto profita</div>', unsafe_allow_html=True)
                            rec_b = _trend_recenica(bruto_po_mes, "Bruto profit")
                            if rec_b: st.markdown(f'<p style="font-size:13px;color:#555;margin-bottom:6px;">{rec_b}</p>', unsafe_allow_html=True)
                            chart_b = _bar_chart_html(bruto_po_mes, max_bruto, "#a855f7", "#ec4899")
                            components.html(f'<!DOCTYPE html><html><body style="margin:0;padding:8px 12px;font-family:sans-serif;">{chart_b}</body></html>', height=len(bruto_po_mes)*28+20)
                        with col_neto:
                            st.markdown('<div class="section-title">📉 Mesečni trend neto profita</div>', unsafe_allow_html=True)
                            rec_n = _trend_recenica(neto_po_mes, "Neto profit")
                            if rec_n: st.markdown(f'<p style="font-size:13px;color:#555;margin-bottom:6px;">{rec_n}</p>', unsafe_allow_html=True)
                            chart_n = _bar_chart_html(neto_po_mes, max_neto, "#7c3aed", "#ec4899")
                            components.html(f'<!DOCTYPE html><html><body style="margin:0;padding:8px 12px;font-family:sans-serif;">{chart_n}</body></html>', height=len(neto_po_mes)*28+20)
                        st.markdown("<div style='margin:20px 0 4px 0;'></div>", unsafe_allow_html=True)
                        st.markdown('<div class="section-title">🏪 Profitabilnost po objektima</div>', unsafe_allow_html=True)
                        ukupno_obj = len(prof)
                        neto_neg = prof[prof['Neto_profit'] <= 0]
                        n_neto_neg = len(neto_neg)
                        oos_neg = prof[(prof['Neto_profit'] <= 0) & (prof['Potencijalni_profit'] > 0)]
                        n_oos_neg = len(oos_neg)
                        pravi_neg = prof[(prof['Neto_profit'] <= 0) & (prof['Potencijalni_profit'] <= 0)]
                        n_pravi_neg = len(pravi_neg)
                        pct_pravi = round(n_pravi_neg / max(ukupno_obj, 1) * 100)
                        trosak_po_obj = engine.trosak_po_objektu
                        trosak_mes_obj = trosak_po_obj / max(n_mes, 1)
                        usteda_trosak = n_pravi_neg * trosak_po_obj
                        usteda_gubitak = abs(pravi_neg['Neto_profit'].sum()) if n_pravi_neg > 0 else 0
                        usteda_mes = (usteda_trosak + usteda_gubitak) / max(n_mes, 1)
                        n_profitabilni = ukupno_obj - n_neto_neg
                        pct_prof = n_profitabilni / max(ukupno_obj, 1)
                        pct_oos_neg_v = n_oos_neg / max(ukupno_obj, 1)
                        pct_pravi_v = n_pravi_neg / max(ukupno_obj, 1)
                        cx, cy, r_out, r_in = 110, 110, 90, 60
                        def _arc_path(cx, cy, r, start_deg, end_deg):
                            s = math.radians(start_deg - 90)
                            e = math.radians(end_deg - 90)
                            large = 1 if (end_deg - start_deg) > 180 else 0
                            x1,y1 = cx+r*math.cos(s), cy+r*math.sin(s)
                            x2,y2 = cx+r*math.cos(e), cy+r*math.sin(e)
                            return f"M {x1:.1f} {y1:.1f} A {r} {r} 0 {large} 1 {x2:.1f} {y2:.1f}"
                        def _donut_seg(cx, cy, ro, ri, start_deg, end_deg, color):
                            if end_deg - start_deg < 0.5: return ""
                            oa = _arc_path(cx, cy, ro, start_deg, end_deg)
                            s2 = math.radians(end_deg - 90); s1 = math.radians(start_deg - 90)
                            x_ie, y_ie = cx+ri*math.cos(s2), cy+ri*math.sin(s2)
                            x_is, y_is = cx+ri*math.cos(s1), cy+ri*math.sin(s1)
                            large = 1 if (end_deg - start_deg) > 180 else 0
                            x2o,y2o = cx+ro*math.cos(s2), cy+ro*math.sin(s2)
                            x1o,y1o = cx+ro*math.cos(s1), cy+ro*math.sin(s1)
                            return f'<path d="{oa} L {x_ie:.1f} {y_ie:.1f} A {ri} {ri} 0 {large} 0 {x_is:.1f} {y_is:.1f} Z" fill="{color}"/>'
                        deg_prof = pct_prof * 360
                        deg_oos = pct_oos_neg_v * 360
                        deg_pravi = pct_pravi_v * 360
                        seg1 = _donut_seg(cx, cy, r_out, r_in, 0, deg_prof, "#10b981")
                        seg2 = _donut_seg(cx, cy, r_out, r_in, deg_prof, deg_prof+deg_pravi, "#ec4899")
                        seg3 = _donut_seg(cx, cy, r_out, r_in, deg_prof+deg_pravi, deg_prof+deg_pravi+deg_oos, "#a855f7")
                        donut_svg = f"""<svg width="220" height="220" xmlns="http://www.w3.org/2000/svg">
                            {seg1}{seg2}{seg3}
                            <circle cx="{cx}" cy="{cy}" r="{r_in}" fill="white"/>
                            <text x="{cx}" y="{cy-8}" text-anchor="middle" font-size="26" font-weight="700" fill="#111" font-family="sans-serif">{n_profitabilni}</text>
                            <text x="{cx}" y="{cy+14}" text-anchor="middle" font-size="12" fill="#888" font-family="sans-serif">profitabilnih</text>
                        </svg>
                        <div style="margin-top:8px;font-size:12px;font-family:sans-serif;">
                            <div style="display:flex;align-items:center;gap:6px;margin-bottom:5px;">
                                <span style="width:12px;height:12px;background:#10b981;border-radius:2px;display:inline-block;flex-shrink:0;"></span>
                                <span style="color:#555;"><strong>{n_profitabilni} profitabilnih</strong> ({round(pct_prof*100)}% mreže)</span>
                            </div>
                            <div style="display:flex;align-items:center;gap:6px;margin-bottom:5px;">
                                <span style="width:12px;height:12px;background:#ec4899;border-radius:2px;display:inline-block;flex-shrink:0;"></span>
                                <span style="color:#555;"><strong>{n_pravi_neg} neprofitabilnih</strong> ({round(pct_pravi_v*100)}% mreže)</span>
                            </div>
                            <div style="display:flex;align-items:center;gap:6px;">
                                <span style="width:12px;height:12px;background:#a855f7;border-radius:2px;display:inline-block;flex-shrink:0;"></span>
                                <span style="color:#555;"><strong>{n_oos_neg} neto-neg. OOS</strong> potencijal</span>
                            </div>
                        </div>"""
                        tekst = f"""
    <div style="background:white;border-radius:12px;padding:20px 24px;box-shadow:0 2px 8px rgba(0,0,0,0.06);margin-bottom:16px;font-size:14px;line-height:1.8;color:#333;">
    <p>Od <strong>{ukupno_obj} objekata</strong>, <strong>{n_neto_neg}</strong> je neto negativno.
    Medjutim, <strong>{n_oos_neg}</strong> od njih ima negativan neto isključivo zbog OOS-a — kada se uračuna izgubljena zarada,
    njihov potencijal je pozitivan. Ovi objekti nisu problem, samo nisu imali robu.</p>
    <p>Pravih neprofitabilnih je <strong>{n_pravi_neg}</strong> ({pct_pravi}% ukupne mreže) — negativni čak i po potencijalu.
    Trošak po objektu je <strong>{trosak_po_obj:,.0f} RSD</strong> za {n_mes} {'mesec' if n_mes==1 else 'meseci'} /
    <strong>{trosak_mes_obj:,.0f} RSD</strong> mesečno.</p>
    <p>Zatvaranjem <strong>{n_pravi_neg} pravih neprofitabilnih</strong> skidamo trošak
    <strong>{n_pravi_neg} × {trosak_po_obj:,.0f} RSD = {usteda_trosak:,.0f} RSD</strong>
    ({usteda_trosak/max(n_mes,1):,.0f} RSD/mes) i prestajemo da gubimo
    <strong>{usteda_gubitak:,.0f} RSD</strong> ({usteda_gubitak/max(n_mes,1):,.0f} RSD/mes) na negativnim objektima.
    Ostaju samo objekti koji su u plusu.</p>
    </div>"""
                        col_tekst, col_donut = st.columns([3, 1])
                        with col_tekst:
                            st.markdown(tekst, unsafe_allow_html=True)
                        with col_donut:
                            components.html(f"""<!DOCTYPE html><html><body style="margin:0;padding:12px 8px;font-family:sans-serif;background:transparent;">
                                {donut_svg}
                            </body></html>""", height=310)
                        a_labels_trend2 = engine.analitika_labels if engine.analitika_labels else engine.mesec_labels
                        a_meseci_trend2 = engine.analitika_meseci if (engine.analitika_meseci and len(engine.analitika_meseci) > 0) else engine.meseci_order
                        chart_mes_data = []
                        for i, (lb, (g, m)) in enumerate(zip(a_labels_trend2, a_meseci_trend2)):
                            col_neto_lb = f'Neto_{lb}'
                            if col_neto_lb in prof.columns:
                                n_prof_mes = int((prof[col_neto_lb] > 0).sum())
                                n_nepr_mes = int((prof[col_neto_lb] <= 0).sum())
                            else:
                                n_prof_mes = 0; n_nepr_mes = 0
                            chart_mes_data.append((lb, n_prof_mes, n_nepr_mes))
                        if chart_mes_data:
                            max_obj_mes = max(a + b for _, a, b in chart_mes_data) if chart_mes_data else 1
                            bar_w = max(30, min(60, 700 // max(len(chart_mes_data), 1)))
                            bars_html = ""
                            for lb, np_v, nn_v in chart_mes_data:
                                h_p = int(np_v / max(max_obj_mes, 1) * 140)
                                h_n = int(nn_v / max(max_obj_mes, 1) * 140)
                                bars_html += f"""
                                <div style="display:flex;flex-direction:column;align-items:center;gap:2px;">
                                    <div style="display:flex;align-items:flex-end;gap:3px;height:160px;">
                                        <div style="width:{bar_w}px;height:{h_p}px;background:#a855f7;border-radius:3px 3px 0 0;position:relative;">
                                            <span style="position:absolute;top:-18px;left:50%;transform:translateX(-50%);font-size:10px;font-weight:700;color:#7c3aed;white-space:nowrap;">{np_v}</span>
                                        </div>
                                        <div style="width:{bar_w}px;height:{h_n}px;background:#ec4899;border-radius:3px 3px 0 0;position:relative;">
                                            <span style="position:absolute;top:-18px;left:50%;transform:translateX(-50%);font-size:10px;font-weight:700;color:#be185d;white-space:nowrap;">{nn_v}</span>
                                        </div>
                                    </div>
                                    <div style="font-size:10px;color:#888;margin-top:4px;text-align:center;width:{bar_w*2+3}px;">{lb}</div>
                                </div>"""
                            chart_html = f"""<!DOCTYPE html><html><body style="margin:0;padding:0;font-family:sans-serif;background:white;">
                            <div style="padding:16px 20px;">
                                <div style="display:flex;gap:16px;margin-bottom:14px;">
                                    <span style="display:flex;align-items:center;gap:5px;font-size:12px;color:#555;">
                                        <span style="width:12px;height:12px;background:#a855f7;border-radius:2px;display:inline-block;"></span> Profitabilni taj mesec (neto &gt; 0)
                                    </span>
                                    <span style="display:flex;align-items:center;gap:5px;font-size:12px;color:#555;">
                                        <span style="width:12px;height:12px;background:#ec4899;border-radius:2px;display:inline-block;"></span> Neprofitabilni taj mesec (neto ≤ 0)
                                    </span>
                                </div>
                                <div style="display:flex;gap:6px;align-items:flex-end;overflow-x:auto;padding-bottom:4px;">
                                    {bars_html}
                                </div>
                            </div>
                            </body></html>"""
                            components.html(chart_html, height=220)
                            st.markdown('''<p style="font-size:12px;color:#9ca3af;margin-top:4px;">
                            ℹ️ Grafikon prikazuje profitabilnost po potencijalu <strong>za svaki mesec posebno</strong> — razlikuje se od ukupnih brojeva iznad koji se odnose na <strong>ceo analizirani period</strong>. Na primer, objekat koji je u poslednjem mesecu neprofitabilan može biti profitabilan gledano kroz ceo period.
                            </p>''', unsafe_allow_html=True)
                        st.markdown("<div style='margin:20px 0 4px 0;'></div>", unsafe_allow_html=True)
                        st.markdown('<div class="section-title">🔴 OOS — Izgubljena zarada zbog nedostatka robe</div>', unsafe_allow_html=True)
                        if len(engine.df_oos) > 0:
                            a_labels_oos = engine.analitika_labels if engine.analitika_labels else engine.mesec_labels
                            oos_ukupno = int(engine.df_oos['Izgubljeni_profit'].sum())
                            oos_mes_avg = oos_ukupno // max(n_mes, 1)
                            oos_kombinacija = int((engine.df_oos['OOS_meseci'] > 0).sum()) if 'OOS_meseci' in engine.df_oos.columns else len(engine.df_oos)
                            oos_0_danas = int((engine.df_oos.get('Lager_danas', 0) == 0).sum()) if 'Lager_danas' in engine.df_oos.columns else oos_kombinacija
                            o1, o2, o3 = st.columns(3)
                            def _oos_kard(col, label, val, suffix=""):
                                col.markdown(f"""<div style="background:white;border-radius:12px;padding:16px 18px;
                                    border-top:3px solid #ec4899;box-shadow:0 2px 8px rgba(0,0,0,0.07);text-align:center;">
                                    <div style="font-size:22px;font-weight:700;color:#ec4899;">{val:,}{suffix}</div>
                                    <div style="font-size:11px;color:#aaa;margin-top:4px;text-transform:uppercase;letter-spacing:.5px;">{label}</div>
                                </div>""", unsafe_allow_html=True)
                            _oos_kard(o1, f"Izgubljen profit · {n_mes} meseci (RSD)", oos_ukupno)
                            _oos_kard(o2, "Prosečno mesečno (RSD)", oos_mes_avg)
                            _oos_kard(o3, "Kombinacija na 0 lagera danas", oos_0_danas)
                            st.markdown("<div style='margin:18px 0 4px 0;'></div>", unsafe_allow_html=True)
                            mes_izgub = []
                            mes_oos_count = []
                            for lb in a_labels_oos:
                                col_izgub = f'Izgub_{lb}'
                                col_oos = f'OOS_{lb}'
                                v_izgub = int(engine.df_oos[col_izgub].sum()) if col_izgub in engine.df_oos.columns else 0
                                v_oos = int((engine.df_oos[col_oos] > 0).sum()) if col_oos in engine.df_oos.columns else 0
                                mes_izgub.append(v_izgub)
                                mes_oos_count.append(v_oos)
                            if any(v > 0 for v in mes_izgub):
                                max_izgub = max(mes_izgub) if mes_izgub else 1
                                chart_w = 860
                                chart_h = 220
                                pad_l, pad_r, pad_t, pad_b = 60, 20, 30, 40
                                plot_w = chart_w - pad_l - pad_r
                                plot_h = chart_h - pad_t - pad_b
                                n_pts = len(a_labels_oos)
                                def px(i): return pad_l + int(i / max(n_pts-1,1) * plot_w)
                                def py(v): return pad_t + plot_h - int(v / max(max_izgub,1) * plot_h)
                                pts_area = " ".join(f"{px(i)},{py(v)}" for i, v in enumerate(mes_izgub))
                                pts_area = f"{px(0)},{pad_t+plot_h} " + pts_area + f" {px(n_pts-1)},{pad_t+plot_h}"
                                pts_line = " ".join(f"{px(i)},{py(v)}" for i, v in enumerate(mes_izgub))
                                dots = ""
                                labels_svg = ""
                                x_labels = ""
                                for i, (lb, v, vc) in enumerate(zip(a_labels_oos, mes_izgub, mes_oos_count)):
                                    x, y = px(i), py(v)
                                    v_k = f"{v//1000}k" if v >= 1000 else str(v)
                                    dots += f'<circle cx="{x}" cy="{y}" r="5" fill="#a855f7" stroke="white" stroke-width="2"/>'
                                    labels_svg += f'<text x="{x}" y="{y-10}" text-anchor="middle" font-size="10" font-weight="700" fill="#7c3aed">{v_k}</text>'
                                    labels_svg += f'<text x="{x}" y="{y+20}" text-anchor="middle" font-size="9" fill="#999">({vc})</text>'
                                    x_labels += f'<text x="{x}" y="{chart_h-6}" text-anchor="middle" font-size="9" fill="#aaa">{lb}</text>'
                                svg = f"""<svg width="{chart_w}" height="{chart_h}" xmlns="http://www.w3.org/2000/svg" style="font-family:sans-serif;">
                                    <text x="{pad_l-5}" y="{pad_t-8}" font-size="10" fill="#888">Izgubljen profit (RSD)</text>
                                    <text x="{chart_w-pad_r}" y="{pad_t-8}" font-size="10" fill="#aaa" text-anchor="end">Broj OOS kombinacija u zagradama</text>
                                    <polygon points="{pts_area}" fill="#a855f7" fill-opacity="0.08"/>
                                    <polyline points="{pts_line}" fill="none" stroke="#a855f7" stroke-width="2.5"/>
                                    {dots}{labels_svg}{x_labels}
                                </svg>"""
                                components.html(f'<!DOCTYPE html><html><body style="margin:0;padding:0;background:white;">{svg}</body></html>', height=chart_h+10)
                            oos_art = engine.df_oos.groupby(['id artikla','Naziv artikla']).agg(
                                Izgubljeni_profit=('Izgubljeni_profit','sum')
                            ).reset_index().sort_values('Izgubljeni_profit', ascending=False).head(5)
                            bar_colors = ["#a855f7","#ec4899","#7c3aed","#c084fc","#f472b6"]
                            top5_max = int(oos_art['Izgubljeni_profit'].max()) if len(oos_art) > 0 else 1
                            bars5 = ""
                            for i, (_, row) in enumerate(oos_art.iterrows()):
                                naziv = str(row['Naziv artikla'])[:35]
                                val = int(row['Izgubljeni_profit'])
                                pct = val / top5_max * 100
                                color = bar_colors[i % len(bar_colors)]
                                bars5 += f"""
                                <div style="display:flex;align-items:center;gap:10px;margin-bottom:10px;">
                                    <div style="width:200px;font-size:12px;color:#444;text-align:right;flex-shrink:0;">{naziv}</div>
                                    <div style="flex:1;background:#f5f0ff;border-radius:4px;height:22px;position:relative;">
                                        <div style="width:{pct:.1f}%;background:{color};height:100%;border-radius:4px;"></div>
                                    </div>
                                    <div style="width:110px;font-size:12px;font-weight:700;color:{color};flex-shrink:0;">{val:,} RSD</div>
                                </div>"""
                            st.markdown("**Top 5 artikala po izgubljenom profitu:**")
                            components.html(f"""<!DOCTYPE html><html><body style="margin:0;padding:8px 12px;font-family:sans-serif;background:white;">
                                {bars5}
                            </body></html>""", height=len(oos_art)*42+20)
                            with st.expander("📋 Svi artikli po izgubljenom profitu"):
                                oos_art_all = engine.df_oos.groupby(['id artikla','Naziv artikla']).agg(
                                    Objekata=('ID KOMITENTA','nunique'),
                                    OOS_meseci=('OOS_meseci','sum'),
                                    Izgubljeni_profit=('Izgubljeni_profit','sum')
                                ).reset_index().sort_values('Izgubljeni_profit', ascending=False)
                                oos_art_all.columns = ['ID Art.','Naziv','Objekata','OOS meseci','Izg. profit (RSD)']
                                st.dataframe(oos_art_all, use_container_width=True, height=300)
                        else:
                            st.success("Nema OOS problema!")
                        st.markdown("<div style='margin:24px 0 4px 0;'></div>", unsafe_allow_html=True)
                        st.markdown('<div class="section-title">⚡ Scenario: Optimalna mreža</div>', unsafe_allow_html=True)
                        prof2 = engine.df_profit_obj.copy()
                        oos_ukupno2 = int(engine.df_oos['Izgubljeni_profit'].sum()) if len(engine.df_oos) > 0 else 0
                        pozitivni = prof2[prof2['Potencijalni_profit'] > 0]
                        neto_pozitivnih = int(pozitivni['Neto_profit'].sum())
                        pravi_neg2 = prof2[(prof2['Neto_profit'] <= 0) & (prof2['Potencijalni_profit'] <= 0)]
                        n_pravi_neg2 = len(pravi_neg2)
                        usteda_trosak2 = int(n_pravi_neg2 * engine.trosak_po_objektu)
                        usteda_gubitak2 = int(abs(pravi_neg2['Neto_profit'].sum()))
                        ukupni_potencijal = neto_pozitivnih + usteda_trosak2 + usteda_gubitak2 + oos_ukupno2
                        stvarni_neto = int(prof2['Neto_profit'].sum())
                        razlika = ukupni_potencijal - stvarni_neto
                        period_sc = period_str2
                        def _red(label, val, color="#10b981", bold_val=True):
                            val_str = f"+{val:,} RSD" if val >= 0 else f"{val:,} RSD"
                            v_style = f"font-weight:{'700' if bold_val else '400'};color:{color};"
                            return f"""<div style="display:flex;justify-content:space-between;align-items:center;
                                padding:8px 0;border-bottom:1px solid #f3f4f6;">
                                <span style="font-size:13px;color:#555;">{label}</span>
                                <span style="{v_style}font-size:13px;">{val_str}</span>
                            </div>"""
                        def _red_bold(label, val, color="#111"):
                            val_str = f"= {val:,} RSD"
                            return f"""<div style="display:flex;justify-content:space-between;align-items:center;
                                padding:10px 0;border-top:2px solid #e5e7eb;margin-top:4px;">
                                <span style="font-size:14px;font-weight:700;color:#111;">{label}</span>
                                <span style="font-size:14px;font-weight:700;color:{color};">{val_str}</span>
                            </div>"""
                        scenario_html = f"""
                        <div style="background:white;border-radius:12px;padding:20px 24px;
                            box-shadow:0 2px 8px rgba(0,0,0,0.07);font-family:sans-serif;">
                            <div style="font-size:12px;font-weight:600;color:#a855f7;margin-bottom:12px;
                                text-transform:uppercase;letter-spacing:.5px;">
                                Period: {period_sc} ({n_mes} meseci)
                            </div>
                            <p style="font-size:13px;color:#666;margin-bottom:14px;">
                                Ako se istovremeno zatvore neprofitabilni objekti i eliminiše OOS, mreža ide sa
                                <strong>{stvarni_neto:,} RSD</strong> neto profita na
                                <strong style="color:#10b981;">+{ukupni_potencijal:,} RSD</strong> za {n_mes} meseci.
                            </p>
                            {_red(f"Neto profit pozitivnih objekata (potencijal > 0)", neto_pozitivnih, "#10b981")}
                            {_red(f"Ušteda: zatvaranje {n_pravi_neg2} neprofitabilnih obj.", usteda_trosak2 + usteda_gubitak2, "#10b981")}
                            {_red(f"Povraćaj izgub. zarade (OOS eliminacija)", oos_ukupno2, "#10b981")}
                            {_red_bold(f"UKUPNI POTENCIJAL ({n_mes} meseci)", ukupni_potencijal, "#10b981")}
                            <div style="height:8px;"></div>
                            {_red(f"Stvarni neto profit ({n_mes} meseci)", stvarni_neto, "#555", False)}
                            {_red(f"Razlika — potencijal koji još nije ostvaren", razlika, "#a855f7")}
                        </div>"""
                        st.markdown(scenario_html, unsafe_allow_html=True)
                        if engine.region_map:
                            st.markdown("<div style='margin:28px 0 6px 0;'></div>", unsafe_allow_html=True)
                            st.markdown('<div class="section-title">🗺️ Profitabilnost po okruzima</div>', unsafe_allow_html=True)
                            prof_reg = prof.copy()
                            prof_reg['Region'] = prof_reg['ID KOMITENTA'].map(engine.region_map).fillna('Ostalo')
                            prof_reg['Profitabilan'] = prof_reg['Neto_profit'] > 0
                            reg_grp = prof_reg.groupby('Region').agg(
                                Ukupno=('ID KOMITENTA','count'),
                                Ostaje=('Profitabilan','sum'),
                            ).reset_index()
                            reg_grp['Zatvara'] = reg_grp['Ukupno'] - reg_grp['Ostaje']
                            reg_grp = reg_grp.sort_values('Ukupno', ascending=False).reset_index(drop=True)
                            mali_okruzi_df = reg_grp[reg_grp['Ostaje'] < 5]
                            mali_okruzi = mali_okruzi_df['Region'].tolist()
                            rows_html = ""
                            for _, r in reg_grp.iterrows():
                                okrug = r['Region']
                                ukupno = int(r['Ukupno'])
                                ostaje = int(r['Ostaje'])
                                zatvara = int(r['Zatvara'])
                                mali = " *" if okrug in mali_okruzi else ""
                                mali_color = "#a855f7" if mali else "#111"
                                pct_o = ostaje / max(ukupno, 1) * 100
                                pct_z = zatvara / max(ukupno, 1) * 100
                                bar = f"""<div style="display:flex;width:120px;height:14px;border-radius:3px;overflow:hidden;">
                                    <div style="width:{pct_o:.0f}%;background:#a855f7;"></div>
                                    <div style="width:{pct_z:.0f}%;background:#ec4899;"></div>
                                </div>"""
                                rows_html += f"""<tr style="border-bottom:1px solid #f3f4f6;">
                                    <td style="padding:7px 10px;font-size:13px;color:{mali_color};font-weight:600;">{okrug}{mali}</td>
                                    <td style="padding:7px 10px;font-size:13px;font-weight:700;text-align:center;">{ukupno}</td>
                                    <td style="padding:7px 10px;font-size:13px;text-align:center;">
                                        <span style="color:#a855f7;font-weight:700;">{ostaje}</span>
                                        <span style="color:#999;"> / </span>
                                        <span style="color:#ec4899;font-weight:700;">{zatvara}</span>
                                    </td>
                                    <td style="padding:7px 16px;">{bar}</td>
                                </tr>"""
                            uk_ukupno = int(reg_grp['Ukupno'].sum())
                            uk_ostaje = int(reg_grp['Ostaje'].sum())
                            uk_zatvara = int(reg_grp['Zatvara'].sum())
                            rows_html += f"""<tr style="border-top:2px solid #e5e7eb;background:#f9fafb;">
                                <td style="padding:9px 10px;font-size:13px;font-weight:700;">UKUPNO</td>
                                <td style="padding:9px 10px;font-size:13px;font-weight:700;text-align:center;">{uk_ukupno}</td>
                                <td style="padding:9px 10px;font-size:13px;text-align:center;">
                                    <span style="color:#a855f7;font-weight:700;">{uk_ostaje}</span>
                                    <span style="color:#999;"> / </span>
                                    <span style="color:#ec4899;font-weight:700;">{uk_zatvara}</span>
                                </td>
                                <td></td>
                            </tr>"""
                            header_html = """<tr style="background:#f9fafb;border-bottom:2px solid #e5e7eb;">
                                <th style="padding:9px 10px;font-size:11px;color:#888;font-weight:600;text-align:left;text-transform:uppercase;letter-spacing:.4px;">Okrug</th>
                                <th style="padding:9px 10px;font-size:11px;color:#888;font-weight:600;text-align:center;text-transform:uppercase;letter-spacing:.4px;">Ukupno obj.</th>
                                <th style="padding:9px 10px;font-size:11px;color:#888;font-weight:600;text-align:center;text-transform:uppercase;letter-spacing:.4px;">✓ Ostaje / ✗ Zatvara</th>
                                <th style="padding:9px 10px;font-size:11px;color:#888;font-weight:600;text-transform:uppercase;letter-spacing:.4px;"></th>
                            </tr>"""
                            tbl_height = len(reg_grp) * 34 + 80
                            components.html(f"""<!DOCTYPE html><html><body style="margin:0;padding:0;font-family:sans-serif;background:white;">
                            <table style="width:100%;border-collapse:collapse;">
                                <thead>{header_html}</thead>
                                <tbody>{rows_html}</tbody>
                            </table>
                            </body></html>""", height=tbl_height)
                            if mali_okruzi:
                                mali_str = ", ".join(mali_okruzi)
                                st.markdown(f'<div style="font-size:12px;color:#a855f7;padding:6px 4px;">* Okruzi sa manje od 5 profitabilnih objekata ({mali_str}): Ne preporučuje se angazovanje komercijalistu isključivo za ove okruge — broj preostalih objekata premali je da bi opravdao redovne obilaske.</div>', unsafe_allow_html=True)
                            if len(mali_okruzi_df) > 0:
                                st.markdown("<div style='margin:20px 0 6px 0;'></div>", unsafe_allow_html=True)
                                prof_reg_mali = prof_reg[prof_reg['Region'].isin(mali_okruzi) & (prof_reg['Neto_profit'] > 0)]
                                n_mali_prof = len(prof_reg_mali)
                                neto_mali_prof = int(prof_reg_mali['Neto_profit'].sum())
                                usteda_mali_trosak = int(n_mali_prof * engine.trosak_po_objektu)
                                scA_potencijal = ukupni_potencijal
                                scB_potencijal = scA_potencijal + usteda_mali_trosak - neto_mali_prof
                                period_label = period_str2
                                def _sc_red(label, val, color="#555", bold=False):
                                    sign = "+" if val >= 0 else ""
                                    fw = "700" if bold else "400"
                                    return f"""<div style="display:flex;justify-content:space-between;padding:7px 0;border-bottom:1px solid #f3f4f6;">
                                        <span style="font-size:13px;color:#555;">{label}</span>
                                        <span style="font-size:13px;font-weight:{fw};color:{color};">{sign}{val:,} RSD</span>
                                    </div>"""
                                def _sc_total(label, val, color="#10b981"):
                                    return f"""<div style="display:flex;justify-content:space-between;padding:9px 0;border-top:2px solid #e5e7eb;margin-top:4px;">
                                        <span style="font-size:14px;font-weight:700;color:#111;">{label}</span>
                                        <span style="font-size:14px;font-weight:700;color:{color};">= {val:,} RSD</span>
                                    </div>"""
                                sc_html = f"""<div style="font-family:sans-serif;background:white;border-radius:12px;
                                    padding:20px 24px;box-shadow:0 2px 8px rgba(0,0,0,0.07);">
                                    <div style="font-size:12px;font-weight:600;color:#a855f7;text-transform:uppercase;
                                        letter-spacing:.5px;margin-bottom:14px;">
                                        Uticaj zatvaranja objekata u malim okruzima ({period_label})
                                    </div>
                                    <p style="font-size:13px;color:#666;margin-bottom:14px;">
                                        Zatvaranjem {n_mali_prof} profitabilnih objekata u {len(mali_okruzi)} malih okruga
                                        štedimo trošak, ali gubimo deo zarade. Poređenje dva scenarija:
                                    </p>
                                    <div style="font-size:12px;font-weight:600;color:#7c3aed;margin:10px 0 6px 0;">
                                        Scenario A: Zatvaramo samo {n_pravi_neg2} neprofitabilnih + OOS eliminacija
                                    </div>
                                    {_sc_red(f"Neto profit pozitivnih objekata ({n_mes}m)", neto_pozitivnih, "#10b981", False)}
                                    {_sc_red(f"Ušteda: zatvaranje {n_pravi_neg2} neprofitabilnih ({n_mes}m)", usteda_trosak2 + usteda_gubitak2, "#10b981", False)}
                                    {_sc_red(f"Povraćaj OOS izgubljene zarade ({n_mes}m)", oos_ukupno2, "#10b981", False)}
                                    {_sc_total(f"POTENCIJAL SCENARIO A", scA_potencijal)}
                                    <div style="font-size:12px;font-weight:600;color:#ec4899;margin:16px 0 6px 0;">
                                        Scenario B: Scenario A + zatvaramo i {n_mali_prof} obj. iz malih okruga
                                    </div>
                                    {_sc_red(f"Potencijal Scenario A", scA_potencijal, "#10b981", False)}
                                    {_sc_red(f"Ušteda troška: {n_mali_prof} obj. × {engine.trosak_po_objektu:,.0f} RSD × {n_mes} mes", usteda_mali_trosak, "#10b981", False)}
                                    {_sc_red(f"Izgubljen profit zatvorenih {n_mali_prof} obj. ({n_mes}m)", -neto_mali_prof, "#ec4899", False)}
                                    {_sc_total(f"POTENCIJAL SCENARIO B", scB_potencijal, "#10b981" if scB_potencijal >= scA_potencijal else "#a855f7")}
                                </div>"""
                                components.html(f'<!DOCTYPE html><html><body style="margin:0;padding:0;">{sc_html}</body></html>', height=420)

                st.markdown("---")
                excel_buf = create_excel(engine)
                fname_xl = f"ANALITIKA_{datetime.date.today().strftime('%Y%m%d')}.xlsx"
                st.download_button(f"📥 Preuzmi Excel — {fname_xl}", data=excel_buf, file_name=fname_xl,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
            except Exception as e:
                st.error(f"Greska: {str(e)}")
                import traceback; st.code(traceback.format_exc())
    else:
        components.html("""<!DOCTYPE html><html><head>
    <link href="https://fonts.googleapis.com/css2?family=Poppins:wght@300;400;600;700&display=swap" rel="stylesheet">
    </head><body style="margin:0;padding:0;background:transparent;font-family:'Poppins',sans-serif;">
    <div style="max-width:680px;margin:32px auto 0 auto;padding:0 16px;">
      <p style="font-size:11px;color:#9ca3af;font-weight:600;letter-spacing:1.5px;text-transform:uppercase;margin-bottom:14px;">
        AMAN d.o.o. &middot; Analiticki sistem
      </p>
      <h1 style="font-size:36px;font-weight:700;color:#1a0533;line-height:1.2;margin-bottom:12px;margin-top:0;">
        Predikcija prodaje<br>
        <span style="background:linear-gradient(135deg,#a855f7,#ec4899);-webkit-background-clip:text;-webkit-text-fill-color:transparent;">
          &amp; Porudzbine
        </span>
      </h1>
      <p style="font-size:15px;color:#6b7280;margin-bottom:28px;line-height:1.6;">
        Profitabilnost objekata &middot; OOS analiza &middot; Trendovi komitenata &middot; Analiza akcije
      </p>
      <div style="display:flex;gap:8px;flex-wrap:wrap;margin-bottom:36px;">
        <span style="font-size:12px;background:rgba(168,85,247,0.10);color:#7c3aed;border-radius:99px;padding:5px 14px;font-weight:600;">Predikcija</span>
        <span style="font-size:12px;background:rgba(236,72,153,0.09);color:#be185d;border-radius:99px;padding:5px 14px;font-weight:600;">Profitabilnost</span>
        <span style="font-size:12px;background:rgba(239,68,68,0.09);color:#b91c1c;border-radius:99px;padding:5px 14px;font-weight:600;">OOS analiza</span>
        <span style="font-size:12px;background:rgba(16,185,129,0.09);color:#065f46;border-radius:99px;padding:5px 14px;font-weight:600;">Trendovi</span>
      </div>
      <div style="height:1px;background:linear-gradient(90deg,rgba(168,85,247,0.3),rgba(236,72,153,0.2),transparent);margin-bottom:28px;"></div>
      <p style="font-size:14px;color:#9ca3af;text-align:center;margin-top:8px;">
        &#8593; Učitaj Excel fajl iznad da pocnes analizu
      </p>
    </div>
    </body></html>""", height=340)
